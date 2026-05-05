"""
Script để parse data từ file Excel data_qlcl.xlsx vào bảng input_qa
Hỗ trợ 3 sheets: QAPL, QANL, QAQT
"""

import os
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import logging

from dotenv import load_dotenv
import psycopg2
import psycopg2.extras
from openpyxl import load_workbook

# Load environment variables
load_dotenv()

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Database connection
DATABASE_URL = os.getenv("DATABASE_URL")
if not DATABASE_URL:
    raise ValueError("DATABASE_URL environment variable is required")

# Excel file path
EXCEL_FILE_PATH = Path(__file__).parent / "templates" / "data_qlcl.xlsx"

# Task IDs cho từng chuc_vu
QAPL_TASK_IDS = [1, 2, 3, 4, 5, 6, 7]  # 7 tasks
QANL_TASK_IDS = [8, 9, 10, 11, 12, 13, 14, 15, 16, 17]  # 10 tasks
QAQT_TASK_IDS = [18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33]  # 16 tasks

# Mapping chuc_vu -> task_ids
CHUC_VU_TASK_IDS = {
    'QAPL': QAPL_TASK_IDS,
    'QANL': QANL_TASK_IDS,
    'QAQT': QAQT_TASK_IDS
}


def get_db_connection():
    """Get database connection"""
    return psycopg2.connect(DATABASE_URL)


def get_task_name_mapping(chuc_vu: str = 'QANL') -> Dict[int, str]:
    """
    Lấy mapping giữa task id và task_name từ bảng tasks_qa
    Returns: Dict[task_id, task_name]
    """
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id, task_name FROM public.tasks_qa WHERE chuc_vu=%s ORDER BY id",
                (chuc_vu,)
            )
            rows = cur.fetchall()
            return {row[0]: row[1] for row in rows}


def find_task_columns(ws, chuc_vu: str) -> List[Tuple[int, int]]:
    """
    Tìm vị trí các task columns từ row 1
    
    Args:
        ws: Worksheet object
        chuc_vu: Chức vụ để xác định task IDs cần tìm
    
    Returns:
        List of (col_idx, task_id) tuples
    """
    task_ids = CHUC_VU_TASK_IDS.get(chuc_vu, [])
    if not task_ids:
        logger.warning(f"No task IDs defined for chuc_vu: {chuc_vu}")
        return []
    
    task_cols = []
    for col_idx in range(5, ws.max_column + 1, 2):  # Bắt đầu từ cột 5, bước 2
        task_id_val = ws.cell(1, col_idx).value
        if task_id_val is not None:
            try:
                task_id = int(task_id_val)
                if task_id in task_ids:
                    task_cols.append((col_idx, task_id))
            except (ValueError, TypeError):
                continue
    return task_cols


def parse_excel_row(ws, row_num: int, task_cols: List[Tuple[int, int]], task_id_to_name: Dict[int, str]) -> List[Dict]:
    """
    Parse một row từ Excel và trả về danh sách các records để insert
    
    Args:
        ws: Worksheet object
        row_num: Số dòng cần parse (bắt đầu từ 1)
        task_cols: List of (col_idx, task_id) tuples đã tìm được từ row 1
        task_id_to_name: Mapping từ task id sang task_name
    
    Returns:
        List of dicts với keys: ma_nv, chuc_vu, from_date, to_date, task_name, thuc_hien
    """
    records = []
    
    # Đọc các cột cơ bản
    from_date_val = ws.cell(row_num, 1).value  # Column A
    end_date_val = ws.cell(row_num, 2).value   # Column B
    ma_nv_val = ws.cell(row_num, 3).value       # Column C
    chuc_vu_val = ws.cell(row_num, 4).value     # Column D
    
    # Validate dữ liệu cơ bản
    if not all([from_date_val, end_date_val, ma_nv_val, chuc_vu_val]):
        return records
    
    # Chỉ xử lý QAPL, QANL, QAQT
    chuc_vu = str(chuc_vu_val).strip()
    if chuc_vu not in ['QAPL', 'QANL', 'QAQT']:
        return records
    
    # Convert dates
    if isinstance(from_date_val, datetime):
        from_date = from_date_val.date()
    elif isinstance(from_date_val, str):
        try:
            from_date = datetime.strptime(from_date_val.strip(), '%Y-%m-%d').date()
        except ValueError:
            try:
                from_date = datetime.strptime(from_date_val.strip(), '%d/%m/%Y').date()
            except ValueError:
                logger.warning(f"Row {row_num}: Invalid from_date format: {from_date_val}")
                return records
    else:
        logger.warning(f"Row {row_num}: Invalid from_date format: {from_date_val}")
        return records
    
    if isinstance(end_date_val, datetime):
        to_date = end_date_val.date()
    elif isinstance(end_date_val, str):
        try:
            to_date = datetime.strptime(end_date_val.strip(), '%Y-%m-%d').date()
        except ValueError:
            try:
                to_date = datetime.strptime(end_date_val.strip(), '%d/%m/%Y').date()
            except ValueError:
                logger.warning(f"Row {row_num}: Invalid end_date format: {end_date_val}")
                return records
    else:
        logger.warning(f"Row {row_num}: Invalid end_date format: {end_date_val}")
        return records
    
    ma_nv = str(ma_nv_val).strip()
    
    # Parse từng task
    # Cột col_idx (lẻ) = thuc_hien, cột col_idx+1 (chẵn) = sai_sot (không dùng cho input_qa)
    for col_idx, task_id in task_cols:
        if task_id not in task_id_to_name:
            logger.warning(f"Row {row_num}: Task id {task_id} not found in mapping")
            continue
        
        task_name = task_id_to_name[task_id]
        
        # Cột thuc_hien là cột col_idx (lẻ)
        thuc_hien_val = ws.cell(row_num, col_idx).value
        
        # Xử lý giá trị thuc_hien
        if thuc_hien_val is None:
            thuc_hien = 0
        else:
            try:
                thuc_hien = int(float(thuc_hien_val))
                if thuc_hien < 0:
                    thuc_hien = 0
            except (ValueError, TypeError):
                logger.warning(f"Row {row_num}, Task {task_id}: Invalid thuc_hien value: {thuc_hien_val}")
                thuc_hien = 0
        
        # Thêm record (cả khi thuc_hien = 0 để có đầy đủ dữ liệu)
        records.append({
            'ma_nv': ma_nv,
            'chuc_vu': chuc_vu,
            'from_date': from_date,
            'to_date': to_date,
            'task_name': task_name,
            'thuc_hien': thuc_hien
        })
    
    return records


def parse_sheet(ws, sheet_name: str) -> List[Dict]:
    """
    Parse một sheet từ workbook
    
    Args:
        ws: Worksheet object
        sheet_name: Tên sheet cần parse (QAPL, QANL, hoặc QAQT)
    
    Returns:
        List of dicts với keys: ma_nv, chuc_vu, from_date, to_date, task_name, thuc_hien
    """
    chuc_vu = sheet_name  # QAPL, QANL, hoặc QAQT
    logger.info(f"Sheet {sheet_name}: {ws.max_row} rows, {ws.max_column} columns")
    
    # Lấy mapping task id -> task_name
    task_id_to_name = get_task_name_mapping(chuc_vu)
    logger.info(f"Found {len(task_id_to_name)} tasks for {chuc_vu}")
    
    # Tìm task columns từ row 1 (một lần duy nhất)
    task_cols = find_task_columns(ws, chuc_vu)
    logger.info(f"Found {len(task_cols)} task columns for {chuc_vu}: {task_cols}")
    
    if not task_cols:
        logger.warning(f"No task columns found for {chuc_vu}")
        return []
    
    # Parse từ row 2 trở đi (row 1 là header)
    all_records = []
    for row_num in range(2, ws.max_row + 1):
        records = parse_excel_row(ws, row_num, task_cols, task_id_to_name)
        all_records.extend(records)
    
    logger.info(f"Parsed {len(all_records)} records from {sheet_name} sheet ({ws.max_row - 1} data rows)")
    return all_records


def parse_all_sheets(excel_path: Path) -> List[Dict]:
    """
    Parse tất cả các sheets (QAPL, QANL, QAQT) từ file Excel
    
    Returns:
        List of dicts với keys: ma_nv, chuc_vu, from_date, to_date, task_name, thuc_hien
    """
    logger.info(f"Loading Excel file: {excel_path}")
    wb = load_workbook(excel_path, data_only=True)
    
    all_records = []
    sheets = ['QAPL', 'QANL', 'QAQT']
    
    for sheet_name in sheets:
        if sheet_name not in wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in Excel file. Available sheets: {wb.sheetnames}")
            continue
        
        logger.info(f"\n{'='*60}")
        logger.info(f"Processing sheet: {sheet_name}")
        logger.info(f"{'='*60}")
        ws = wb[sheet_name]
        records = parse_sheet(ws, sheet_name)
        all_records.extend(records)
        logger.info(f"Total records so far: {len(all_records)}")
    
    logger.info(f"\n{'='*60}")
    logger.info(f"Total records from all sheets: {len(all_records)}")
    logger.info(f"{'='*60}")
    return all_records


def insert_records(records: List[Dict], dry_run: bool = False):
    """
    Insert records vào bảng input_qa
    
    Args:
        records: List of dicts với keys: ma_nv, chuc_vu, from_date, to_date, task_name, thuc_hien
        dry_run: Nếu True, chỉ log mà không insert
    """
    if not records:
        logger.info("No records to insert")
        return
    
    insert_sql = """
        INSERT INTO public.input_qa (ma_nv, chuc_vu, from_date, to_date, task_name, thuc_hien)
        VALUES (%(ma_nv)s, %(chuc_vu)s, %(from_date)s, %(to_date)s, %(task_name)s, %(thuc_hien)s)
        RETURNING id
    """
    
    if dry_run:
        logger.info(f"DRY RUN: Would insert {len(records)} records")
        for i, rec in enumerate(records[:5], 1):  # Show first 5
            logger.info(f"  Record {i}: {rec}")
        if len(records) > 5:
            logger.info(f"  ... and {len(records) - 5} more records")
        return
    
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            inserted_count = 0
            error_count = 0
            
            for rec in records:
                try:
                    cur.execute(insert_sql, rec)
                    inserted_count += 1
                except Exception as e:
                    error_count += 1
                    logger.error(f"Error inserting record {rec}: {e}")
            
            conn.commit()
            logger.info(f"Inserted {inserted_count} records successfully")
            if error_count > 0:
                logger.warning(f"Failed to insert {error_count} records")


def main(dry_run: bool = False):
    """
    Main function
    
    Args:
        dry_run: Nếu True, chỉ parse và log, không insert vào database
    """
    if not EXCEL_FILE_PATH.exists():
        logger.error(f"Excel file not found: {EXCEL_FILE_PATH}")
        return
    
    try:
        # Parse Excel - tất cả sheets
        records = parse_all_sheets(EXCEL_FILE_PATH)
        
        # Insert vào database
        insert_records(records, dry_run=dry_run)
        
        logger.info("\n✅ Done!")
        
    except Exception as e:
        logger.error(f"Error: {e}", exc_info=True)


if __name__ == "__main__":
    import sys
    
    # Kiểm tra argument để chạy dry-run
    dry_run = "--dry-run" in sys.argv or "-d" in sys.argv
    
    if dry_run:
        logger.info("Running in DRY RUN mode (no database changes)")
    
    main(dry_run=dry_run)

