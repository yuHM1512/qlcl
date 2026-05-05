import argparse
import os
import unicodedata
from typing import Dict, Tuple

import psycopg2
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()


def normalize_text(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    if not s:
        return ""
    s = ''.join(ch for ch in unicodedata.normalize('NFKD', s) if not unicodedata.combining(ch))
    return s.lower()


def get_col_indices(header_row) -> Dict[str, int]:
    indices = {}
    for idx, cell in enumerate(header_row, start=1):
        name = normalize_text(cell.value or "")
        if not name:
            continue
        if name in ("loai hang", "loaihang", "loai_hang"):
            indices["loai_hang"] = idx
        elif name in ("bo phan", "bophan", "bo_phan"):
            indices["bo_phan"] = idx
        elif name in ("chi tiet", "chitiet", "chi_tiet"):
            indices["chi_tiet"] = idx
    return indices


def main():
    parser = argparse.ArgumentParser(description="Import dm_loai_hang, dm_bo_phan, dm_chi_tiet from Excel")
    parser.add_argument("--file", default=r"D:\Data Analyst\Tools\kpi\qlcl\dm_import.xlsx")
    args = parser.parse_args()

    database_url = os.getenv("DATABASE_URL")
    if not database_url:
        raise ValueError("DATABASE_URL environment variable is required")

    wb = load_workbook(args.file)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    cols = get_col_indices(header_row)
    missing = [k for k in ("loai_hang", "bo_phan", "chi_tiet") if k not in cols]
    if missing:
        raise ValueError(f"Missing columns in header: {', '.join(missing)}")

    loai_cache: Dict[str, int] = {}
    bo_phan_cache: Dict[Tuple[int, str], int] = {}

    inserted_loai = 0
    inserted_bp = 0
    inserted_ct = 0

    with psycopg2.connect(database_url) as conn:
        with conn.cursor() as cur:
            for row in ws.iter_rows(min_row=2):
                loai_val = row[cols["loai_hang"] - 1].value
                bp_val = row[cols["bo_phan"] - 1].value
                ct_val = row[cols["chi_tiet"] - 1].value

                loai = str(loai_val).strip() if loai_val is not None else ""
                bp = str(bp_val).strip() if bp_val is not None else ""
                ct = str(ct_val).strip() if ct_val is not None else ""

                if not loai or not bp:
                    continue

                if loai not in loai_cache:
                    cur.execute(
                        "INSERT INTO public.dm_loai_hang (ten_loai) VALUES (%s) ON CONFLICT DO NOTHING",
                        (loai,)
                    )
                    if cur.rowcount:
                        inserted_loai += 1
                    cur.execute("SELECT id FROM public.dm_loai_hang WHERE ten_loai = %s", (loai,))
                    loai_cache[loai] = cur.fetchone()[0]

                loai_id = loai_cache[loai]
                bp_key = (loai_id, bp)
                if bp_key not in bo_phan_cache:
                    cur.execute(
                        """
                        INSERT INTO public.dm_bo_phan (loai_hang_id, ten_bo_phan)
                        VALUES (%s, %s)
                        ON CONFLICT DO NOTHING
                        """,
                        (loai_id, bp)
                    )
                    if cur.rowcount:
                        inserted_bp += 1
                    cur.execute(
                        "SELECT id FROM public.dm_bo_phan WHERE loai_hang_id = %s AND ten_bo_phan = %s",
                        (loai_id, bp)
                    )
                    bo_phan_cache[bp_key] = cur.fetchone()[0]

                if ct:
                    bp_id = bo_phan_cache[bp_key]
                    cur.execute(
                        """
                        INSERT INTO public.dm_chi_tiet (bo_phan_id, ten_chi_tiet)
                        VALUES (%s, %s)
                        ON CONFLICT DO NOTHING
                        """,
                        (bp_id, ct)
                    )
                    if cur.rowcount:
                        inserted_ct += 1

        conn.commit()

    print("Import completed")
    print(f"Inserted loai_hang: {inserted_loai}")
    print(f"Inserted bo_phan: {inserted_bp}")
    print(f"Inserted chi_tiet: {inserted_ct}")


if __name__ == "__main__":
    main()
