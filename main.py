from datetime import date, datetime
from typing import Any, List, Optional, Dict, Tuple
from datetime import timedelta
from pydantic import BaseModel
import os
import re
import shutil
import logging
import json
import threading
import time
from pathlib import Path as PathLib
import unicodedata
from urllib.parse import quote, unquote

# Load environment variables from .env file
from dotenv import load_dotenv
load_dotenv()

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

from fastapi import FastAPI, Request, Form, Depends, HTTPException, Query, Path, File, UploadFile
from fastapi.responses import RedirectResponse, JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
# from starlette.middleware.sessions import SessionMiddleware

import psycopg2
import psycopg2.extras
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
try:
    import win32com.client  # type: ignore
    WIN32_AVAILABLE = True
except ImportError:
    WIN32_AVAILABLE = False


# Lấy cấu hình từ environment variables hoặc file .env
# Tất cả các biến này là bắt buộc, phải được cấu hình trong file .env hoặc environment variables
DATABASE_URL = os.getenv("DATABASE_URL")
if not DATABASE_URL:
    raise ValueError("DATABASE_URL environment variable is required. Please set it in .env file or environment variables.")

PROD_FACTORY_DATABASE_URL = os.getenv("PROD_FACTORY_DATABASE_URL")
if not PROD_FACTORY_DATABASE_URL:
    raise ValueError("PROD_FACTORY_DATABASE_URL environment variable is required. Please set it in .env file or environment variables.")

TEMPLATES_DIR = os.getenv("TEMPLATES_DIR")
if not TEMPLATES_DIR:
    raise ValueError("TEMPLATES_DIR environment variable is required. Please set it in .env file or environment variables.")

EXCEL_TEMPLATE_PATH = os.getenv("EXCEL_TEMPLATE_PATH")
if not EXCEL_TEMPLATE_PATH:
    raise ValueError("EXCEL_TEMPLATE_PATH environment variable is required. Please set it in .env file or environment variables.")

PDF_STORAGE_DIR = os.getenv("PDF_STORAGE_DIR")
if not PDF_STORAGE_DIR:
    raise ValueError("PDF_STORAGE_DIR environment variable is required. Please set it in .env file or environment variables.")

IMAGES_STORAGE_DIR = os.getenv("IMAGES_STORAGE_DIR")
if not IMAGES_STORAGE_DIR:
    raise ValueError("IMAGES_STORAGE_DIR environment variable is required. Please set it in .env file or environment variables.")


def env_flag(name: str, default: bool = False) -> bool:
    raw = os.getenv(name)
    if raw is None:
        return default
    return str(raw).strip().lower() in {"1", "true", "yes", "on"}


QTCN_AUTO_SYNC_ENABLED = env_flag("QTCN_AUTO_SYNC_ENABLED", default=False)
try:
    QTCN_AUTO_SYNC_INTERVAL_MINUTES = max(1, int(os.getenv("QTCN_AUTO_SYNC_INTERVAL_MINUTES", "60")))
except ValueError:
    QTCN_AUTO_SYNC_INTERVAL_MINUTES = 60

DB_SCRIPTS_DIR = PathLib(__file__).resolve().parent / "db"
GEMBA_CP_BASE_DIR = PathLib(__file__).resolve().parent / "gemba_cp"
GEMBA_CP_STATIC_DIR = GEMBA_CP_BASE_DIR / "static"
GEMBA_CP_TEMPLATES_DIR = GEMBA_CP_BASE_DIR / "templates"

from gemba_cp.config import get_settings as get_gemba_settings
from gemba_cp.db import Base as GembaCPBase, engine as gemba_cp_engine
from gemba_cp.routers import admin as gemba_admin_router, dashboard as gemba_dashboard_router

gemba_cp_settings = get_gemba_settings()

SCHEMA_BOOTSTRAP_FILES = [
    "initialize_qlcl.sql",
    "migrate_ds_qa_to_quality_employees.sql",
    "alter_quality_employees_add_station.sql",
    "create_customer_hierarchy.sql",
    "create_error_classification.sql",
    "create_hdkp_tables.sql",
    "create_prod_plan.sql",
    "migrate_prod_plan_bo_phan_json.sql",
    "create_qc_output_sp_log.sql",
    "migrate_qc_output_sp_log_status.sql",
    "create_qc_error_hierarchy.sql",
    "create_qc_cum.sql",
    "migrate_dm_loai_hang_type.sql",
    "create_qc_error_log_sp.sql",
    "alter_qc_output_sp_log_add_station.sql",
    "alter_qc_sp_add_ma_nv.sql",
    "create_qc_defect_multi.sql",
    "alter_qc_defect_multi_add_station.sql",
    "alter_qc_defect_add_image_path.sql",
    "create_qc_error_dps.sql",
    "alter_qc_error_dps_add_keys.sql",
    "alter_qc_error_dps_add_station.sql",
    "alter_qc_error_dps_add_bo_phan.sql",
    "create_qc_hdkp_endline.sql",
    "alter_prod_plan_add_po_info.sql",
    "alter_prod_plan_add_sync_fields.sql",
]

# Tạo thư mục lưu PDF và images nếu chưa có
os.makedirs(PDF_STORAGE_DIR, exist_ok=True)
os.makedirs(IMAGES_STORAGE_DIR, exist_ok=True)
os.makedirs(os.path.join(IMAGES_STORAGE_DIR, "qc_sp"), exist_ok=True)


def get_db_connection():
    return psycopg2.connect(DATABASE_URL)


def get_prod_factory_connection():
    return psycopg2.connect(PROD_FACTORY_DATABASE_URL)


def bootstrap_qlcl_schema():
    """Run bundled idempotent SQL scripts so a fresh environment can self-bootstrap."""
    with get_db_connection() as conn:
        conn.autocommit = True
        with conn.cursor() as cur:
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS public.app_schema_bootstrap (
                    script_name TEXT PRIMARY KEY,
                    executed_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
                )
                """
            )
            for script_name in SCHEMA_BOOTSTRAP_FILES:
                script_path = DB_SCRIPTS_DIR / script_name
                if not script_path.exists():
                    raise FileNotFoundError(f"Missing schema bootstrap script: {script_path}")
                cur.execute(
                    "SELECT 1 FROM public.app_schema_bootstrap WHERE script_name = %s",
                    (script_name,),
                )
                if cur.fetchone():
                    continue
                logger.info("Bootstrapping schema with %s", script_name)
                sql_text = script_path.read_text(encoding="utf-8-sig")
                cur.execute(sql_text)
                cur.execute(
                    """
                    INSERT INTO public.app_schema_bootstrap (script_name)
                    VALUES (%s)
                    ON CONFLICT (script_name) DO NOTHING
                    """,
                    (script_name,),
                )


def generate_ma_nv_variants(ma_nv: str) -> List[str]:
    """Generate uppercase variants for mã nhân viên, handling Vietnamese characters like Đ."""
    if not ma_nv:
        return []
    base = ma_nv.strip().upper()
    variants = []
    if base:
        variants.append(base)
    simplified = ''.join(
        ch for ch in unicodedata.normalize('NFKD', base)
        if not unicodedata.combining(ch)
    )
    simplified = simplified.replace('Đ', 'D').replace('đ', 'D').upper()
    if simplified and simplified not in variants:
        variants.append(simplified)
    return variants if variants else [base]


def encode_ma_nv_cookie(ma_nv: str) -> str:
    """Percent-encode mã nhân viên để lưu trong cookie (tránh lỗi latin-1)."""
    if not ma_nv:
        return ""
    try:
        return quote(ma_nv.strip())
    except Exception:
        return quote(str(ma_nv))


def decode_ma_nv_cookie(value: Optional[str]) -> str:
    """Giải mã giá trị mã nhân viên từ cookie."""
    if not value:
        return ""
    try:
        return unquote(value)
    except Exception:
        return value


def get_authenticated_user(request: Request, qc_only: bool = False) -> Optional[Dict]:
    """Lấy thông tin nhân viên từ cookie; trả về None nếu chưa đăng nhập hoặc cookie không hợp lệ."""
    ma_nv_cookie = request.cookies.get("ma_nv")
    if not ma_nv_cookie:
        return None

    ma_nv = decode_ma_nv_cookie(ma_nv_cookie)
    ma_nv_variants = generate_ma_nv_variants(ma_nv)
    if not ma_nv_variants:
        return None

    query = """
        SELECT ma_nv, ho_ten as name, chuc_vu as department, don_vi, bo_phan, station
        FROM public.quality_employees
        WHERE ma_nv = ANY(%s)
    """
    params: List[Any] = [ma_nv_variants]
    if qc_only:
        query += " AND chuc_vu = 'QC'"
    query += """
        ORDER BY CASE WHEN ma_nv = %s THEN 0 ELSE 1 END
        LIMIT 1
    """
    params.append(ma_nv)

    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(query, tuple(params))
            user = cur.fetchone()

    return dict(user) if user else None


def require_authenticated_api_user(request: Request) -> Dict:
    """Dependency bảo vệ các API chỉ cho người đã đăng nhập."""
    user = get_authenticated_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Chưa đăng nhập")
    return user


def require_qaqt_api_user(request: Request) -> Dict:
    """Dependency chỉ cho QAQT (admin QA) — dùng cho thao tác tinh chỉnh master data."""
    user = get_authenticated_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Chưa đăng nhập")
    if (user.get("department") or "").upper() != "QAQT":
        raise HTTPException(status_code=403, detail="Chỉ tài khoản QAQT mới có quyền thao tác này.")
    return user


QC_DON_VI_SCOPED_ROLES = {"FACTORY_MANAGER", "FACTORY_DIRECTOR"}


def is_qc_don_vi_scoped_role(user: Optional[Dict]) -> bool:
    role = ((user or {}).get("department") or "").upper()
    return role in QC_DON_VI_SCOPED_ROLES


def resolve_qc_don_vi_scope(request: Request, requested_don_vi: Optional[str]) -> Tuple[Optional[str], Dict]:
    """Return the effective QC don_vi filter for view-only factory roles."""
    user = require_authenticated_api_user(request)
    if is_qc_don_vi_scoped_role(user):
        scoped_don_vi = (user.get("don_vi") or "").strip()
        if not scoped_don_vi:
            raise HTTPException(status_code=403, detail="Tài khoản chưa được gán đơn vị quản lý.")
        return scoped_don_vi, user
    return requested_don_vi, user


def resolve_dashboard_default_station(
    cur,
    *,
    station: Optional[str] = None,
    scoped_user: Optional[Dict] = None,
    don_vi: Optional[str] = None,
    bo_phan: Optional[str] = None,
    ma_hang: Optional[str] = None,
    type_name: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    date_before: Optional[str] = None,
    resolved_bo_phan_expr: str = "COALESCE(NULLIF(qe.bo_phan, ''), '')",
) -> str:
    station_clean = (station or "").strip()
    if station_clean:
        return station_clean

    if not is_qc_don_vi_scoped_role(scoped_user):
        return "Trạm sau seam"

    where_clauses = ["o.station IS NOT NULL", "o.station <> ''"]
    params: List[Any] = []
    if date_from and date_to:
        where_clauses.append("o.date BETWEEN %s AND %s")
        params.extend([date_from, date_to])
    elif date_before:
        where_clauses.append("o.date < %s")
        params.append(date_before)
    if don_vi:
        where_clauses.append("p.don_vi = %s")
        params.append(don_vi)
    if bo_phan:
        where_clauses.append(f"{resolved_bo_phan_expr} = %s")
        params.append(bo_phan)
    if ma_hang:
        where_clauses.append("COALESCE(p.ma_hang, '') = %s")
        params.append(ma_hang)
    if type_name:
        where_clauses.append("COALESCE(tgt.type, '') = %s")
        params.append(type_name)

    cur.execute(
        f"""
        SELECT o.station
        FROM public.qc_output_sp_log o
        JOIN public.prod_plan p ON p.id = o.plan_id
        LEFT JOIN public.quality_employees qe ON qe.ma_nv = o.ma_nv
        LEFT JOIN public.dm_loai_hang lh ON lh.ten_loai = p.loai_hang
        LEFT JOIN public.dm_loai_hang_target tgt ON tgt.id_type = lh.id_type
        WHERE {" AND ".join(where_clauses)}
        GROUP BY o.station
        ORDER BY SUM(CASE WHEN o.delta > 0 THEN o.delta ELSE 0 END) DESC, o.station
        LIMIT 1
        """,
        tuple(params),
    )
    row = cur.fetchone() or {}
    return (row.get("station") or "").strip()


def build_qc_template_context(request: Request, user: Optional[Dict], **extra) -> Dict[str, Any]:
    role = ((user or {}).get("department") or "").upper()
    context: Dict[str, Any] = {
        "request": request,
        "user": user,
        "qc_role": role,
        "is_qc_role": role == "QC",
        "is_qaqt_role": role == "QAQT",
        "is_qc_viewer_role": role in QC_DON_VI_SCOPED_ROLES,
        "qc_scope_don_vi": ((user or {}).get("don_vi") or "") if role in QC_DON_VI_SCOPED_ROLES else "",
    }
    context.update(extra)
    return context


def build_static_asset_version(asset_path: PathLib) -> str:
    """Use file mtime as a simple cache-busting version for static assets."""
    try:
        return str(int(asset_path.stat().st_mtime))
    except FileNotFoundError:
        return "0"


def _is_merged_cell(ws, cell):
    """
    Kiểm tra xem cell có phải là merged cell không.
    Trả về True nếu là merged cell, False nếu không.
    """
    try:
        # Kiểm tra xem cell có nằm trong merged range không
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # Kiểm tra xem có phải là top-left cell không
                min_col, min_row, max_col, max_row = merged_range.bounds
                if cell.row == min_row and cell.column == min_col:
                    return False  # Đây là top-left cell, có thể ghi được
                else:
                    return True  # Đây là merged cell nhưng không phải top-left
        return False
    except Exception:
        return False


def fill_excel_template(template_path: str, data: Dict, output_path: str):
    """
    Điền biến {{variable}} hoặc {{variable_N}} vào Excel template.
    
    Template có thể có:
    - Biến có số: {{a_root_cause_1}}, {{a_root_cause_2}}, ...
    - Biến không có số: {{a_root_cause}} ở nhiều dòng (mỗi dòng sẽ điền theo số thứ tự)
    
    Logic:
    1. Xử lý các section (A, B, III) - tìm các dòng có biến không có số và điền theo thứ tự
    2. Xử lý các biến có số - điền trực tiếp
    3. Xử lý các biến khác (Section I)
    4. Nếu có biến nhưng không có dữ liệu, để trống (không điền "trống")
    5. Tự động bật wrap text cho các cell được điền dữ liệu
    """
    try:
        wb = load_workbook(template_path)
        ws = wb.active
        
        # Xử lý Section II A: tìm các dòng có {{a_root_cause}} và điền theo thứ tự
        _fill_section_with_sequential_rows(ws, 'a_root_cause', data, 
                                           ['a_root_cause', 'a_hđkp_tuc_thoi', 'a_hđ_phong_ngua', 
                                            'a_tg_theo_doi', 'a_trach_nhiem', 'a_tg_thuc_hien'])
        
        # Xử lý Section II B
        _fill_section_with_sequential_rows(ws, 'b_root_cause', data,
                                           ['b_root_cause', 'b_hđkp_tuc_thoi', 'b_hđ_phong_ngua',
                                            'b_tg_theo_doi', 'b_trach_nhiem', 'b_tg_thuc_hien'])
        
        # Xử lý Section III
        _fill_section_with_sequential_rows(ws, 'cong_viec', data,
                                           ['stt', 'cong_viec', 'trach_nhiem', 'ngay_bat_dau',
                                            'ngay_hoan_thanh', 'giam_sat', 'ket_qua_dat', 'ket_qua_khong_dat', 'ket_luan'])
        
        # Xử lý hình ảnh (vd_image) - chèn hình ảnh vào cell
        _insert_image_to_cell(ws, data.get('vd_image', ''))
        
        # Xử lý các biến khác (Section I và các biến có số hoặc đơn lẻ)
        for row in ws.iter_rows():
            for cell in row:
                # B? qua merged cells (tr? top-left)
                if _is_merged_cell(ws, cell):
                    continue
                
                if cell.value and isinstance(cell.value, str):
                    # Tìm tất cả các biến {{variable}} hoặc {{variable_N}}
                    pattern = r'\{\{(\w+(?:_\d+)?)\}\}'
                    matches = re.findall(pattern, cell.value)
                    
                    if not matches:
                        continue
                    
                    original_value = cell.value
                    has_replacement = False
                    
                    for var_name in matches:
                        # Bỏ qua vd_image (đã xử lý riêng bằng _insert_image_to_cell)
                        if var_name == 'vd_image':
                            continue
                        
                        # Bỏ qua các biến đã xử lý trong sections (nếu không có số)
                        if var_name in ['a_root_cause', 'b_root_cause', 'cong_viec'] or \
                           var_name.startswith('a_root_cause') or var_name.startswith('b_root_cause') or \
                           var_name.startswith('cong_viec') or var_name.startswith('stt') or \
                           var_name.startswith('ket_qua_dat') or var_name.startswith('ket_qua_khong_dat'):
                            # Chỉ xử lý nếu có số (đã xử lý ở trên)
                            if '_' in var_name and var_name.split('_')[-1].isdigit():
                                continue
                            # Nếu không có số, đã xử lý ở _fill_section_with_sequential_rows
                            continue
                        
                        # Bỏ qua ket_qua: đã được xử lý trong _fill_section_with_sequential_rows cho Section III
                        # Chỉ xử lý ở đây nếu không phải là Section III (không có {{cong_viec}} trong dòng)
                        if var_name == 'ket_qua' or (var_name.startswith('ket_qua') and var_name != 'ket_qua_dat' and var_name != 'ket_qua_khong_dat'):
                            # Kiểm tra xem dòng này có thuộc Section III không (có {{cong_viec}})
                            is_section_iii = False
                            for check_cell in ws[cell.row]:
                                if check_cell.value and isinstance(check_cell.value, str):
                                    if '{{cong_viec}}' in check_cell.value or '{{cong_viec_' in str(check_cell.value):
                                        is_section_iii = True
                                        break
                            
                            # Nếu thuộc Section III, bỏ qua (đã xử lý trong _fill_section_with_sequential_rows)
                            if is_section_iii:
                                continue
                            
                            # Nếu không thuộc Section III, xử lý như bình thường (trường hợp đặc biệt)
                            # Tìm số thứ tự (nếu có)
                            if '_' in var_name and var_name.split('_')[-1].isdigit():
                                idx = int(var_name.split('_')[-1])
                            else:
                                idx = 1  # Mặc định là 1
                            
                            # Chỉ xử lý nếu có dữ liệu tương ứng
                            ket_qua_dat = data.get(f'ket_qua_dat_{idx}', '')
                            ket_qua_khong_dat = data.get(f'ket_qua_khong_dat_{idx}', '')
                            
                            # Chỉ xử lý nếu có dữ liệu
                            if not ket_qua_dat and not ket_qua_khong_dat:
                                continue
                            
                            # Tìm tất cả các cell trong cùng dòng có chứa biến {{ket_qua}} hoặc {{ket_qua_N}}
                            pattern_str = '{{' + var_name + '}}'
                            ket_qua_cells = []
                            
                            # Quét toàn bộ dòng để tìm tất cả các cell có chứa biến này
                            for col in range(1, ws.max_column + 1):
                                try:
                                    row_cell = ws.cell(row=cell.row, column=col)
                                    if row_cell.value and isinstance(row_cell.value, str):
                                        if pattern_str in str(row_cell.value):
                                            ket_qua_cells.append(col)
                                except Exception:
                                    pass
                            
                            # Nếu tìm thấy 2 cell, điền vào: cell đầu tiên = Đạt, cell thứ hai = Không đạt
                            if len(ket_qua_cells) >= 2:
                                # Cell đầu tiên: cột trái (Đạt)
                                try:
                                    cell1 = ws.cell(row=cell.row, column=ket_qua_cells[0])
                                    if cell1.value and isinstance(cell1.value, str):
                                        cell1.value = cell1.value.replace(pattern_str, ket_qua_dat)
                                        if not cell1.value or cell1.value.strip() == '':
                                            cell1.value = None
                                except Exception:
                                    pass
                                
                                # Cell thứ hai: cột phải (Không đạt)
                                try:
                                    cell2 = ws.cell(row=cell.row, column=ket_qua_cells[1])
                                    if cell2.value and isinstance(cell2.value, str):
                                        cell2.value = cell2.value.replace(pattern_str, ket_qua_khong_dat)
                                        if not cell2.value or cell2.value.strip() == '':
                                            cell2.value = None
                                except Exception:
                                    pass
                            elif len(ket_qua_cells) == 1:
                                # Chỉ có 1 cell, điền vào cell hiện tại (cột trái - Đạt)
                                if pattern_str in original_value:
                                    original_value = original_value.replace(pattern_str, ket_qua_dat)
                                    has_replacement = True
                            
                            # Bỏ qua biến này trong vòng lặp tiếp theo
                            continue
                        
                        # Lấy giá trị từ data dict
                        value = data.get(var_name, '')
                        
                        # Nếu không có giá trị, để trống (không điền "trống")
                        if not value:
                            value = ''
                        
                        # Nếu biến không có số, thử lấy từ _1
                        if '_' not in var_name or not var_name.split('_')[-1].isdigit():
                            var_with_1 = var_name + '_1'
                            if var_with_1 in data:
                                value = data.get(var_with_1, '')
                        
                        # Thay thế biến bằng giá trị
                        pattern_str = '{{' + var_name + '}}'
                        if pattern_str in original_value:
                            original_value = original_value.replace(pattern_str, str(value))
                            has_replacement = True
                    
                    # Cập nhật cell và bật wrap text
                    if has_replacement:
                        try:
                            cell.value = original_value if original_value else None
                            if cell.value:
                                _set_wrap_text(cell)
                        except Exception as cell_error:
                            # Nếu có lỗi (merged cell, etc.), bỏ qua
                            logger.debug(f"B? qua cell {cell.coordinate}: {cell_error}")
        
        wb.save(output_path)
        logger.info(f"Đã lưu Excel file: {output_path}")
        return output_path
    except Exception as e:
        logger.error(f"Lá»—i trong fill_excel_template: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Lỗi khi điền Excel template: {str(e)}")


def _fill_section_with_sequential_rows(ws, base_var_name: str, data: Dict, var_patterns: List[str]):
    """
    Tìm các dòng có biến không có số (ví dụ {{a_root_cause}}) và điền theo thứ tự:
    - Dòng đầu tiên: điền với _1
    - Dòng thứ hai: điền với _2
    - Dòng thứ ba: điền với _3
    - Nếu không có dữ liệu, để trống
    """
    # Tìm tất cả các dòng có chứa biến không có số
    rows_with_base_var = []
    
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            if _is_merged_cell(ws, cell):
                continue
            if cell.value and isinstance(cell.value, str):
                if '{{' + base_var_name + '}}' in cell.value:
                    rows_with_base_var.append(row_idx)
                    break
    
    # Tìm số thứ tự tối đa có trong data (tính một lần ở đầu)
    max_idx = 0
    for key in data.keys():
        if key.startswith(base_var_name + '_'):
            try:
                key_idx = int(key.split('_')[-1])
                max_idx = max(max_idx, key_idx)
            except ValueError:
                pass
    
    # Điền dữ liệu vào từng dòng theo thứ tự
    for idx, row_idx in enumerate(rows_with_base_var, start=1):
        # Nếu dòng này vượt quá số lượng items có trong data, để trống
        if idx > max_idx:
            # Xóa tất cả các biến trong dòng này (để trống), bao gồm cả ket_qua
            for cell in ws[row_idx]:
                if _is_merged_cell(ws, cell):
                    continue
                if cell.value and isinstance(cell.value, str):
                    # Xóa tất cả các biến bao gồm ket_qua
                    for var_name in var_patterns:
                        pattern_str = '{{' + var_name + '}}'
                        if pattern_str in cell.value:
                            cell.value = cell.value.replace(pattern_str, '')
                    # Xóa {{ket_qua}} và {{ket_qua_N}} nếu có
                    if '{{ket_qua}}' in cell.value:
                        cell.value = cell.value.replace('{{ket_qua}}', '')
                    if f'{{ket_qua_{idx}}}' in cell.value:
                        cell.value = cell.value.replace(f'{{ket_qua_{idx}}}', '')
                    if not cell.value or cell.value.strip() == '':
                        cell.value = None
        else:
            # Xử lý đặc biệt cho ket_qua: tìm tất cả các cell có {{ket_qua}} hoặc {{ket_qua_N}} trong dòng này
            ket_qua_cells = []
            for cell_temp in ws[row_idx]:
                if _is_merged_cell(ws, cell_temp):
                    continue
                if cell_temp.value and isinstance(cell_temp.value, str):
                    # Tìm {{ket_qua}} hoặc {{ket_qua_N}}
                    if '{{ket_qua}}' in cell_temp.value or f'{{ket_qua_{idx}}}' in cell_temp.value:
                        ket_qua_cells.append(cell_temp.column)
            
            # Nếu tìm thấy 2 cell có {{ket_qua}}, điền vào 2 cell đó (chỉ khi có dữ liệu)
            if len(ket_qua_cells) >= 2:
                ket_qua_dat = data.get(f'ket_qua_dat_{idx}', '')
                ket_qua_khong_dat = data.get(f'ket_qua_khong_dat_{idx}', '')
                
                # Cell đầu tiên: cột trái (Đạt)
                try:
                    cell1 = ws.cell(row=row_idx, column=ket_qua_cells[0])
                    if cell1.value and isinstance(cell1.value, str):
                        cell1.value = cell1.value.replace('{{ket_qua}}', ket_qua_dat).replace(f'{{ket_qua_{idx}}}', ket_qua_dat)
                        if not cell1.value or cell1.value.strip() == '':
                            cell1.value = None
                except Exception:
                    pass
                
                # Cell thứ hai: cột phải (Không đạt)
                try:
                    cell2 = ws.cell(row=row_idx, column=ket_qua_cells[1])
                    if cell2.value and isinstance(cell2.value, str):
                        cell2.value = cell2.value.replace('{{ket_qua}}', ket_qua_khong_dat).replace(f'{{ket_qua_{idx}}}', ket_qua_khong_dat)
                        if not cell2.value or cell2.value.strip() == '':
                            cell2.value = None
                except Exception:
                    pass
            
            # Điền dữ liệu với số thứ tự idx cho các biến khác
            for cell in ws[row_idx]:
                if _is_merged_cell(ws, cell):
                    continue
                if cell.value and isinstance(cell.value, str):
                    original_value = cell.value
                    has_replacement = False
                    
                    # Bỏ qua nếu cell này đã được xử lý cho ket_qua
                    if cell.column in ket_qua_cells:
                        continue
                    
                    for var_name in var_patterns:
                        # Bỏ qua ket_qua_dat và ket_qua_khong_dat vì đã xử lý ở trên
                        if var_name == 'ket_qua_dat' or var_name == 'ket_qua_khong_dat':
                            continue
                        
                        # Tạo biến với số thứ tự
                        var_with_idx = var_name + '_' + str(idx)
                        value = data.get(var_with_idx, '')
                        
                        # Nếu không có giá trị, để trống
                        if not value:
                            value = ''
                        
                        # Thay thế biến (có thể là không có số hoặc có số)
                        pattern_no_num = '{{' + var_name + '}}'
                        pattern_with_num = '{{' + var_name + '_' + str(idx) + '}}'
                        
                        if pattern_no_num in original_value:
                            original_value = original_value.replace(pattern_no_num, str(value))
                            has_replacement = True
                        if pattern_with_num in original_value:
                            original_value = original_value.replace(pattern_with_num, str(value))
                            has_replacement = True
                    
                    if has_replacement:
                        try:
                            cell.value = original_value if original_value else None
                            if cell.value:
                                _set_wrap_text(cell)
                        except Exception as cell_error:
                            logger.debug(f"B? qua cell {cell.coordinate}: {cell_error}")


def _process_section_rows(ws, base_var_name: str, max_items: int, data: Dict, var_patterns: List[str]):
    """
    Xử lý một section: tìm dòng template, điền dữ liệu và duplicate nếu cần.
    """
    # Tìm dòng template (dòng đầu tiên có biến không có số)
    template_row_idx = None
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if '{{' + base_var_name + '}}' in cell.value:
                    template_row_idx = row_idx
                    break
        if template_row_idx:
            break
    
    if not template_row_idx:
        return  # Không tìm thấy template row
    
    # Điền dữ liệu vào dòng đầu tiên (item 1)
    _fill_row_data(ws, template_row_idx, base_var_name, 1, data, var_patterns)
    
    # Nếu có nhiều items, duplicate và điền vào các dòng tiếp theo
    if max_items > 1:
        for idx in range(2, max_items + 1):
            # Tìm dòng trống tiếp theo hoặc duplicate dòng template
            new_row_idx = template_row_idx + (idx - 1)
            
            # Kiểm tra xem dòng này có tồn tại không, nếu không thì insert
            if new_row_idx > ws.max_row:
                ws.insert_rows(new_row_idx)
            
            # Copy format và giá trị từ dòng template
            _copy_row_format(ws, template_row_idx, new_row_idx)
            
            # Điền dữ liệu với số thứ tự idx
            _fill_row_data(ws, new_row_idx, base_var_name, idx, data, var_patterns)


def _fill_row_data(ws, row_idx: int, base_var_name: str, item_idx: int, data: Dict, var_patterns: List[str]):
    """
    Điền dữ liệu vào một dòng với số thứ tự cụ thể.
    """
    for col_idx, cell in enumerate(ws[row_idx], start=1):
        if cell.value and isinstance(cell.value, str):
            new_value = cell.value
            
            # Thay thế các biến trong dòng
            for var_name in var_patterns:
                # Tạo biến với số thứ tự
                var_with_idx = var_name + '_' + str(item_idx)
                value = data.get(var_with_idx, '')
                
                # Thay thế biến (có thể là không có số hoặc có số)
                pattern_no_num = '{{' + var_name + '}}'
                pattern_with_num = '{{' + var_name + '_' + str(item_idx) + '}}'
                
                if pattern_no_num in new_value:
                    new_value = new_value.replace(pattern_no_num, str(value))
                if pattern_with_num in new_value:
                    new_value = new_value.replace(pattern_with_num, str(value))
            
            # Chỉ cập nhật nếu không phải là merged cell (trừ top-left)
            try:
                if not _is_merged_cell(ws, cell):
                    cell.value = new_value
                    _set_wrap_text(cell)
            except Exception as cell_error:
                logger.debug(f"B? qua merged cell {cell.coordinate}: {cell_error}")


def _copy_row_format(ws, source_row_idx: int, dest_row_idx: int):
    """
    Copy format từ dòng nguồn sang dòng đích.
    """
    for col_idx, source_cell in enumerate(ws[source_row_idx], start=1):
        dest_cell = ws.cell(row=dest_row_idx, column=col_idx)
        
        # Copy value (sẽ được thay thế sau)
        if source_cell.value:
            dest_cell.value = source_cell.value
        
        # Copy alignment (giá»¯ wrap_text)
        try:
            if source_cell.alignment:
                dest_cell.alignment = Alignment(
                    horizontal=source_cell.alignment.horizontal,
                    vertical=source_cell.alignment.vertical,
                    wrap_text=True,  # Bật wrap text
                    shrink_to_fit=source_cell.alignment.shrink_to_fit if source_cell.alignment.shrink_to_fit else False,
                    text_rotation=source_cell.alignment.text_rotation if source_cell.alignment.text_rotation else 0,
                    indent=source_cell.alignment.indent if source_cell.alignment.indent else 0
                )
            else:
                dest_cell.alignment = Alignment(wrap_text=True)
        except Exception as e:
            logger.warning(f"Không thể copy alignment: {e}")
        
        # Copy number format
        dest_cell.number_format = source_cell.number_format


def _set_wrap_text(cell):
    """
    Bật wrap text cho cell nếu có dữ liệu.
    """
    if cell.value and str(cell.value).strip():
        try:
            if cell.alignment:
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical,
                    wrap_text=True,
                    shrink_to_fit=cell.alignment.shrink_to_fit if cell.alignment.shrink_to_fit else False,
                    text_rotation=cell.alignment.text_rotation if cell.alignment.text_rotation else 0,
                    indent=cell.alignment.indent if cell.alignment.indent else 0
                )
            else:
                cell.alignment = Alignment(wrap_text=True)
        except Exception as e:
            logger.warning(f"Không thể set wrap_text: {e}")


def _insert_image_to_cell(ws, image_url: str):
    """
    Tìm cell có chứa {{vd_image}} và chèn hình ảnh vào.
    Nếu image_url rỗng hoặc None, chỉ xóa biến {{vd_image}} khỏi cell.
    
    Args:
        ws: Worksheet object
        image_url: URL của hình ảnh (ví dụ: /api/images/filename.png) hoặc rỗng/None
    """
    # Tìm cell có chứa {{vd_image}}
    image_cell = None
    for row in ws.iter_rows():
        for cell in row:
            if _is_merged_cell(ws, cell):
                continue
            if cell.value and isinstance(cell.value, str):
                if '{{vd_image}}' in cell.value:
                    image_cell = cell
                    break
        if image_cell:
            break
    
    if not image_cell:
        logger.debug("Không tìm thấy cell có chứa {{vd_image}}")
        return
    
    # Xóa text {{vd_image}} trong cell (luôn luôn xóa, kể cả khi không có hình ảnh)
    cell_value = str(image_cell.value) if image_cell.value else ''
    cell_value = cell_value.replace('{{vd_image}}', '')
    image_cell.value = cell_value if cell_value.strip() else None
    
    # Nếu không có image_url hoặc rỗng, chỉ xóa biến và return
    if not image_url or not image_url.strip():
        logger.debug("Không có hình ảnh để chèn, đã xóa biến {{vd_image}}")
        return
    
    # Chuyển đổi URL thành đường dẫn file system
    # URL có dạng: /api/images/filename.png
    # Cần chuyển thành: D:\Data Analyst\Tools\kpi\qlcl\images\filename.png
    try:
        if image_url.startswith('/api/images/'):
            filename = image_url.replace('/api/images/', '')
            image_path = os.path.join(IMAGES_STORAGE_DIR, filename)
        else:
            # Nếu là đường dẫn tuyệt đối hoặc relative
            image_path = image_url
        
        # Kiểm tra file có tồn tại không
        if not os.path.exists(image_path):
            logger.warning(f"Hình ảnh không tồn tại: {image_path}")
            return

        # Chèn hình ảnh vào cell (image_cell đã được tìm ở trên)
        try:
            img = Image(image_path)
            
            # Lấy kích thước cell (width và height)
            col_letter = image_cell.column_letter
            row_idx = image_cell.row
            
            # Lấy column width (nếu không có, dùng default width = 8.43 characters)
            col_width = ws.column_dimensions[col_letter].width
            if col_width is None:
                col_width = 8.43  # Default column width
            
            # Lấy row height (nếu không có, dùng default height = 15 points)
            row_height = ws.row_dimensions[row_idx].height
            if row_height is None:
                row_height = 15.0  # Default row height
            
            # Chuyển đổi sang pixels - ĐIỀU CHỈNH HỆ SỐ CHUYỂN ĐỔI
            # Column width: 1 unit ≈ 7-8 pixels (tùy DPI)
            # Row height: 1 point â‰ˆ 1.33-1.5 pixels
            cell_width_px = col_width * 7.5  # Tăng hệ số từ 7 lên 7.5
            cell_height_px = row_height * 1.5  # Tăng hệ số từ 1.33 lên 1.5
            
            # Tính kích thước hình ảnh
            # Yêu cầu: chiều rộng ảnh = 120 pixels, chiều cao scale theo tỉ lệ
            img_width = img.width
            img_height = img.height
            aspect_ratio = img_width / img_height if img_height > 0 else 1
            
            # Đặt chiều rộng = 120 pixels (cố định)
            new_width = 120
            
            # Tính chiều cao theo tỉ lệ thuận (giữ aspect ratio)
            new_height = int(new_width / aspect_ratio)
            
            # Đảm bảo không vượt quá chiều cao cell
            # Nếu chiều cao vượt quá chiều cao cell, scale lại theo chiều cao cell
            if new_height > cell_height_px:
                new_height = int(cell_height_px)
                new_width = int(new_height * aspect_ratio)
            
            # HOẶC CÁCH 2: Fill toàn bộ cell (giữ tỉ lệ, fit vào cạnh lớn hơn)
            # Uncomment phần dưới nếu muốn ảnh lấp đầy cell hơn
            """
            width_ratio = cell_width_px / img_width
            height_ratio = cell_height_px / img_height
            
            # Chọn tỉ lệ LỚN HƠN để lấp đầy (thay vì min như trước)
            scale_ratio = max(width_ratio, height_ratio) * 0.95  # 0.95 để có chút padding
            
            new_width = int(img_width * scale_ratio)
            new_height = int(img_height * scale_ratio)
            """
            
            img.width = new_width
            img.height = new_height
            
            # Đặt vị trí hình ảnh tại cell
            image_cell_name = image_cell.coordinate
            img.anchor = image_cell_name
            
            # Thêm hình ảnh vào worksheet
            ws.add_image(img)
            
            logger.info(f"Đã chèn hình ảnh vào cell {image_cell_name}: {image_path} "
                       f"(kích thước: {new_width}x{new_height}px, cell: {cell_width_px:.1f}x{cell_height_px:.1f}px)")
        except Exception as img_error:
            logger.error(f"Lỗi khi chèn hình ảnh: {img_error}", exc_info=True)
            
    except Exception as e:
        logger.error(f"Lỗi khi xử lý hình ảnh: {e}", exc_info=True)


def _duplicate_section_rows(ws, base_var_name: str, max_items: int, data: Dict):
    """
    Tìm dòng có chứa biến {{base_var_name}} và duplicate nó thành nhiều dòng với số thứ tự.
    Template có biến không có số (ví dụ {{a_root_cause}}), code sẽ:
    1. Thay thế biến trong dòng template thành {{a_root_cause_1}}
    2. Duplicate dòng và tạo {{a_root_cause_2}}, {{a_root_cause_3}}, ...
    """
    template_row_idx = None
    
    # Tìm dòng template (dòng đầu tiên chứa biến {{base_var_name}} không có số)
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                # Tìm biến không có số (ví dụ {{a_root_cause}})
                if f'{{{{{base_var_name}}}}}' in cell.value:
                    template_row_idx = row_idx
                    break
        if template_row_idx:
            break
    
    if not template_row_idx:
        return  # Không tìm thấy template row
    
    # Xác định các biến cần thay thế dựa trên section
    var_patterns = []
    if base_var_name == 'a_root_cause':
        var_patterns = ['a_root_cause', 'a_hđkp_tuc_thoi', 'a_hđ_phong_ngua', 
                       'a_tg_theo_doi', 'a_trach_nhiem', 'a_tg_thuc_hien']
    elif base_var_name == 'b_root_cause':
        var_patterns = ['b_root_cause', 'b_hđkp_tuc_thoi', 'b_hđ_phong_ngua',
                       'b_tg_theo_doi', 'b_trach_nhiem', 'b_tg_thuc_hien']
    elif base_var_name == 'cong_viec':
        var_patterns = ['stt', 'cong_viec', 'trach_nhiem', 'ngay_bat_dau', 
                       'ngay_hoan_thanh', 'giam_sat', 'ket_qua', 'ket_luan']
    
    # Thay thế biến trong dòng template thành số 1
    for col_idx, template_cell in enumerate(ws[template_row_idx], start=1):
        if template_cell.value and isinstance(template_cell.value, str):
            new_value = template_cell.value
            for var_name in var_patterns:
                # Thay thế {{var_name}} thành {{var_name_1}}
                old_pattern = '{{' + var_name + '}}'
                new_pattern = '{{' + var_name + '_1' + '}}'
                new_value = new_value.replace(old_pattern, new_pattern)
            template_cell.value = new_value
    
    # Nếu chỉ có 1 item, không cần duplicate
    if max_items <= 1:
        return
    
    # Duplicate rows: chèn các dòng mới sau dòng template
    for idx in range(2, max_items + 1):
        # Insert row sau dòng template + (idx - 2)
        new_row_idx = template_row_idx + (idx - 1)
        ws.insert_rows(new_row_idx)
        
        # Copy format và công thức từ template row
        for col_idx, template_cell in enumerate(ws[template_row_idx], start=1):
            new_cell = ws.cell(row=new_row_idx, column=col_idx)
            
            # Copy value và thay thế biến
            if template_cell.value:
                if isinstance(template_cell.value, str):
                    # Thay thế tất cả các biến từ {{var_name_1}} thành {{var_name_idx}}
                    new_value = template_cell.value
                    for var_name in var_patterns:
                        # Tránh dùng f-string với quá nhiều dấu ngoặc nhọn
                        # Pattern: {{{{var_name}_1}}}} â†’ {{{{var_name}_{idx}}}}}}
                        old_pattern = '{{' + var_name + '_1' + '}}'
                        new_pattern = '{{' + var_name + '_' + str(idx) + '}}'
                        new_value = new_value.replace(old_pattern, new_pattern)
                    new_cell.value = new_value
                else:
                    new_cell.value = template_cell.value
            
            # Copy style - chỉ copy các thuộc tính cơ bản để tránh lỗi StyleProxy
            try:
                # Copy alignment (quan trọng nhất cho format)
                if template_cell.alignment and template_cell.alignment.horizontal is not None:
                    new_cell.alignment = Alignment(
                        horizontal=template_cell.alignment.horizontal,
                        vertical=template_cell.alignment.vertical,
                        wrap_text=template_cell.alignment.wrap_text if template_cell.alignment.wrap_text else False,
                        shrink_to_fit=template_cell.alignment.shrink_to_fit if template_cell.alignment.shrink_to_fit else False
                    )
                # Copy number format (string, an toàn)
                if template_cell.number_format:
                    new_cell.number_format = template_cell.number_format
            except Exception as style_error:
                # Nếu copy style lỗi, chỉ copy number_format
                logger.warning(f"Không thể copy style: {style_error}")
                try:
                    new_cell.number_format = template_cell.number_format
                except:
                    pass


def excel_to_pdf(excel_path: str, pdf_path: str):
    """
    Chuyển đổi Excel sang PDF sử dụng win32com (Windows only).
    Thiết lập page setup để fit to page width và loại bỏ margins thừa.
    """
    if not WIN32_AVAILABLE:
        raise HTTPException(status_code=500, detail="Không thể xuất PDF: win32com không khả dụng. Chỉ hỗ trợ trên Windows.")
    
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        # Má»Ÿ file Excel
        workbook = excel.Workbooks.Open(os.path.abspath(excel_path))
        
        # Thiết lập page setup cho tất cả các sheets
        for sheet in workbook.Worksheets:
            # Thiết lập margins nhỏ để tránh lề thừa
            sheet.PageSetup.LeftMargin = excel.InchesToPoints(0.25)  # 0.25 inch
            sheet.PageSetup.RightMargin = excel.InchesToPoints(0.25)
            sheet.PageSetup.TopMargin = excel.InchesToPoints(0.25)
            sheet.PageSetup.BottomMargin = excel.InchesToPoints(0.25)
            sheet.PageSetup.HeaderMargin = excel.InchesToPoints(0.1)
            sheet.PageSetup.FooterMargin = excel.InchesToPoints(0.1)
            
            # Fit to page width (1 page wide) để tránh tạo trang thừa do lề
            sheet.PageSetup.FitToPagesWide = 1
            sheet.PageSetup.FitToPagesTall = False  # Không giới hạn số trang theo chiều cao
            
            # Đảm bảo sử dụng fit to page
            sheet.PageSetup.Zoom = False
        
        # Xuất sang PDF
        workbook.ExportAsFixedFormat(0, os.path.abspath(pdf_path))  # 0 = xlTypePDF
        
        workbook.Close(False)
        excel.Quit()
        
        return pdf_path
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lá»—i khi chuyá»ƒn Excel sang PDF: {str(e)}")


def create_hdkp_pdf(error_id: int, hdkp_data: Dict) -> str:
    """
    Tạo file PDF HĐKP từ template Excel và dữ liệu.
    Returns: đường dẫn file PDF
    """
    try:
        # Format tên file
        filename = f"SAISOT_{hdkp_data.get('ma_nv', 'UNKNOWN')}_{hdkp_data.get('ngay_ghi_nhan', '').replace('-', '')}.pdf"
        pdf_path = os.path.join(PDF_STORAGE_DIR, filename)
        logger.info(f"Tạo PDF: {pdf_path}")
        
        # Tạo file Excel tạm
        temp_excel_path = pdf_path.replace('.pdf', '.xlsx')
        
        # Kiểm tra template có tồn tại không
        if not os.path.exists(EXCEL_TEMPLATE_PATH):
            raise FileNotFoundError(f"Template Excel không tồn tại: {EXCEL_TEMPLATE_PATH}")
        
        logger.info(f"Điền dữ liệu vào Excel template...")
        # Điền dữ liệu vào Excel template
        fill_excel_template(EXCEL_TEMPLATE_PATH, hdkp_data, temp_excel_path)
        logger.info(f"Đã điền xong, chuyển sang PDF...")
        
        # Chuyá»ƒn Excel sang PDF
        excel_to_pdf(temp_excel_path, pdf_path)
        logger.info(f"Đã tạo PDF thành công: {pdf_path}")
        
        # Xóa file Excel tạm
        if os.path.exists(temp_excel_path):
            os.remove(temp_excel_path)
        
        # Trả về URL để truy cập file
        return f"/api/pdf/{filename}"
        
    except Exception as e:
        logger.error(f"Lá»—i trong create_hdkp_pdf: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Lỗi khi tạo file PDF: {str(e)}")


app = FastAPI(title="QLCL KPI")
# app.add_middleware(SessionMiddleware, secret_key="qlcl-kpi-secret-please-change")
templates = Jinja2Templates(directory=TEMPLATES_DIR)
gemba_cp_templates = Jinja2Templates(directory=str(GEMBA_CP_TEMPLATES_DIR))
# Serve logo and other assets from templates folder under /templates
app.mount("/templates", StaticFiles(directory=TEMPLATES_DIR), name="templates")
# Serve PDF files
app.mount("/api/pdf", StaticFiles(directory=PDF_STORAGE_DIR), name="pdf")
# Serve uploaded images
app.mount("/api/images", StaticFiles(directory=IMAGES_STORAGE_DIR), name="images")
app.mount("/static", StaticFiles(directory=str(GEMBA_CP_STATIC_DIR)), name="gemba-static")

bootstrap_qlcl_schema()
GembaCPBase.metadata.create_all(bind=gemba_cp_engine)
app.include_router(gemba_dashboard_router.router, dependencies=[Depends(require_authenticated_api_user)])
app.include_router(gemba_admin_router.router, dependencies=[Depends(require_authenticated_api_user)])


@app.on_event("startup")
def startup_qtcn_auto_sync():
    start_qtcn_auto_sync_if_enabled()


@app.get("/")
def home(request: Request):
    user_data = get_authenticated_user(request)
    if not user_data:
        response = RedirectResponse(url="/login", status_code=303)
        response.delete_cookie("ma_nv")
        return response

    return templates.TemplateResponse("index.html", {"request": request, "user": user_data})


@app.get("/kpi")
def kpi_page(request: Request):
    if not get_authenticated_user(request):
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("kpi.html", {"request": request})


@app.get("/view-kpi")
def view_kpi(request: Request):
    if not get_authenticated_user(request):
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("view_kpi.html", {"request": request})


@app.get("/dashboard-summary")
def dashboard_summary(request: Request):
    if not get_authenticated_user(request):
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("dashboard_summary.html", {"request": request})


@app.get("/gemba-control-plan")
def gemba_control_plan_dashboard(request: Request):
    if not get_authenticated_user(request):
        return RedirectResponse(url=f"/login?next={quote(str(request.url.path), safe='/')}", status_code=303)
    dashboard_js_version = build_static_asset_version(GEMBA_CP_STATIC_DIR / "dashboard.js")
    return gemba_cp_templates.TemplateResponse(
        "index.html",
        {
            "request": request,
            "app_name": gemba_cp_settings.app_name,
            "sheet_name": gemba_cp_settings.google_sheets_worksheet,
            "dashboard_js_version": dashboard_js_version,
        },
    )


@app.get("/login")
def login_page(request: Request):
    next_url = (request.query_params.get("next") or "").strip()
    if get_authenticated_user(request):
        redirect_target = next_url if next_url.startswith("/") else "/"
        return RedirectResponse(url=redirect_target, status_code=303)
    return templates.TemplateResponse("login.html", {"request": request, "next_url": next_url})


@app.get("/kpi-input")
def kpi_input(request: Request):
    ma_nv_cookie = request.cookies.get("ma_nv")
    if not ma_nv_cookie:
        return RedirectResponse(url="/login", status_code=303)
    ma_nv = decode_ma_nv_cookie(ma_nv_cookie)
    ma_nv_variants = generate_ma_nv_variants(ma_nv)

    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT ma_nv, ho_ten, chuc_vu
                FROM public.quality_employees
                WHERE ma_nv = ANY(%s)
                ORDER BY CASE WHEN ma_nv = %s THEN 0 ELSE 1 END
                LIMIT 1
                """,
                (ma_nv_variants, ma_nv)
            )
            user_row = cur.fetchone()
            if not user_row:
                response = RedirectResponse(url="/login", status_code=303)
                response.delete_cookie("ma_nv")
                return response

    return templates.TemplateResponse("kpi_input.html", {"request": request, "step": "form", "user": dict(user_row)})


@app.post("/login")
def login(request: Request, ma_nv: str = Form(...), next_url: str = Form("")):
    ma_nv = (ma_nv or "").strip()
    next_url = (next_url or "").strip()
    redirect_target = next_url if next_url.startswith("/") else "/"
    if not ma_nv:
        return templates.TemplateResponse(
            "login.html",
            {"request": request, "error": "Vui lòng nhập mã nhân viên", "next_url": next_url},
            status_code=400,
        )
    ma_nv_upper = ma_nv.upper()
    ma_nv_variants = generate_ma_nv_variants(ma_nv)
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT ma_nv, ho_ten, chuc_vu
                FROM public.quality_employees
                WHERE ma_nv = ANY(%s)
                ORDER BY CASE WHEN ma_nv = %s THEN 0 ELSE 1 END
                LIMIT 1
                """,
                (ma_nv_variants, ma_nv_upper)
            )
            user_row = cur.fetchone()
            if not user_row:
                return templates.TemplateResponse(
                    "login.html",
                    {"request": request, "error": "Mã nhân viên không tồn tại", "next_url": next_url},
                    status_code=401,
                )

    response = RedirectResponse(url=redirect_target, status_code=303)
    response.set_cookie(
        key="ma_nv",
        value=encode_ma_nv_cookie(user_row["ma_nv"]),
        httponly=True,
        samesite="lax",
        max_age=2*24*60*60,
    )
    return response


@app.post("/logout")
def logout(request: Request):
    response = RedirectResponse(url="/login", status_code=303)
    response.delete_cookie("ma_nv")
    return response


@app.get("/api/tasks")
def api_tasks(request: Request):
    ma_nv_cookie = request.cookies.get("ma_nv")
    if not ma_nv_cookie:
        raise HTTPException(status_code=401, detail="Chưa đăng nhập")
    ma_nv = decode_ma_nv_cookie(ma_nv_cookie)
    ma_nv_variants = generate_ma_nv_variants(ma_nv)
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT chuc_vu
                FROM public.quality_employees
                WHERE ma_nv = ANY(%s)
                ORDER BY CASE WHEN ma_nv = %s THEN 0 ELSE 1 END
                LIMIT 1
                """,
                (ma_nv_variants, ma_nv)
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=403, detail="Không tìm thấy nhân viên")
            chuc_vu = row["chuc_vu"]
            cur.execute("SELECT task_name FROM public.tasks_qa WHERE chuc_vu=%s ORDER BY task_name", (chuc_vu,))
            tasks = [r["task_name"] for r in cur.fetchall()]
    return {"chuc_vu": chuc_vu, "tasks": tasks}


@app.post("/api/input")
async def api_input(request: Request):
    ma_nv_cookie = request.cookies.get("ma_nv")
    if not ma_nv_cookie:
        raise HTTPException(status_code=401, detail="Chưa đăng nhập")
    ma_nv = decode_ma_nv_cookie(ma_nv_cookie)
    ma_nv_variants = generate_ma_nv_variants(ma_nv)
    payload = await request.json()
    try:
        from_date_str = payload.get("from_date")
        to_date_str = payload.get("to_date")
        items = payload.get("items", [])
        if not from_date_str or not to_date_str or not items:
            raise ValueError("Thiếu dữ liệu bắt buộc")

        from_date_val = date.fromisoformat(from_date_str)
        to_date_val = date.fromisoformat(to_date_str)
        if to_date_val < from_date_val:
            raise ValueError("Khoảng thời gian không hợp lệ")

        for it in items:
            name = (it.get("task_name") or "").strip()
            qty = int(it.get("thuc_hien"))
            if not name:
                raise ValueError("Thiếu tên task")
            if qty < 0:
                raise ValueError("Số lượng phải >= 0")
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT ma_nv, chuc_vu
                FROM public.quality_employees
                WHERE ma_nv = ANY(%s)
                ORDER BY CASE WHEN ma_nv = %s THEN 0 ELSE 1 END
                LIMIT 1
                """,
                (ma_nv_variants, ma_nv)
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=403, detail="Không tìm thấy nhân viên")
            ma_nv_db = row["ma_nv"]
            chuc_vu = row["chuc_vu"]

            insert_sql = (
                "INSERT INTO public.input_qa (ma_nv, chuc_vu, from_date, to_date, task_name, thuc_hien)"
                " VALUES (%s, %s, %s, %s, %s, %s)"
                " RETURNING id"
            )
            ids: List[int] = []
            for it in items:
                cur.execute(
                    insert_sql,
                    (
                        ma_nv_db,
                        chuc_vu,
                        from_date_val,
                        to_date_val,
                        it["task_name"],
                        int(it["thuc_hien"]),
                    ),
                )
                ids.append(cur.fetchone()["id"])
            conn.commit()

    return {"status": "ok", "inserted": len(ids), "ids": ids}


@app.post("/api/input-error")
async def api_input_error(request: Request):
    ma_nv_cookie = request.cookies.get("ma_nv")
    if not ma_nv_cookie:
        raise HTTPException(status_code=401, detail="Chưa đăng nhập")
    ma_nv = decode_ma_nv_cookie(ma_nv_cookie)
    ma_nv_variants = generate_ma_nv_variants(ma_nv)

    payload = await request.json()
    # Optional fields allowed; minimal validation
    ngay_ghi_nhan = payload.get("ngay_ghi_nhan")  # ISO date or empty
    task_name = payload.get("task_name")
    phan_loai_loi = payload.get("phan_loai_loi")
    mo_ta = payload.get("mo_ta")
    muc_do_anh_huong = payload.get("muc_do_anh_huong")
    huong_giai_quyet = payload.get("huong_giai_quyet")

    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            # Resolve chuc_vu
            cur.execute(
                """
                SELECT ma_nv, chuc_vu
                FROM public.quality_employees
                WHERE ma_nv = ANY(%s)
                ORDER BY CASE WHEN ma_nv = %s THEN 0 ELSE 1 END
                LIMIT 1
                """,
                (ma_nv_variants, ma_nv)
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=403, detail="Không tìm thấy nhân viên")
            ma_nv_db = row["ma_nv"]
            chuc_vu = row["chuc_vu"]

            cur.execute(
                """
                INSERT INTO public.input_error (
                    ma_nv, chuc_vu, ngay_ghi_nhan, task_name, phan_loai_loi, mo_ta,
                    muc_do_anh_huong, huong_giai_quyet
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s
                ) RETURNING id
                """,
                (
                    ma_nv_db,
                    chuc_vu,
                    date.fromisoformat(ngay_ghi_nhan) if ngay_ghi_nhan else None,
                    task_name,
                    phan_loai_loi,
                    mo_ta,
                    muc_do_anh_huong,
                    huong_giai_quyet,
                ),
            )
            new_id = cur.fetchone()["id"]
            conn.commit()

    return {"status": "ok", "id": new_id}


@app.get("/api/input-edit")
def api_input_edit(
    request: Request,
    from_date: Optional[str] = Query(None, description="Lọc từ ngày (YYYY-MM-DD)"),
    to_date: Optional[str] = Query(None, description="Lọc đến ngày (YYYY-MM-DD)")
):
    """Lấy danh sách dữ liệu input_qa của người đăng nhập, sắp xếp từ mới nhất đến cũ nhất.
    Có thể lọc theo khoảng thời gian (from_date, to_date) - lọc các bản ghi có khoảng thời gian giao với khoảng filter.
    """
    ma_nv_cookie = request.cookies.get("ma_nv")
    if not ma_nv_cookie:
        raise HTTPException(status_code=401, detail="Chưa đăng nhập")
    ma_nv = decode_ma_nv_cookie(ma_nv_cookie)
    ma_nv_variants = generate_ma_nv_variants(ma_nv)
    
    # Parse filter dates
    from_date_val = None
    to_date_val = None
    if from_date:
        try:
            from_date_val = date.fromisoformat(from_date)
        except ValueError:
            raise HTTPException(status_code=400, detail="Định dạng from_date không hợp lệ (YYYY-MM-DD)")
    if to_date:
        try:
            to_date_val = date.fromisoformat(to_date)
        except ValueError:
            raise HTTPException(status_code=400, detail="Định dạng to_date không hợp lệ (YYYY-MM-DD)")
    
    if from_date_val and to_date_val and to_date_val < from_date_val:
        raise HTTPException(status_code=400, detail="Ngày Đến phải >= Ngày Từ")
    
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            # Xác định ma_nv chính xác từ database
            cur.execute(
                """
                SELECT ma_nv
                FROM public.quality_employees
                WHERE ma_nv = ANY(%s)
                ORDER BY CASE WHEN ma_nv = %s THEN 0 ELSE 1 END
                LIMIT 1
                """,
                (ma_nv_variants, ma_nv)
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=403, detail="Không tìm thấy nhân viên")
            ma_nv_db = row["ma_nv"]
            
            # Xây dựng query với filter theo ngày
            # Lọc các bản ghi có khoảng thời gian giao với khoảng filter
            # (from_date <= record.to_date AND to_date >= record.from_date)
            query = """
                SELECT id, from_date, to_date, task_name, thuc_hien, created_at
                FROM public.input_qa
                WHERE ma_nv = %s
            """
            params = [ma_nv_db]
            
            if from_date_val and to_date_val:
                # Cả hai ngày đều có: lọc các bản ghi có khoảng thời gian giao với khoảng filter
                query += " AND from_date <= %s AND to_date >= %s"
                params.extend([to_date_val, from_date_val])
            elif from_date_val:
                # Chỉ có from_date: lọc các bản ghi có to_date >= from_date
                query += " AND to_date >= %s"
                params.append(from_date_val)
            elif to_date_val:
                # Chỉ có to_date: lọc các bản ghi có from_date <= to_date
                query += " AND from_date <= %s"
                params.append(to_date_val)
            
            query += " ORDER BY created_at DESC"
            
            cur.execute(query, tuple(params))
            items = cur.fetchall()
            
            # Chuyển đổi date và datetime thành string
            result_items = []
            for item in items:
                result_items.append({
                    "id": item["id"],
                    "from_date": item["from_date"].isoformat() if item["from_date"] else None,
                    "to_date": item["to_date"].isoformat() if item["to_date"] else None,
                    "task_name": item["task_name"],
                    "thuc_hien": item["thuc_hien"],
                    "created_at": item["created_at"].isoformat() if item["created_at"] else None
                })
    
    return {"status": "ok", "items": result_items}


@app.put("/api/input-edit/{item_id}")
async def api_input_edit_update(item_id: int, request: Request):
    """Cập nhật dữ liệu input_qa theo ID"""
    ma_nv_cookie = request.cookies.get("ma_nv")
    if not ma_nv_cookie:
        raise HTTPException(status_code=401, detail="Chưa đăng nhập")
    ma_nv = decode_ma_nv_cookie(ma_nv_cookie)
    ma_nv_variants = generate_ma_nv_variants(ma_nv)
    
    payload = await request.json()
    try:
        from_date_str = payload.get("from_date")
        to_date_str = payload.get("to_date")
        task_name = payload.get("task_name")
        thuc_hien = payload.get("thuc_hien")
        
        if not from_date_str or not to_date_str or not task_name:
            raise ValueError("Thiếu dữ liệu bắt buộc")
        
        from_date_val = date.fromisoformat(from_date_str)
        to_date_val = date.fromisoformat(to_date_str)
        if to_date_val < from_date_val:
            raise ValueError("Khoảng thời gian không hợp lệ")
        
        task_name = task_name.strip()
        if not task_name:
            raise ValueError("Tên công việc không được để trống")
        
        thuc_hien = int(thuc_hien) if thuc_hien is not None else 0
        if thuc_hien < 0:
            raise ValueError("Số lượng phải >= 0")
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            # Xác định ma_nv chính xác từ database
            cur.execute(
                """
                SELECT ma_nv
                FROM public.quality_employees
                WHERE ma_nv = ANY(%s)
                ORDER BY CASE WHEN ma_nv = %s THEN 0 ELSE 1 END
                LIMIT 1
                """,
                (ma_nv_variants, ma_nv)
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=403, detail="Không tìm thấy nhân viên")
            ma_nv_db = row["ma_nv"]
            
            # Kiểm tra xem record có thuộc về người đăng nhập không
            cur.execute(
                """
                SELECT id, ma_nv
                FROM public.input_qa
                WHERE id = %s
                """,
                (item_id,)
            )
            record = cur.fetchone()
            if not record:
                raise HTTPException(status_code=404, detail="Không tìm thấy bản ghi")
            
            if record["ma_nv"] != ma_nv_db:
                raise HTTPException(status_code=403, detail="Bạn không có quyền chỉnh sửa bản ghi này")
            
            # Cập nhật dữ liệu
            cur.execute(
                """
                UPDATE public.input_qa
                SET from_date = %s, to_date = %s, task_name = %s, thuc_hien = %s
                WHERE id = %s AND ma_nv = %s
                RETURNING id, from_date, to_date, task_name, thuc_hien, created_at
                """,
                (from_date_val, to_date_val, task_name, thuc_hien, item_id, ma_nv_db)
            )
            updated = cur.fetchone()
            if not updated:
                raise HTTPException(status_code=500, detail="Không thể cập nhật dữ liệu")
            conn.commit()
            
            return {
                "status": "ok",
                "id": updated["id"],
                "from_date": updated["from_date"].isoformat() if updated["from_date"] else None,
                "to_date": updated["to_date"].isoformat() if updated["to_date"] else None,
                "task_name": updated["task_name"],
                "thuc_hien": updated["thuc_hien"],
                "created_at": updated["created_at"].isoformat() if updated["created_at"] else None
            }


@app.get("/api/summary")
def api_summary(
    chuc_vu: str = Query(..., regex="^(QAPL|QANL|QAQT)$"),
    week: Optional[int] = Query(None, ge=1, le=53),
    year: Optional[int] = Query(None, ge=2000, le=2100)
):
    week_clause = ""
    params_qa = [chuc_vu]
    if week is not None:
        week_clause += " AND to_char(GREATEST(from_date, to_date), 'IW') = %s"
        params_qa.append(f"{int(week):02d}")
    if year is not None:
        week_clause += " AND to_char(GREATEST(from_date, to_date), 'YYYY') = %s"
        params_qa.append(str(year))

    err_clause = ""
    err_params = [chuc_vu]
    if week is not None:
        err_clause += " AND to_char(ngay_ghi_nhan, 'IW') = %s"
        err_params.append(f"{int(week):02d}")
    if year is not None:
        err_clause += " AND to_char(ngay_ghi_nhan, 'YYYY') = %s"
        err_params.append(str(year))

    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT task_name FROM public.tasks_qa WHERE chuc_vu=%s ORDER BY task_name", (chuc_vu,))
            tasks = [r["task_name"] for r in cur.fetchall()]

            cur.execute(
                f"""
                WITH qa AS (
                  SELECT 
                    to_char(GREATEST(from_date, to_date), '"W"IW-YYYY') AS week_year,
                    ma_nv,
                    task_name,
                    SUM(thuc_hien) AS thuc_hien
                  FROM public.input_qa
                  WHERE chuc_vu = %s
                    AND from_date IS NOT NULL
                    AND to_date IS NOT NULL
                  {week_clause}
                  GROUP BY 1,2,3
                )
                SELECT qa.week_year, d.ma_nv, d.ho_ten, qa.task_name, qa.thuc_hien
                FROM qa
                JOIN public.quality_employees d ON d.ma_nv = qa.ma_nv
                WHERE d.chuc_vu = %s
                  AND qa.week_year IS NOT NULL
                ORDER BY qa.week_year, d.ma_nv, qa.task_name
                """,
                (*params_qa, chuc_vu),
            )
            qa_rows = cur.fetchall()

            cur.execute(
                f"""
                SELECT 
                  to_char(ngay_ghi_nhan, '"W"IW-YYYY') AS week_year,
                  e.ma_nv,
                  e.task_name,
                  COUNT(*) AS sai_sot
                FROM public.input_error e
                JOIN public.quality_employees d ON d.ma_nv = e.ma_nv
                WHERE d.chuc_vu = %s
                  AND e.ngay_ghi_nhan IS NOT NULL
                  {err_clause}
                GROUP BY 1,2,3
                HAVING to_char(ngay_ghi_nhan, '"W"IW-YYYY') IS NOT NULL
                ORDER BY 1,2,3
                """,
                tuple(err_params),
            )
            err_rows = cur.fetchall()

    rows_map: Dict[Tuple[str, str], Dict] = {}
    for r in qa_rows:
        if not r.get("week_year") or not r.get("ma_nv"):
            continue
        key = (r["week_year"], r["ma_nv"])
        entry = rows_map.setdefault(key, {"week_year": r["week_year"], "ma_nv": r["ma_nv"], "ho_ten": r.get("ho_ten"), "metrics": {}})
        entry["metrics"].setdefault(r["task_name"], {"thuc_hien": 0, "sai_sot": 0})
        entry["metrics"][r["task_name"]]["thuc_hien"] += int(r["thuc_hien"] or 0)
    
    for r in err_rows:
        if not r.get("week_year") or not r.get("ma_nv"):
            continue
        key = (r["week_year"], r["ma_nv"])
        entry = rows_map.setdefault(key, {"week_year": r["week_year"], "ma_nv": r["ma_nv"], "ho_ten": None, "metrics": {}})
        entry["metrics"].setdefault(r["task_name"], {"thuc_hien": 0, "sai_sot": 0})
        entry["metrics"][r["task_name"]]["sai_sot"] += int(r["sai_sot"] or 0)

    rows = list(rows_map.values())
    rows.sort(key=lambda x: (
        x.get("week_year") or "",
        x.get("ma_nv") or ""
    ))

    return {"tasks": tasks, "rows": rows}


def _normalize_month_filter(month_str: Optional[str]) -> Optional[str]:
    if not month_str:
        return None
    month_str = month_str.strip()
    for fmt in ("%y-%m", "%Y-%m"):
        try:
            dt = datetime.strptime(month_str, fmt)
            return dt.strftime("%Y-%m")
        except ValueError:
            continue
    return None


@app.get("/api/summary/monthly")
def api_summary_monthly(
    chuc_vu: str = Query(..., regex="^(QAPL|QANL|QAQT)$"),
    month: Optional[str] = Query(None)
):
    month_filter = _normalize_month_filter(month)

    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT task_name FROM public.tasks_qa WHERE chuc_vu=%s ORDER BY task_name", (chuc_vu,))
            tasks = [r["task_name"] for r in cur.fetchall()]

            qa_params = [chuc_vu]
            qa_month_clause = ""
            if month_filter:
                qa_month_clause = " AND to_char(from_date, 'YYYY-MM') = %s"
                qa_params.append(month_filter)

            cur.execute(
                f"""
                WITH qa AS (
                  SELECT 
                    to_char(from_date, 'YYYY-MM') AS month_key,
                    ma_nv,
                    task_name,
                    SUM(thuc_hien) AS thuc_hien
                  FROM public.input_qa
                  WHERE chuc_vu = %s
                    AND from_date IS NOT NULL
                  {qa_month_clause}
                  GROUP BY 1,2,3
                )
                SELECT qa.month_key, d.ma_nv, d.ho_ten, qa.task_name, qa.thuc_hien
                FROM qa
                JOIN public.quality_employees d ON d.ma_nv = qa.ma_nv
                WHERE d.chuc_vu = %s
                  AND qa.month_key IS NOT NULL
                ORDER BY qa.month_key, d.ma_nv, qa.task_name
                """,
                (*qa_params, chuc_vu),
            )
            qa_rows = cur.fetchall()

            err_params = [chuc_vu]
            err_month_clause = ""
            if month_filter:
                err_month_clause = " AND to_char(ngay_ghi_nhan, 'YYYY-MM') = %s"
                err_params.append(month_filter)

            cur.execute(
                f"""
                SELECT 
                  to_char(ngay_ghi_nhan, 'YYYY-MM') AS month_key,
                  e.ma_nv,
                  e.task_name,
                  COUNT(*) AS sai_sot
                FROM public.input_error e
                JOIN public.quality_employees d ON d.ma_nv = e.ma_nv
                WHERE d.chuc_vu = %s
                  AND e.ngay_ghi_nhan IS NOT NULL
                  {err_month_clause}
                GROUP BY 1,2,3
                HAVING to_char(ngay_ghi_nhan, 'YYYY-MM') IS NOT NULL
                ORDER BY 1,2,3
                """,
                tuple(err_params),
            )
            err_rows = cur.fetchall()

    rows_map: Dict[Tuple[str, str], Dict] = {}

    def convert_month_key(month_key: Optional[str]) -> Tuple[str, str]:
        if not month_key:
            return "", ""
        try:
            dt = datetime.strptime(month_key, "%Y-%m")
            label = dt.strftime("%y-%m")
            iso_first_day = dt.replace(day=1).date().isoformat()
            return label, iso_first_day
        except ValueError:
            return month_key, ""

    for r in qa_rows:
        month_label, first_day = convert_month_key(r.get("month_key"))
        if not month_label or not r.get("ma_nv"):
            continue
        key = (month_label, r["ma_nv"])
        entry = rows_map.setdefault(
            key,
            {
                "month_year": month_label,
                "from_date": first_day,
                "ma_nv": r["ma_nv"],
                "ho_ten": r.get("ho_ten"),
                "metrics": {}
            }
        )
        entry["metrics"].setdefault(r["task_name"], {"thuc_hien": 0, "sai_sot": 0})
        entry["metrics"][r["task_name"]]["thuc_hien"] += int(r["thuc_hien"] or 0)

    for r in err_rows:
        month_label, first_day = convert_month_key(r.get("month_key"))
        if not month_label or not r.get("ma_nv"):
            continue
        key = (month_label, r["ma_nv"])
        entry = rows_map.setdefault(
            key,
            {
                "month_year": month_label,
                "from_date": first_day,
                "ma_nv": r["ma_nv"],
                "ho_ten": None,
                "metrics": {}
            }
        )
        entry["from_date"] = entry.get("from_date") or first_day
        entry["metrics"].setdefault(r["task_name"], {"thuc_hien": 0, "sai_sot": 0})
        entry["metrics"][r["task_name"]]["sai_sot"] += int(r["sai_sot"] or 0)

    rows = list(rows_map.values())
    rows.sort(key=lambda x: (
        x.get("month_year") or "",
        x.get("ma_nv") or ""
    ))

    return {"tasks": tasks, "rows": rows}


@app.get("/api/errors")
def api_errors(chuc_vu: Optional[str] = Query(None, regex="^(QAPL|QANL|QAQT)$")):
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            base = (
                "SELECT id, ngay_ghi_nhan, chuc_vu, ma_nv, task_name, phan_loai_loi, mo_ta, muc_do_anh_huong, huong_giai_quyet, "
                "hanh_dong, trach_nhiem, thoi_han, tien_do, ngay_hoan_thanh, ket_luan, cap_form, ghi_chu, created_at "
                "FROM public.input_error"
            )
            if chuc_vu:
                cur.execute(base + " WHERE chuc_vu=%s ORDER BY created_at DESC", (chuc_vu,))
            else:
                cur.execute(base + " ORDER BY created_at DESC")
            rows = cur.fetchall()
            
            # Thêm flag để biết có cần HĐKP không
            for row in rows:
                row['needs_hdkp'] = row.get('huong_giai_quyet', '') and 'HĐKP' in str(row.get('huong_giai_quyet', '')).upper()
    return {"rows": rows}


@app.patch("/api/error/{error_id}")
async def api_error_update(error_id: int = Path(...), request: Request = None):
    payload = await request.json()
    # 'tien_do' is derived from 'ngay_hoan_thanh' and MUST NOT be set directly
    allowed_fields = {
        "ngay_ghi_nhan", "task_name", "phan_loai_loi", "mo_ta", "muc_do_anh_huong", "huong_giai_quyet",
        "hanh_dong", "trach_nhiem", "thoi_han", "ngay_hoan_thanh", "ket_luan", "cap_form", "ghi_chu"
    }
    sets = []
    values = []
    for k, v in payload.items():
        if k in allowed_fields:
            if k in ("ngay_ghi_nhan", "thoi_han", "ngay_hoan_thanh") and v:
                try:
                    v = date.fromisoformat(v)
                except Exception:
                    raise HTTPException(status_code=400, detail=f"Ngày không hợp lệ cho {k}")
            sets.append(f"{k} = %s")
            values.append(v)
    # Auto set tien_do based on ngay_hoan_thanh when provided (rule)
    if "ngay_hoan_thanh" in payload:
        sets.append("tien_do = %s")
        values.append("Hoàn thành" if payload.get("ngay_hoan_thanh") else "Chưa hoàn thành")
    if not sets:
        raise HTTPException(status_code=400, detail="Không có trường hợp lệ để cập nhật")

    values.append(error_id)
    sql = f"UPDATE public.input_error SET {', '.join(sets)} WHERE id = %s RETURNING id"

    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(sql, tuple(values))
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Không tìm thấy lỗi")
            conn.commit()
    return {"status": "ok", "id": row["id"]}


@app.get("/api/error/{error_id}")
def api_get_error(error_id: int = Path(...)):
    """Lấy thông tin chi tiết của một error"""
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT id, ngay_ghi_nhan, chuc_vu, ma_nv, task_name, phan_loai_loi, mo_ta, 
                       muc_do_anh_huong, huong_giai_quyet, hanh_dong, trach_nhiem, 
                       thoi_han, tien_do, ngay_hoan_thanh, ket_luan, cap_form, ghi_chu, created_at
                FROM public.input_error WHERE id = %s
                """,
                (error_id,)
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Không tìm thấy lỗi")
            
            # Thêm flag để biết có cần HĐKP không
            row['needs_hdkp'] = row.get('huong_giai_quyet', '') and 'HĐKP' in str(row.get('huong_giai_quyet', '')).upper()
            
            return {"error": dict(row)}


@app.get("/api/hdkp/{error_id}")
def api_get_hdkp(error_id: int = Path(...)):
    """Lấy dữ liệu HĐKP đã lưu cho một error"""
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            # Lấy Section I: Mô tả vấn đề
            cur.execute("SELECT * FROM public.cap_mota WHERE error_id = %s", (error_id,))
            mota = cur.fetchone()
            
            # Lấy Section II: Kế hoạch
            cur.execute(
                "SELECT * FROM public.cap_kehoach WHERE error_id = %s ORDER BY section, thu_tu",
                (error_id,)
            )
            kehoach = cur.fetchall()
            
            # Lấy Section III: Chi tiết
            cur.execute(
                "SELECT * FROM public.cap_chitiet WHERE error_id = %s ORDER BY thu_tu",
                (error_id,)
            )
            chitiet = cur.fetchall()
    
    return {
        "mota": dict(mota) if mota else None,
        "kehoach": [dict(r) for r in kehoach],
        "chitiet": [dict(r) for r in chitiet]
    }


@app.get("/hdkp-form/{error_id}")
def hdkp_form_page(error_id: int = Path(...), request: Request = None):
    """Trang form nhập HĐKP cho một error"""
    ma_nv = request.cookies.get("ma_nv") if request else None
    if not ma_nv:
        return RedirectResponse(url="/login", status_code=303)
    
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                "SELECT id, ngay_ghi_nhan, ma_nv, task_name, phan_loai_loi, mo_ta, huong_giai_quyet FROM public.input_error WHERE id = %s",
                (error_id,)
            )
            error_row = cur.fetchone()
            if not error_row:
                raise HTTPException(status_code=404, detail="Không tìm thấy lỗi")
            
            # Kiểm tra có cần HĐKP không
            needs_hdkp = error_row.get('huong_giai_quyet', '') and 'HĐKP' in str(error_row.get('huong_giai_quyet', '')).upper()
            if not needs_hdkp:
                raise HTTPException(status_code=400, detail="Lỗi này không yêu cầu HĐKP")
    
    return templates.TemplateResponse("hdkp_form.html", {
        "request": request,
        "error": dict(error_row),
        "error_id": error_id
    })


@app.post("/api/hdkp/{error_id}")
async def api_save_hdkp(error_id: int = Path(...), request: Request = None):
    """
    Lưu thông tin HĐKP cho một error vào 3 tables: cap_mota, cap_kehoach, cap_chitiet
    """
    ma_nv = request.cookies.get("ma_nv")
    if not ma_nv:
        raise HTTPException(status_code=401, detail="Chưa đăng nhập")
    
    payload = await request.json()
    
    # Lấy thông tin error
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                "SELECT id, ma_nv, ngay_ghi_nhan FROM public.input_error WHERE id = %s",
                (error_id,)
            )
            error_row = cur.fetchone()
            if not error_row:
                raise HTTPException(status_code=404, detail="Không tìm thấy lỗi")
            
            # 1. Lưu Section I: Mô tả vấn đề vào cap_mota
            mota_data = payload.get('mota', {})
            cur.execute(
                """
                INSERT INTO public.cap_mota (
                    error_id, vd_what, vd_when, vd_who, vd_where, 
                    vd_how, vd_before, vd_importance, vd_image
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (error_id) DO UPDATE SET
                    vd_what = EXCLUDED.vd_what,
                    vd_when = EXCLUDED.vd_when,
                    vd_who = EXCLUDED.vd_who,
                    vd_where = EXCLUDED.vd_where,
                    vd_how = EXCLUDED.vd_how,
                    vd_before = EXCLUDED.vd_before,
                    vd_importance = EXCLUDED.vd_importance,
                    vd_image = EXCLUDED.vd_image,
                    updated_at = NOW()
                RETURNING id
                """,
                (
                    error_id,
                    mota_data.get('vd_what'),
                    mota_data.get('vd_when'),
                    mota_data.get('vd_who'),
                    mota_data.get('vd_where'),
                    mota_data.get('vd_how'),
                    mota_data.get('vd_before'),
                    mota_data.get('vd_importance'),
                    mota_data.get('vd_image'),
                )
            )
            
            # 2. Lưu Section II: Kế hoạch khắc phục vào cap_kehoach
            # Xóa các rows cũ
            cur.execute("DELETE FROM public.cap_kehoach WHERE error_id = %s", (error_id,))
            
            # Lưu các rows mới
            kehoach_list = payload.get('kehoach', [])
            for idx, item in enumerate(kehoach_list, 1):
                cur.execute(
                    """
                    INSERT INTO public.cap_kehoach (
                        error_id, section, root_cause, hdkp_tuc_thoi, hd_phong_ngua,
                        tg_theo_doi, trach_nhiem, tg_thuc_hien, thu_tu
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        error_id,
                        item.get('section'),  # 'A' hoặc 'B'
                        item.get('root_cause'),
                        item.get('hdkp_tuc_thoi'),
                        item.get('hd_phong_ngua'),
                        item.get('tg_theo_doi'),
                        item.get('trach_nhiem'),
                        item.get('tg_thuc_hien'),
                        idx,
                    )
                )
            
            # 3. Lưu Section III: Chi tiết kế hoạch vào cap_chitiet
            # Xóa các rows cũ
            cur.execute("DELETE FROM public.cap_chitiet WHERE error_id = %s", (error_id,))
            
            # Lưu các rows mới
            chitiet_list = payload.get('chitiet', [])
            for idx, item in enumerate(chitiet_list, 1):
                ngay_bat_dau = date.fromisoformat(item['ngay_bat_dau']) if item.get('ngay_bat_dau') else None
                ngay_hoan_thanh = date.fromisoformat(item['ngay_hoan_thanh']) if item.get('ngay_hoan_thanh') else None
                
                cur.execute(
                    """
                    INSERT INTO public.cap_chitiet (
                        error_id, cong_viec, trach_nhiem, ngay_bat_dau, ngay_hoan_thanh,
                        giam_sat, ket_qua, ket_luan, thu_tu
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        error_id,
                        item.get('cong_viec'),
                        item.get('trach_nhiem'),
                        ngay_bat_dau,
                        ngay_hoan_thanh,
                        item.get('giam_sat'),
                        item.get('ket_qua'),  # 'Đạt' hoặc 'Không đạt'
                        item.get('ket_luan'),
                        idx,
                    )
                )
            
            conn.commit()
            
            # 4. Tạo PDF nếu có đủ dữ liệu
            pdf_url = None
            if error_row.get('ngay_ghi_nhan'):
                try:
                    # Lấy dữ liệu từ 3 tables để tạo PDF
                    logger.info(f"Bắt đầu tạo PDF cho error_id={error_id}")
                    hdkp_data = prepare_hdkp_data_for_pdf(error_id, error_row)
                    logger.info(f"Dữ liệu HĐKP đã chuẩn bị: {len(hdkp_data)} fields")
                    pdf_url = create_hdkp_pdf(error_id, hdkp_data)
                    logger.info(f"PDF đã được tạo: {pdf_url}")
                    
                    # Cập nhật cap_form với URL PDF
                    cur.execute(
                        "UPDATE public.input_error SET cap_form = %s WHERE id = %s",
                        (pdf_url, error_id)
                    )
                    conn.commit()
                    logger.info(f"Đã cập nhật cap_form cho error_id={error_id}")
                except Exception as e:
                    # Log error để debug
                    logger.error(f"Lỗi khi tạo PDF cho error_id={error_id}: {str(e)}", exc_info=True)
                    # Không raise để không làm gián đoạn việc lưu dữ liệu
    
    return {"status": "ok", "id": error_id, "pdf_url": pdf_url}


@app.post("/api/hdkp/{error_id}/regenerate-pdf")
async def api_regenerate_hdkp_pdf(error_id: int = Path(...), request: Request = None):
    """
    Tạo lại PDF HĐKP từ dữ liệu đã có (để kiểm tra lề và format).
    """
    ma_nv = request.cookies.get("ma_nv")
    if not ma_nv:
        raise HTTPException(status_code=401, detail="Chưa đăng nhập")
    
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            # Lấy thông tin error
            cur.execute(
                "SELECT id, ma_nv, ngay_ghi_nhan FROM public.input_error WHERE id = %s",
                (error_id,)
            )
            error_row = cur.fetchone()
            if not error_row:
                raise HTTPException(status_code=404, detail="Không tìm thấy lỗi")
            
            # Kiểm tra có dữ liệu HĐKP không
            cur.execute("SELECT COUNT(*) as cnt FROM public.cap_mota WHERE error_id = %s", (error_id,))
            has_data = cur.fetchone()['cnt'] > 0
            if not has_data:
                raise HTTPException(status_code=400, detail="Chưa có dữ liệu HĐKP để tạo PDF")
            
            # Tạo lại PDF
            try:
                hdkp_data = prepare_hdkp_data_for_pdf(error_id, error_row)
                pdf_url = create_hdkp_pdf(error_id, hdkp_data)
                
                # Cập nhật cap_form với URL PDF mới
                cur.execute(
                    "UPDATE public.input_error SET cap_form = %s WHERE id = %s",
                    (pdf_url, error_id)
                )
                conn.commit()
                
                return {"status": "ok", "pdf_url": pdf_url, "message": "PDF đã được tạo lại thành công"}
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"Lỗi khi tạo lại PDF: {str(e)}")


@app.get("/api/qc/hdkp/{dps_id}")
def api_get_qc_hdkp(dps_id: int = Path(...)):
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT * FROM public.qc_hdkp_mota WHERE qc_error_dps_id = %s", (dps_id,))
            mota = cur.fetchone()
            cur.execute(
                "SELECT * FROM public.qc_hdkp_kehoach WHERE qc_error_dps_id = %s ORDER BY section, thu_tu",
                (dps_id,),
            )
            kehoach = cur.fetchall()
            cur.execute(
                "SELECT * FROM public.qc_hdkp_chitiet WHERE qc_error_dps_id = %s ORDER BY thu_tu",
                (dps_id,),
            )
            chitiet = cur.fetchall()

    return {
        "mota": dict(mota) if mota else None,
        "kehoach": [dict(r) for r in kehoach],
        "chitiet": [dict(r) for r in chitiet],
    }


@app.get("/hdkp-form-endline/{dps_id}")
def hdkp_form_endline_page(dps_id: int = Path(...), request: Request = None):
    ma_nv = request.cookies.get("ma_nv") if request else None
    if not ma_nv:
        return RedirectResponse(url="/login", status_code=303)

    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT dps.id, dps.plan_id, dps.date, dps.time_bucket, dps.loai_loi, dps.ma_loi, dps.vi_tri
                FROM public.qc_error_dps dps
                WHERE dps.id = %s
                """,
                (dps_id,),
            )
            dps_row = cur.fetchone()
            if not dps_row:
                raise HTTPException(status_code=404, detail="Khong tim thay DPS")

    return templates.TemplateResponse("hdkp_form_endline.html", {
        "request": request,
        "dps": dict(dps_row),
        "dps_id": dps_id,
    })


@app.post("/api/qc/hdkp/{dps_id}")
async def api_save_qc_hdkp(dps_id: int = Path(...), request: Request = None):
    ma_nv = request.cookies.get("ma_nv")
    if not ma_nv:
        raise HTTPException(status_code=401, detail="Chua dang nhap")

    payload = await request.json()

    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                "SELECT id, date FROM public.qc_error_dps WHERE id = %s",
                (dps_id,),
            )
            dps_row = cur.fetchone()
            if not dps_row:
                raise HTTPException(status_code=404, detail="Khong tim thay DPS")

            mota_data = payload.get('mota', {})
            cur.execute(
                """
                INSERT INTO public.qc_hdkp_mota (
                    qc_error_dps_id, vd_what, vd_when, vd_who, vd_where,
                    vd_how, vd_before, vd_importance, vd_image
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (qc_error_dps_id) DO UPDATE SET
                    vd_what = EXCLUDED.vd_what,
                    vd_when = EXCLUDED.vd_when,
                    vd_who = EXCLUDED.vd_who,
                    vd_where = EXCLUDED.vd_where,
                    vd_how = EXCLUDED.vd_how,
                    vd_before = EXCLUDED.vd_before,
                    vd_importance = EXCLUDED.vd_importance,
                    vd_image = EXCLUDED.vd_image,
                    updated_at = NOW()
                """,
                (
                    dps_id,
                    mota_data.get('vd_what'),
                    mota_data.get('vd_when'),
                    mota_data.get('vd_who'),
                    mota_data.get('vd_where'),
                    mota_data.get('vd_how'),
                    mota_data.get('vd_before'),
                    mota_data.get('vd_importance'),
                    mota_data.get('vd_image'),
                ),
            )

            cur.execute("DELETE FROM public.qc_hdkp_kehoach WHERE qc_error_dps_id = %s", (dps_id,))
            kehoach_list = payload.get('kehoach', [])
            for idx, item in enumerate(kehoach_list, 1):
                cur.execute(
                    """
                    INSERT INTO public.qc_hdkp_kehoach (
                        qc_error_dps_id, section, root_cause, hdkp_tuc_thoi, hd_phong_ngua,
                        tg_theo_doi, trach_nhiem, tg_thuc_hien, thu_tu
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        dps_id,
                        item.get('section'),
                        item.get('root_cause'),
                        item.get('hdkp_tuc_thoi'),
                        item.get('hd_phong_ngua'),
                        item.get('tg_theo_doi'),
                        item.get('trach_nhiem'),
                        item.get('tg_thuc_hien'),
                        idx,
                    ),
                )

            cur.execute("DELETE FROM public.qc_hdkp_chitiet WHERE qc_error_dps_id = %s", (dps_id,))
            chitiet_list = payload.get('chitiet', [])
            for idx, item in enumerate(chitiet_list, 1):
                ngay_bat_dau = date.fromisoformat(item['ngay_bat_dau']) if item.get('ngay_bat_dau') else None
                ngay_hoan_thanh = date.fromisoformat(item['ngay_hoan_thanh']) if item.get('ngay_hoan_thanh') else None
                cur.execute(
                    """
                    INSERT INTO public.qc_hdkp_chitiet (
                        qc_error_dps_id, cong_viec, trach_nhiem, ngay_bat_dau, ngay_hoan_thanh,
                        giam_sat, ket_qua, ket_luan, thu_tu
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        dps_id,
                        item.get('cong_viec'),
                        item.get('trach_nhiem'),
                        ngay_bat_dau,
                        ngay_hoan_thanh,
                        item.get('giam_sat'),
                        item.get('ket_qua'),
                        item.get('ket_luan'),
                        idx,
                    ),
                )

            conn.commit()

            pdf_url = None
            try:
                hdkp_data = prepare_qc_hdkp_data_for_pdf(dps_id, dps_row)
                pdf_url = create_hdkp_pdf(dps_id, hdkp_data)
                cur.execute(
                    "UPDATE public.qc_error_dps SET hdkp_pdf = %s WHERE id = %s",
                    (pdf_url, dps_id),
                )
                conn.commit()
            except Exception as e:
                logger.error(f"Loi khi tao PDF HDKP QC dps_id={dps_id}: {str(e)}", exc_info=True)

    return {"status": "ok", "id": dps_id, "pdf_url": pdf_url}


def prepare_hdkp_data_for_pdf(error_id: int, error_row: Dict) -> Dict:
    """
    Chuẩn bị dữ liệu từ 3 tables để tạo PDF.
    Tập hợp tất cả dữ liệu và format cho mapping vào Excel template.
    """
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            # Lấy Section I: Mô tả vấn đề
            cur.execute("SELECT * FROM public.cap_mota WHERE error_id = %s", (error_id,))
            mota = cur.fetchone()
            
            # Lấy Section II: Kế hoạch (section A và B)
            cur.execute(
                "SELECT * FROM public.cap_kehoach WHERE error_id = %s ORDER BY section, thu_tu",
                (error_id,)
            )
            kehoach_rows = cur.fetchall()
            
            # Lấy Section III: Chi tiết
            cur.execute(
                "SELECT * FROM public.cap_chitiet WHERE error_id = %s ORDER BY thu_tu",
                (error_id,)
            )
            chitiet_rows = cur.fetchall()
            
            # Lấy họ và tên của người điền input error
            ho_ten = ''
            ma_nv_error = error_row.get('ma_nv')
            if ma_nv_error:
                cur.execute("SELECT ho_ten FROM public.quality_employees WHERE ma_nv = %s", (ma_nv_error,))
                ho_ten_row = cur.fetchone()
                if ho_ten_row:
                    ho_ten = ho_ten_row.get('ho_ten', '')
    
    # Format created_at từ cap_mota thành dd/mm/yyyy
    created_at_formatted = ''
    if mota and mota.get('created_at'):
        try:
            # created_at là datetime object từ psycopg2
            created_at = mota.get('created_at')
            if isinstance(created_at, str):
                # Parse string thành datetime (nếu là string)
                created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
            
            # Format thành dd/mm/yyyy
            if isinstance(created_at, datetime):
                created_at_formatted = created_at.strftime('%d/%m/%Y')
            elif isinstance(created_at, date):
                created_at_formatted = created_at.strftime('%d/%m/%Y')
        except Exception as e:
            logger.warning(f"Không thể format created_at: {e}")
            created_at_formatted = ''
    
    # Tạo data dict với các biến {{...}}
    data = {
        'ma_nv': error_row.get('ma_nv', ''),
        'ngay_ghi_nhan': str(error_row.get('ngay_ghi_nhan', '')),
        'created_at': created_at_formatted,
        'ho_ten': ho_ten,
    }
    
    # Section I: Mô tả vấn đề
    if mota:
        data.update({
            'vd_what': mota.get('vd_what', ''),
            'vd_when': mota.get('vd_when', ''),
            'vd_who': mota.get('vd_who', ''),
            'vd_where': mota.get('vd_where', ''),
            'vd_how': mota.get('vd_how', ''),
            'vd_before': mota.get('vd_before', ''),
            'vd_importance': mota.get('vd_importance', ''),
            'vd_image': mota.get('vd_image', ''),
        })
    
    # Section II: Kế hoạch - Format theo section A và B với số thứ tự
    section_a_items = [r for r in kehoach_rows if r.get('section') == 'A']
    section_b_items = [r for r in kehoach_rows if r.get('section') == 'B']
    
    # Format section A vá»›i sá»‘ thá»© tá»± (1, 2, 3, ...)
    for idx, item in enumerate(section_a_items, start=1):
        data.update({
            f'a_root_cause_{idx}': item.get('root_cause', ''),
            f'a_hđkp_tuc_thoi_{idx}': item.get('hdkp_tuc_thoi', ''),
            f'a_hđ_phong_ngua_{idx}': item.get('hd_phong_ngua', ''),
            f'a_tg_theo_doi_{idx}': item.get('tg_theo_doi', ''),
            f'a_trach_nhiem_{idx}': item.get('trach_nhiem', ''),
            f'a_tg_thuc_hien_{idx}': item.get('tg_thuc_hien', ''),
        })
    
    # Format section B vá»›i sá»‘ thá»© tá»± (1, 2, 3, ...)
    for idx, item in enumerate(section_b_items, start=1):
        data.update({
            f'b_root_cause_{idx}': item.get('root_cause', ''),
            f'b_hđkp_tuc_thoi_{idx}': item.get('hdkp_tuc_thoi', ''),
            f'b_hđ_phong_ngua_{idx}': item.get('hd_phong_ngua', ''),
            f'b_tg_theo_doi_{idx}': item.get('tg_theo_doi', ''),
            f'b_trach_nhiem_{idx}': item.get('trach_nhiem', ''),
            f'b_tg_thuc_hien_{idx}': item.get('tg_thuc_hien', ''),
        })
    
    # Section III: Chi tiết với số thứ tự (1, 2, 3, ...) và thêm {{stt}}
    for idx, item in enumerate(chitiet_rows, start=1):
        ket_qua = (item.get('ket_qua') or '').strip()
        # Tách ket_qua thành 2 cột: Đạt (trái) và Không đạt (phải)
        ket_qua_dat = ''  # Cột trái (Đạt)
        ket_qua_khong_dat = ''  # Cột phải (Không đạt)
        if ket_qua == 'Đạt':
            ket_qua_dat = 'x'
        elif ket_qua == 'Không đạt':
            ket_qua_khong_dat = 'x'
        
        data.update({
            f'stt_{idx}': str(idx),  # Số thứ tự tăng dần
            f'cong_viec_{idx}': item.get('cong_viec', ''),
            f'trach_nhiem_{idx}': item.get('trach_nhiem', ''),
            f'ngay_bat_dau_{idx}': str(item.get('ngay_bat_dau', '')) if item.get('ngay_bat_dau') else '',
            f'ngay_hoan_thanh_{idx}': str(item.get('ngay_hoan_thanh', '')) if item.get('ngay_hoan_thanh') else '',
            f'giam_sat_{idx}': item.get('giam_sat', ''),
            f'ket_qua_dat_{idx}': ket_qua_dat,  # Cột trái: "x" nếu Đạt, trống nếu không
            f'ket_qua_khong_dat_{idx}': ket_qua_khong_dat,  # Cột phải: "x" nếu Không đạt, trống nếu không
            f'ket_luan_{idx}': item.get('ket_luan', ''),
        })
    
    return data


def prepare_qc_hdkp_data_for_pdf(dps_id: int, dps_row: Dict) -> Dict:
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT * FROM public.qc_hdkp_mota WHERE qc_error_dps_id = %s", (dps_id,))
            mota = cur.fetchone()

            cur.execute(
                "SELECT * FROM public.qc_hdkp_kehoach WHERE qc_error_dps_id = %s ORDER BY section, thu_tu",
                (dps_id,),
            )
            kehoach_rows = cur.fetchall()

            cur.execute(
                "SELECT * FROM public.qc_hdkp_chitiet WHERE qc_error_dps_id = %s ORDER BY thu_tu",
                (dps_id,),
            )
            chitiet_rows = cur.fetchall()

    created_at_formatted = ''
    if mota and mota.get('created_at'):
        try:
            created_at = mota.get('created_at')
            if isinstance(created_at, str):
                created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
            if isinstance(created_at, datetime):
                created_at_formatted = created_at.strftime('%d/%m/%Y')
            elif isinstance(created_at, date):
                created_at_formatted = created_at.strftime('%d/%m/%Y')
        except Exception as e:
            logger.warning(f"Khong the format created_at QC HDKP: {e}")
            created_at_formatted = ''

    data = {
        'ma_nv': '',
        'ngay_ghi_nhan': str(dps_row.get('date', '')),
        'created_at': created_at_formatted,
        'ho_ten': '',
    }

    if mota:
        data.update({
            'vd_what': mota.get('vd_what', ''),
            'vd_when': mota.get('vd_when', ''),
            'vd_who': mota.get('vd_who', ''),
            'vd_where': mota.get('vd_where', ''),
            'vd_how': mota.get('vd_how', ''),
            'vd_before': mota.get('vd_before', ''),
            'vd_importance': mota.get('vd_importance', ''),
            'vd_image': mota.get('vd_image', ''),
        })

    section_a_items = [r for r in kehoach_rows if r.get('section') == 'A']
    section_b_items = [r for r in kehoach_rows if r.get('section') == 'B']

    for idx, item in enumerate(section_a_items, start=1):
        data.update({
            f'a_root_cause_{idx}': item.get('root_cause', ''),
            f'a_hdkp_tuc_thoi_{idx}': item.get('hdkp_tuc_thoi', ''),
            f'a_hd_phong_ngua_{idx}': item.get('hd_phong_ngua', ''),
            f'a_tg_theo_doi_{idx}': item.get('tg_theo_doi', ''),
            f'a_trach_nhiem_{idx}': item.get('trach_nhiem', ''),
            f'a_tg_thuc_hien_{idx}': item.get('tg_thuc_hien', ''),
        })

    for idx, item in enumerate(section_b_items, start=1):
        data.update({
            f'b_root_cause_{idx}': item.get('root_cause', ''),
            f'b_hdkp_tuc_thoi_{idx}': item.get('hdkp_tuc_thoi', ''),
            f'b_hd_phong_ngua_{idx}': item.get('hd_phong_ngua', ''),
            f'b_tg_theo_doi_{idx}': item.get('tg_theo_doi', ''),
            f'b_trach_nhiem_{idx}': item.get('trach_nhiem', ''),
            f'b_tg_thuc_hien_{idx}': item.get('tg_thuc_hien', ''),
        })

    for idx, item in enumerate(chitiet_rows, start=1):
        ket_qua = (item.get('ket_qua') or '').strip()
        ket_qua_dat = ''
        ket_qua_khong_dat = ''
        if ket_qua == 'Đạt':
            ket_qua_dat = 'x'
        elif ket_qua == 'Không đạt':
            ket_qua_khong_dat = 'x'

        data.update({
            f'stt_{idx}': str(idx),
            f'cong_viec_{idx}': item.get('cong_viec', ''),
            f'trach_nhiem_{idx}': item.get('trach_nhiem', ''),
            f'ngay_bat_dau_{idx}': str(item.get('ngay_bat_dau', '')) if item.get('ngay_bat_dau') else '',
            f'ngay_hoan_thanh_{idx}': str(item.get('ngay_hoan_thanh', '')) if item.get('ngay_hoan_thanh') else '',
            f'giam_sat_{idx}': item.get('giam_sat', ''),
            f'ket_qua_dat_{idx}': ket_qua_dat,
            f'ket_qua_khong_dat_{idx}': ket_qua_khong_dat,
            f'ket_luan_{idx}': item.get('ket_luan', ''),
        })

    return data


@app.post("/api/upload-image")
async def upload_image(request: Request, file: UploadFile = File(...)):
    """
    Upload hình ảnh cho HĐKP.
    Returns: URL của file đã upload
    """
    # Kiá»ƒm tra file type
    allowed_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp'}
    file_ext = PathLib(file.filename).suffix.lower()
    if file_ext not in allowed_extensions:
        raise HTTPException(status_code=400, detail=f"File không hợp lệ. Chỉ chấp nhận: {', '.join(allowed_extensions)}")
    
    # Tạo tên file unique
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_filename = f"{timestamp}_{file.filename}"
    file_path = os.path.join(IMAGES_STORAGE_DIR, safe_filename)
    
    # LÆ°u file
    try:
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        # Return URL
        image_url = f"/api/images/{safe_filename}"
        return {"status": "ok", "url": image_url, "filename": safe_filename}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lá»—i khi upload file: {str(e)}")


# ==================== QC MODULE ====================

DON_VI_OPTIONS = ["XN1-V1", "XN2", "XN3", "XNDT", "XNV2"]


@app.get("/qc")
def qc_page(request: Request):
    user_data = get_authenticated_user(request)
    if not user_data:
        return RedirectResponse(url="/login", status_code=303)
    role = (user_data.get("department") or "").upper()
    if role == "QC":
        return RedirectResponse(url="/qc-input", status_code=303)
    if role in QC_DON_VI_SCOPED_ROLES:
        return RedirectResponse(url="/qc/dashboard", status_code=303)
    return templates.TemplateResponse(
        "qc.html",
        build_qc_template_context(request, user_data, don_vi_options=DON_VI_OPTIONS),
    )


@app.get("/qc-login")
def qc_login_page(request: Request):
    if get_authenticated_user(request, qc_only=True):
        return RedirectResponse(url="/qc", status_code=303)
    return templates.TemplateResponse("login.html", {"request": request, "qc_mode": True})

@app.post("/qc-login")
def qc_login(request: Request, ma_nv: str = Form(...)):
    ma_nv = (ma_nv or "").strip()
    if not ma_nv:
        return templates.TemplateResponse("login.html", {"request": request, "error": "Vui lòng nhập mã nhân viên", "qc_mode": True}, status_code=400)
    ma_nv_upper = ma_nv.upper()
    ma_nv_variants = generate_ma_nv_variants(ma_nv)
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT ma_nv, ho_ten, chuc_vu
                FROM public.quality_employees
                WHERE ma_nv = ANY(%s) AND chuc_vu = 'QC'
                ORDER BY CASE WHEN ma_nv = %s THEN 0 ELSE 1 END
                LIMIT 1
                """,
                (ma_nv_variants, ma_nv_upper)
            )
            user_row = cur.fetchone()
            if not user_row:
                return templates.TemplateResponse("login.html", {"request": request, "error": "Bạn không có quyền QC (chuc_vu != 'QC') hoặc mã nhân viên không đúng", "qc_mode": True}, status_code=401)

    response = RedirectResponse(url="/qc", status_code=303)
    response.set_cookie(
        key="ma_nv",
        value=encode_ma_nv_cookie(user_row["ma_nv"]),
        httponly=True,
        samesite="lax",
        max_age=2*24*60*60,
    )
    return response


@app.get("/qc-input")
def qc_input_page(request: Request):
    user_data = get_authenticated_user(request, qc_only=True)
    if not user_data:
        # Require login if missing or not QC
        return RedirectResponse(url="/qc-login", status_code=303)

    return templates.TemplateResponse(
        "qc_input_sp.html",
        build_qc_template_context(request, user_data, don_vi_options=DON_VI_OPTIONS),
    )


@app.get("/qc-input-2")
def qc_input_sp_page(request: Request):
    return RedirectResponse(url="/qc-input", status_code=303)

@app.get("/qc/settings/customer")
def qc_settings_customer(request: Request):
    user = get_authenticated_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    if (user.get("department") or "").upper() != "QAQT":
        return RedirectResponse(url="/qc", status_code=303)
    return templates.TemplateResponse(
        "qc_settings_customer.html",
        build_qc_template_context(request, user),
    )

@app.get("/qc/settings/details")
def qc_settings_details(request: Request):
    user = get_authenticated_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    if (user.get("department") or "").upper() != "QAQT":
        return RedirectResponse(url="/qc", status_code=303)
    return templates.TemplateResponse(
        "qc_settings_details.html",
        build_qc_template_context(request, user),
    )

@app.get("/qc/settings/qc-list")
def qc_settings_qc_list(request: Request):
    user = get_authenticated_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    if (user.get("department") or "").upper() != "QAQT":
        return RedirectResponse(url="/qc", status_code=303)
    return templates.TemplateResponse(
        "qc_settings_qcs.html",
        build_qc_template_context(request, user, don_vi_options=DON_VI_OPTIONS),
    )

@app.get("/qc/cap")
def qc_cap_page(request: Request):
    user = get_authenticated_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("cap.html", build_qc_template_context(request, user))

@app.get("/qc/dashboard")
def qc_dashboard_page(request: Request):
    user = get_authenticated_user(request)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    return templates.TemplateResponse("qc_dashboard.html", build_qc_template_context(request, user))

@app.get("/api/qc/dashboard/filters")
def api_qc_dashboard_filters(
    request: Request,
    don_vi: Optional[str] = Query(None),
    bo_phan: Optional[str] = Query(None),
    ma_hang: Optional[str] = Query(None),
    type_name: Optional[str] = Query(None),
):
    """Return filter options for QC Dashboard."""
    don_vi, scoped_user = resolve_qc_don_vi_scope(request, don_vi)
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            don_vi_filter_params: List[Any] = []
            don_vi_filter_sql = ""
            if don_vi:
                don_vi_filter_sql = "AND don_vi = %s"
                don_vi_filter_params.append(don_vi)
            cur.execute(
                f"""
                SELECT DISTINCT don_vi
                FROM public.prod_plan
                WHERE don_vi IS NOT NULL AND don_vi <> ''
                  {don_vi_filter_sql}
                ORDER BY don_vi
                """,
                tuple(don_vi_filter_params),
            )
            don_vi_options = [r["don_vi"] for r in cur.fetchall()]

            station_params = []
            station_where = []
            if don_vi:
                station_where.append("p.don_vi = %s")
                station_params.append(don_vi)
            cur.execute(f"""
                SELECT DISTINCT o.station
                FROM public.qc_output_sp_log o
                JOIN public.prod_plan p ON p.id = o.plan_id
                WHERE o.station IS NOT NULL AND o.station <> ''
                {("AND " + " AND ".join(station_where)) if station_where else ""}
                ORDER BY o.station
            """, tuple(station_params))
            station_options = [r["station"] for r in cur.fetchall()]

            # "Tổ" on QC dashboard must reflect QC employee team (quality_employees.bo_phan),
            # i.e. filter by the team of the QC employee code (ma_nv) that created each row.
            bo_phan_params: List = []
            bo_phan_where = ["qe.bo_phan IS NOT NULL", "qe.bo_phan <> ''"]
            if don_vi:
                bo_phan_where.append("p.don_vi = %s")
                bo_phan_params.append(don_vi)
            cur.execute(
                f"""
                SELECT DISTINCT qe.bo_phan AS bo_phan
                FROM public.qc_output_sp_log o
                JOIN public.prod_plan p ON p.id = o.plan_id
                LEFT JOIN public.quality_employees qe ON qe.ma_nv = o.ma_nv
                WHERE {" AND ".join(bo_phan_where)}
                ORDER BY qe.bo_phan
                """,
                tuple(bo_phan_params),
            )
            bo_phan_options = [r["bo_phan"] for r in cur.fetchall()]
            if not bo_phan_options:
                fallback_params: List = []
                fallback_where = ["chuc_vu = 'QC'", "bo_phan IS NOT NULL", "bo_phan <> ''"]
                if don_vi:
                    fallback_where.append("don_vi = %s")
                    fallback_params.append(don_vi)
                cur.execute(
                    f"""
                    SELECT DISTINCT bo_phan
                    FROM public.quality_employees
                    WHERE {" AND ".join(fallback_where)}
                    ORDER BY bo_phan
                    """,
                    tuple(fallback_params),
                )
                bo_phan_options = [r["bo_phan"] for r in cur.fetchall()]

            ma_hang_params = []
            ma_hang_where = []
            if don_vi:
                ma_hang_where.append("p.don_vi = %s")
                ma_hang_params.append(don_vi)
            cur.execute(f"""
                SELECT DISTINCT p.ma_hang
                FROM public.prod_plan p
                WHERE p.ma_hang IS NOT NULL AND p.ma_hang <> ''
                {("AND " + " AND ".join(ma_hang_where)) if ma_hang_where else ""}
                ORDER BY p.ma_hang
            """, tuple(ma_hang_params))
            ma_hang_options = [r["ma_hang"] for r in cur.fetchall()]

            type_params = []
            type_where = []
            if don_vi:
                type_where.append("p.don_vi = %s")
                type_params.append(don_vi)
            cur.execute(f"""
                SELECT DISTINCT tgt.type
                FROM public.prod_plan p
                JOIN public.dm_loai_hang lh ON lh.ten_loai = p.loai_hang
                JOIN public.dm_loai_hang_target tgt ON tgt.id_type = lh.id_type
                WHERE tgt.type IS NOT NULL AND tgt.type <> ''
                {("AND " + " AND ".join(type_where)) if type_where else ""}
                ORDER BY tgt.type
            """, tuple(type_params))
            type_options = [r["type"] for r in cur.fetchall()]
            default_type_name = type_name or ("Thường" if "Thường" in type_options else None)
            default_station = resolve_dashboard_default_station(
                cur,
                scoped_user=scoped_user,
                don_vi=don_vi,
                bo_phan=bo_phan,
                ma_hang=ma_hang,
                type_name=default_type_name,
            )

    return {
        "don_vi": don_vi_options,
        "bo_phan": bo_phan_options,
        "station": station_options,
        "default_station": default_station,
        "ma_hang": ma_hang_options,
        "type": type_options,
    }

@app.get("/api/qc/dashboard/prev-date")
def api_qc_dashboard_prev_date(
    request: Request,
    date_before: Optional[str] = Query(None),
    don_vi: Optional[str] = Query(None),
    bo_phan: Optional[str] = Query(None),
    station: Optional[str] = Query(None),
    ma_hang: Optional[str] = Query(None),
    type_name: Optional[str] = Query(None),
):
    """Return the nearest previous day (before date_before) that has data, to skip off days."""
    don_vi, scoped_user = resolve_qc_don_vi_scope(request, don_vi)
    if not date_before:
        date_before = datetime.now().strftime("%Y-%m-%d")

    # "Tổ" filter is based on QC employee team only.
    resolved_bo_phan_expr = "COALESCE(NULLIF(qe.bo_phan, ''), '')"

    where_clauses = ["o.date < %s"]
    params: List = [date_before]

    if don_vi:
        where_clauses.append("p.don_vi = %s")
        params.append(don_vi)
    if bo_phan:
        where_clauses.append(f"{resolved_bo_phan_expr} = %s")
        params.append(bo_phan)
    if ma_hang:
        where_clauses.append("COALESCE(p.ma_hang, '') = %s")
        params.append(ma_hang)
    if type_name:
        where_clauses.append("COALESCE(tgt.type, '') = %s")
        params.append(type_name)

    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            station_for_calc = resolve_dashboard_default_station(
                cur,
                station=station,
                scoped_user=scoped_user,
                don_vi=don_vi,
                bo_phan=bo_phan,
                ma_hang=ma_hang,
                type_name=type_name,
                date_before=date_before,
                resolved_bo_phan_expr=resolved_bo_phan_expr,
            )
            if station_for_calc:
                where_clauses.append("COALESCE(o.station, '') = %s")
                params.append(station_for_calc)

            where_sql = " AND ".join(where_clauses)
            cur.execute(
                f"""
                SELECT MAX(o.date) AS prev_date
                FROM public.qc_output_sp_log o
                JOIN public.prod_plan p ON p.id = o.plan_id
                LEFT JOIN public.quality_employees qe ON qe.ma_nv = o.ma_nv
                LEFT JOIN public.dm_loai_hang lh ON lh.ten_loai = p.loai_hang
                LEFT JOIN public.dm_loai_hang_target tgt ON tgt.id_type = lh.id_type
                WHERE {where_sql}
                """,
                tuple(params),
            )
            row = cur.fetchone() or {}

    prev_date = row.get("prev_date")
    return {"status": "ok", "date": str(prev_date) if prev_date else None}

@app.get("/api/qc/dashboard")
def api_qc_dashboard(
    request: Request,
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    focus_date: Optional[str] = Query(None),
    don_vi: Optional[str] = Query(None),
    bo_phan: Optional[str] = Query(None),
    station: Optional[str] = Query(None),
    ma_hang: Optional[str] = Query(None),
    type_name: Optional[str] = Query(None),
    detail_bo_phan: Optional[str] = Query(None),
    ma_loi_bo_phan: Optional[str] = Query(None),
    ma_loi_chi_tiet: Optional[str] = Query(None),
):
    """QC Dashboard data source."""
    don_vi, scoped_user = resolve_qc_don_vi_scope(request, don_vi)
    if not date_to:
        date_to = datetime.now().strftime("%Y-%m-%d")
    if not date_from:
        # Default: last 7 days (inclusive) to match dashboard UX.
        try:
            dt_to = datetime.strptime(date_to, "%Y-%m-%d")
            date_from = (dt_to - timedelta(days=7)).strftime("%Y-%m-%d")
        except Exception:
            date_from = date_to
    if not focus_date:
        focus_date = date_to

    def _bucket_case(time_expr: str) -> str:
        return f"""
            CASE
                WHEN {time_expr} < '09:30' THEN '7H30 - 9H30'
                WHEN {time_expr} >= '09:30' AND {time_expr} < '11:30' THEN '9H30 - 11H30'
                WHEN {time_expr} >= '11:30' AND {time_expr} < '14:30' THEN '12H30 - 14H30'
                WHEN {time_expr} >= '14:30' AND {time_expr} < '16:30' THEN '14H30 - 16H30'
                ELSE 'Sau 16H30'
            END
        """

    bucket_order = ["7H30 - 9H30", "9H30 - 11H30", "12H30 - 14H30", "14H30 - 16H30", "Sau 16H30"]

    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            # Detect prod_plan.bo_phan type for compatibility
            cur.execute("""
                SELECT data_type
                FROM information_schema.columns
                WHERE table_schema = 'public'
                  AND table_name = 'prod_plan'
                  AND column_name = 'bo_phan'
                LIMIT 1
            """)
            bo_phan_type = (cur.fetchone() or {}).get("data_type")
            use_jsonb_bo_phan = bo_phan_type == "jsonb"

            # NOTE: Use a non-conflicting alias for this LATERAL join because
            # other queries in this endpoint also join dm_bo_phan as "bp".
            bo_phan_expr = "bp_plan.bo_phan_text" if use_jsonb_bo_phan else "p.bo_phan"
            bo_phan_join = (
                """
                LEFT JOIN LATERAL (
                    SELECT string_agg(value, ', ') AS bo_phan_text
                    FROM jsonb_array_elements_text(COALESCE(p.bo_phan, '[]'::jsonb)) AS t(value)
                ) bp_plan ON true
                """
                if use_jsonb_bo_phan
                else ""
            )
            # "Tổ" filter must be based on QC employee team only (qe.bo_phan).
            resolved_bo_phan_expr = "COALESCE(NULLIF(qe.bo_phan, ''), '')"

            station_for_calc = resolve_dashboard_default_station(
                cur,
                station=station,
                scoped_user=scoped_user,
                don_vi=don_vi,
                bo_phan=bo_phan,
                ma_hang=ma_hang,
                type_name=type_name,
                date_from=date_from,
                date_to=date_to,
                resolved_bo_phan_expr=resolved_bo_phan_expr,
            )

            # Filters used for KPI/charts (force station_for_calc).
            base_clauses_calc = ["o.date BETWEEN %s AND %s"]
            base_params_calc: List = [date_from, date_to]
            if don_vi:
                base_clauses_calc.append("p.don_vi = %s")
                base_params_calc.append(don_vi)
            if bo_phan:
                base_clauses_calc.append(f"{resolved_bo_phan_expr} = %s")
                base_params_calc.append(bo_phan)
            if station_for_calc:
                base_clauses_calc.append("COALESCE(o.station, '') = %s")
                base_params_calc.append(station_for_calc)
            if ma_hang:
                base_clauses_calc.append("COALESCE(p.ma_hang, '') = %s")
                base_params_calc.append(ma_hang)
            if type_name:
                base_clauses_calc.append("COALESCE(tgt.type, '') = %s")
                base_params_calc.append(type_name)

            where_sql_calc = " AND ".join(base_clauses_calc)

            # Filters used for the station summary table (show all stations).
            base_clauses_station = ["o.date BETWEEN %s AND %s"]
            base_params_station: List = [date_from, date_to]
            if don_vi:
                base_clauses_station.append("p.don_vi = %s")
                base_params_station.append(don_vi)
            if bo_phan:
                base_clauses_station.append(f"{resolved_bo_phan_expr} = %s")
                base_params_station.append(bo_phan)
            if ma_hang:
                base_clauses_station.append("COALESCE(p.ma_hang, '') = %s")
                base_params_station.append(ma_hang)
            if type_name:
                base_clauses_station.append("COALESCE(tgt.type, '') = %s")
                base_params_station.append(type_name)

            where_sql_station = " AND ".join(base_clauses_station)

            # A. Mixed chart by date (defect rate + target line)
            cur.execute(f"""
                SELECT
                    o.date,
                    SUM(o.delta) AS inspected_total,
                    SUM(CASE WHEN o.status = 'Failed' THEN o.delta ELSE 0 END) AS defect_total,
                    CASE
                        WHEN SUM(o.delta) > 0
                        THEN ROUND(SUM(CASE WHEN o.status = 'Failed' THEN o.delta ELSE 0 END)::numeric * 100.0 / SUM(o.delta), 4)
                        ELSE 0
                    END AS defect_rate,
                    CASE
                        WHEN SUM(o.delta) > 0
                        THEN ROUND(SUM(o.delta * COALESCE(tgt.target_percent, 0))::numeric / SUM(o.delta), 4)
                        ELSE 0
                    END AS target_rate
                FROM public.qc_output_sp_log o
                JOIN public.prod_plan p ON p.id = o.plan_id
                LEFT JOIN public.quality_employees qe ON qe.ma_nv = o.ma_nv
                {bo_phan_join}
                LEFT JOIN public.dm_loai_hang lh ON lh.ten_loai = p.loai_hang
                LEFT JOIN public.dm_loai_hang_target tgt ON tgt.id_type = lh.id_type
                WHERE {where_sql_calc}
                GROUP BY o.date
                ORDER BY o.date
            """, tuple(base_params_calc))
            daily_rows = cur.fetchall()

            # By station summary
            cur.execute(f"""
                SELECT
                    COALESCE(NULLIF(o.station, ''), 'N/A') AS station,
                    SUM(o.delta) AS inspected_total,
                    SUM(CASE WHEN o.status = 'Failed' THEN o.delta ELSE 0 END) AS defect_total,
                    CASE
                        WHEN SUM(o.delta) > 0
                        THEN ROUND(SUM(CASE WHEN o.status = 'Failed' THEN o.delta ELSE 0 END)::numeric * 100.0 / SUM(o.delta), 4)
                        ELSE 0
                    END AS defect_rate,
                    CASE
                        WHEN SUM(o.delta) > 0
                        THEN ROUND(SUM(o.delta * COALESCE(tgt.target_percent, 0))::numeric / SUM(o.delta), 4)
                        ELSE 0
                    END AS target_rate
                FROM public.qc_output_sp_log o
                JOIN public.prod_plan p ON p.id = o.plan_id
                LEFT JOIN public.quality_employees qe ON qe.ma_nv = o.ma_nv
                {bo_phan_join}
                LEFT JOIN public.dm_loai_hang lh ON lh.ten_loai = p.loai_hang
                LEFT JOIN public.dm_loai_hang_target tgt ON tgt.id_type = lh.id_type
                WHERE {where_sql_station}
                GROUP BY COALESCE(NULLIF(o.station, ''), 'N/A')
                ORDER BY inspected_total DESC, station
            """, tuple(base_params_station))
            station_rows = cur.fetchall()

            # B. Defect rate by time buckets in selected day
            bucket_clauses = ["o.date = %s"]
            bucket_params: List = [focus_date]
            if don_vi:
                bucket_clauses.append("p.don_vi = %s")
                bucket_params.append(don_vi)
            if bo_phan:
                bucket_clauses.append(f"{resolved_bo_phan_expr} = %s")
                bucket_params.append(bo_phan)
            if station_for_calc:
                bucket_clauses.append("COALESCE(o.station, '') = %s")
                bucket_params.append(station_for_calc)
            if ma_hang:
                bucket_clauses.append("COALESCE(p.ma_hang, '') = %s")
                bucket_params.append(ma_hang)
            if type_name:
                bucket_clauses.append("COALESCE(tgt.type, '') = %s")
                bucket_params.append(type_name)
            bucket_where_sql = " AND ".join(bucket_clauses)

            cur.execute(f"""
                SELECT
                    {_bucket_case("timezone('Asia/Bangkok', o.created_at)::time")} AS time_bucket,
                    SUM(o.delta) AS inspected_total,
                    SUM(CASE WHEN o.status = 'Failed' THEN o.delta ELSE 0 END) AS defect_total,
                    CASE
                        WHEN SUM(o.delta) > 0
                        THEN ROUND(SUM(CASE WHEN o.status = 'Failed' THEN o.delta ELSE 0 END)::numeric * 100.0 / SUM(o.delta), 4)
                        ELSE 0
                    END AS defect_rate,
                    CASE
                        WHEN SUM(o.delta) > 0
                        THEN ROUND(SUM(o.delta * COALESCE(tgt.target_percent, 0))::numeric / SUM(o.delta), 4)
                        ELSE 0
                    END AS target_rate
                FROM public.qc_output_sp_log o
                JOIN public.prod_plan p ON p.id = o.plan_id
                LEFT JOIN public.quality_employees qe ON qe.ma_nv = o.ma_nv
                {bo_phan_join}
                LEFT JOIN public.dm_loai_hang lh ON lh.ten_loai = p.loai_hang
                LEFT JOIN public.dm_loai_hang_target tgt ON tgt.id_type = lh.id_type
                WHERE {bucket_where_sql}
                GROUP BY time_bucket
            """, tuple(bucket_params))
            bucket_rows = {r["time_bucket"]: r for r in cur.fetchall()}

            # "Thống kê theo vị trí" (panel cạnh chart B): defects theo (Bộ phận -> Chi tiết)
            # cho focus_date, dùng chung bộ lọc với chart B.
            pos_clauses = ["sp.date = %s"]
            pos_params: List = [focus_date]
            if don_vi:
                pos_clauses.append("p.don_vi = %s")
                pos_params.append(don_vi)
            if bo_phan:
                pos_clauses.append(f"{resolved_bo_phan_expr} = %s")
                pos_params.append(bo_phan)
            if station_for_calc:
                pos_clauses.append("COALESCE(sp.station, '') = %s")
                pos_params.append(station_for_calc)
            if ma_hang:
                pos_clauses.append("COALESCE(p.ma_hang, '') = %s")
                pos_params.append(ma_hang)
            if type_name:
                pos_clauses.append("COALESCE(tgt.type, '') = %s")
                pos_params.append(type_name)
            pos_where_sql = " AND ".join(pos_clauses)

            cur.execute(
                f"""
                SELECT
                    COALESCE(NULLIF(bp.ten_bo_phan, ''), '(Trống)') AS bo_phan_name,
                    COALESCE(NULLIF(ct.ten_chi_tiet, ''), '(Trống)') AS chi_tiet_name,
                    COUNT(*) AS qty
                FROM public.qc_defect d
                JOIN public.qc_error_log_sp sp ON sp.id = d.error_log_sp_id
                JOIN public.prod_plan p ON p.id = sp.plan_id
                LEFT JOIN public.quality_employees qe ON qe.ma_nv = sp.ma_nv
                {bo_phan_join}
                LEFT JOIN public.dm_loai_hang lh ON lh.ten_loai = p.loai_hang
                LEFT JOIN public.dm_loai_hang_target tgt ON tgt.id_type = lh.id_type
                LEFT JOIN public.dm_bo_phan bp ON bp.id = d.bo_phan_id
                LEFT JOIN public.dm_chi_tiet ct ON ct.id = d.chi_tiet_id
                WHERE {pos_where_sql}
                GROUP BY bo_phan_name, chi_tiet_name
                """,
                tuple(pos_params),
            )
            pos_rows = cur.fetchall()

            # Defect aggregates for C/D
            defect_clauses = ["sp.date BETWEEN %s AND %s"]
            defect_params: List = [date_from, date_to]
            if don_vi:
                defect_clauses.append("p.don_vi = %s")
                defect_params.append(don_vi)
            if bo_phan:
                defect_clauses.append(f"{resolved_bo_phan_expr} = %s")
                defect_params.append(bo_phan)
            if station_for_calc:
                defect_clauses.append("COALESCE(sp.station, '') = %s")
                defect_params.append(station_for_calc)
            if ma_hang:
                defect_clauses.append("COALESCE(p.ma_hang, '') = %s")
                defect_params.append(ma_hang)
            if type_name:
                defect_clauses.append("COALESCE(tgt.type, '') = %s")
                defect_params.append(type_name)
            defect_where_sql = " AND ".join(defect_clauses)

            cur.execute(f"""
                SELECT
                    COALESCE(NULLIF(bp.ten_bo_phan, ''), '(Trống)') AS bo_phan_name,
                    COALESCE(NULLIF(ct.ten_chi_tiet, ''), '(Trống)') AS chi_tiet_name,
                    COALESCE(NULLIF(ml.ten_ma, ''), d.ma_loi_id::text, '(Trống)') AS ma_loi_name,
                    COUNT(*) AS qty
                FROM public.qc_defect d
                JOIN public.qc_error_log_sp sp ON sp.id = d.error_log_sp_id
                JOIN public.prod_plan p ON p.id = sp.plan_id
                LEFT JOIN public.quality_employees qe ON qe.ma_nv = sp.ma_nv
                {bo_phan_join}
                LEFT JOIN public.dm_loai_hang lh ON lh.ten_loai = p.loai_hang
                LEFT JOIN public.dm_loai_hang_target tgt ON tgt.id_type = lh.id_type
                LEFT JOIN public.dm_bo_phan bp ON bp.id = d.bo_phan_id
                LEFT JOIN public.dm_chi_tiet ct ON ct.id = d.chi_tiet_id
                LEFT JOIN public.dm_ma_loi ml ON ml.id = d.ma_loi_id
                WHERE {defect_where_sql}
                GROUP BY bo_phan_name, chi_tiet_name, ma_loi_name
            """, tuple(defect_params))
            defect_agg_rows = cur.fetchall()

            # Rework done / recycle stats based on product-level defects
            cur.execute(f"""
                WITH per_sp AS (
                    SELECT
                        d.error_log_sp_id,
                        d.sp_index,
                        BOOL_AND(COALESCE(d.rework_done, FALSE)) AS done
                    FROM public.qc_defect d
                    WHERE d.error_log_sp_id IS NOT NULL
                      AND d.sp_index IS NOT NULL
                    GROUP BY d.error_log_sp_id, d.sp_index
                )
                SELECT
                    COUNT(*) AS defect_sp_total,
                    COUNT(*) FILTER (WHERE per_sp.done) AS rework_done_total
                FROM per_sp
                JOIN public.qc_error_log_sp sp ON sp.id = per_sp.error_log_sp_id
                JOIN public.prod_plan p ON p.id = sp.plan_id
                LEFT JOIN public.quality_employees qe ON qe.ma_nv = sp.ma_nv
                {bo_phan_join}
                LEFT JOIN public.dm_loai_hang lh ON lh.ten_loai = p.loai_hang
                LEFT JOIN public.dm_loai_hang_target tgt ON tgt.id_type = lh.id_type
                WHERE {defect_where_sql}
            """, tuple(defect_params))
            rework_stats_row = cur.fetchone() or {}

    def _to_float(v) -> float:
        try:
            return float(v or 0)
        except Exception:
            return 0.0

    daily = []
    for r in daily_rows:
        daily.append({
            "date": str(r["date"]) if r.get("date") else None,
            "inspected_total": int(r.get("inspected_total") or 0),
            "defect_total": int(r.get("defect_total") or 0),
            "defect_rate": _to_float(r.get("defect_rate")),
            "target_rate": _to_float(r.get("target_rate")),
        })

    station_summary = []
    for r in station_rows:
        station_summary.append({
            "station": r.get("station") or "N/A",
            "inspected_total": int(r.get("inspected_total") or 0),
            "defect_total": int(r.get("defect_total") or 0),
            "defect_rate": _to_float(r.get("defect_rate")),
            "target_rate": _to_float(r.get("target_rate")),
        })

    pos_bp: Dict[str, Dict[str, int]] = {}
    pos_bp_totals: Dict[str, int] = {}
    for r in (pos_rows or []):
        bp_name = r.get("bo_phan_name") or "(Trống)"
        ct_name = r.get("chi_tiet_name") or "(Trống)"
        qty = int(r.get("qty") or 0)
        if qty <= 0:
            continue
        if bp_name not in pos_bp:
            pos_bp[bp_name] = {}
        pos_bp[bp_name][ct_name] = pos_bp[bp_name].get(ct_name, 0) + qty
        pos_bp_totals[bp_name] = pos_bp_totals.get(bp_name, 0) + qty

    pos_items = []
    for bp_name, total in sorted(pos_bp_totals.items(), key=lambda x: (-x[1], x[0])):
        details_map = pos_bp.get(bp_name, {})
        details = []
        for ct_name, cnt in sorted(details_map.items(), key=lambda x: (-x[1], x[0])):
            details.append({"chi_tiet": ct_name, "count": int(cnt)})
        pos_items.append({"bo_phan": bp_name, "count": int(total), "details": details})
    pos_stats = {"date": focus_date, "items": pos_items}

    by_bucket = []
    for b in bucket_order:
        r = bucket_rows.get(b) or {}
        by_bucket.append({
            "time_bucket": b,
            "inspected_total": int(r.get("inspected_total") or 0),
            "defect_total": int(r.get("defect_total") or 0),
            "defect_rate": _to_float(r.get("defect_rate")),
            "target_rate": _to_float(r.get("target_rate")),
        })

    bo_phan_counts: Dict[str, int] = {}
    chi_tiet_by_bo_phan: Dict[str, Dict[str, int]] = {}
    ma_loi_counts_all: Dict[str, int] = {}
    all_chi_tiet: Dict[str, int] = {}

    for r in defect_agg_rows:
        bp_name = r.get("bo_phan_name") or "(Trống)"
        ct_name = r.get("chi_tiet_name") or "(Trống)"
        ml_name = r.get("ma_loi_name") or "(Trống)"
        qty = int(r.get("qty") or 0)

        bo_phan_counts[bp_name] = bo_phan_counts.get(bp_name, 0) + qty
        chi_tiet_by_bo_phan.setdefault(bp_name, {})
        chi_tiet_by_bo_phan[bp_name][ct_name] = chi_tiet_by_bo_phan[bp_name].get(ct_name, 0) + qty
        ma_loi_counts_all[ml_name] = ma_loi_counts_all.get(ml_name, 0) + qty
        all_chi_tiet[ct_name] = all_chi_tiet.get(ct_name, 0) + qty

    bo_phan_total = sum(bo_phan_counts.values())
    top_bo_phan_items = sorted(bo_phan_counts.items(), key=lambda x: (-x[1], x[0]))[:5]
    top_bo_phan = [
        {
            "name": name,
            "count": cnt,
            "percent_total": round((cnt * 100.0 / bo_phan_total), 4) if bo_phan_total > 0 else 0.0,
        }
        for name, cnt in top_bo_phan_items
    ]

    selected_detail_bo_phan = detail_bo_phan or (top_bo_phan[0]["name"] if top_bo_phan else "")
    selected_chi_tiet_map = chi_tiet_by_bo_phan.get(selected_detail_bo_phan, {})
    chi_tiet_total = sum(selected_chi_tiet_map.values())
    top_chi_tiet_items = sorted(selected_chi_tiet_map.items(), key=lambda x: (-x[1], x[0]))[:5]
    top_chi_tiet = [
        {
            "name": name,
            "count": cnt,
            "percent_total": round((cnt * 100.0 / chi_tiet_total), 4) if chi_tiet_total > 0 else 0.0,
        }
        for name, cnt in top_chi_tiet_items
    ]

    ma_loi_filtered_counts: Dict[str, int] = {}
    for r in defect_agg_rows:
        bp_name = r.get("bo_phan_name") or "(Trống)"
        ct_name = r.get("chi_tiet_name") or "(Trống)"
        ml_name = r.get("ma_loi_name") or "(Trống)"
        qty = int(r.get("qty") or 0)
        if ma_loi_bo_phan and bp_name != ma_loi_bo_phan:
            continue
        if ma_loi_chi_tiet and ct_name != ma_loi_chi_tiet:
            continue
        ma_loi_filtered_counts[ml_name] = ma_loi_filtered_counts.get(ml_name, 0) + qty

    ma_loi_total = sum(ma_loi_filtered_counts.values())
    # Return a longer list for table display (still sorted by frequency).
    top_ma_loi_items = sorted(ma_loi_filtered_counts.items(), key=lambda x: (-x[1], x[0]))[:20]
    top_ma_loi = [
        {
            "name": name,
            "count": cnt,
            "percent_total": round((cnt * 100.0 / ma_loi_total), 4) if ma_loi_total > 0 else 0.0,
        }
        for name, cnt in top_ma_loi_items
    ]

    defect_filter_bo_phan = sorted(bo_phan_counts.keys())
    if ma_loi_bo_phan:
        defect_filter_chi_tiet = sorted((chi_tiet_by_bo_phan.get(ma_loi_bo_phan) or {}).keys())
    else:
        defect_filter_chi_tiet = sorted(all_chi_tiet.keys())

    totals = {
        "inspected_total": sum(x["inspected_total"] for x in daily),
        "defect_total": sum(x["defect_total"] for x in daily),
    }
    totals["defect_rate"] = round((totals["defect_total"] * 100.0 / totals["inspected_total"]), 4) if totals["inspected_total"] > 0 else 0.0
    totals["defect_sp_total"] = int(rework_stats_row.get("defect_sp_total") or 0)
    totals["rework_done_total"] = int(rework_stats_row.get("rework_done_total") or 0)
    totals["recycle_rate"] = round((totals["rework_done_total"] * 100.0 / totals["defect_sp_total"]), 4) if totals["defect_sp_total"] > 0 else 0.0

    return {
        "range": {"date_from": date_from, "date_to": date_to, "focus_date": focus_date},
        "active_station": station_for_calc,
        "totals": totals,
        "daily": daily,
        "by_bucket": by_bucket,
        "pos_stats": pos_stats,
        "station_summary": station_summary,
        "top_bo_phan": {
            "total": bo_phan_total,
            "items": top_bo_phan,
        },
        "top_chi_tiet": {
            "selected_bo_phan": selected_detail_bo_phan,
            "total": chi_tiet_total,
            "items": top_chi_tiet,
        },
        "top_ma_loi": {
            "filter_bo_phan": ma_loi_bo_phan or "",
            "filter_chi_tiet": ma_loi_chi_tiet or "",
            "total": ma_loi_total,
            "items": top_ma_loi,
        },
        "defect_filters": {
            "bo_phan": defect_filter_bo_phan,
            "chi_tiet": defect_filter_chi_tiet,
        },
    }

class QCEmployee(BaseModel):
    ma_nv: str
    ho_ten: str
    chuc_vu: str = 'QC'
    don_vi: str
    bo_phan: str
    station: Optional[List[str]] = None

class QcErrorDpsPayload(BaseModel):
    plan_id: int
    date: str
    time_bucket: str
    station: Optional[str] = None
    bo_phan: Optional[str] = None
    loai_loi: str
    ma_loi: Optional[str] = None
    vi_tri: Optional[str] = None
    nn_cong_nhan: bool = False
    nn_may_moc: bool = False
    nn_phuong_phap: bool = False
    nn_nguyen_phu_lieu: bool = False
    nn_moi_truong: bool = False
    mo_ta: Optional[str] = None
    giai_phap: Optional[str] = None
    tram_ap_dung: Optional[str] = None
    tien_do: Optional[str] = None
    ngay_hoan_thanh: Optional[str] = None
    ket_luan: Optional[str] = None
    ghi_chu: Optional[str] = None


COMBO_SHARED_CAP_TYPES = {
    "Lỗi đại trà",
    "Lỗi vượt mục tiêu",
    "Tỉ lệ theo dạng lỗi",
    "CAP theo combo lỗi",
}


def normalize_qc_cap_loai_loi(loai_loi: Optional[str], ma_loi: Optional[str], vi_tri: Optional[str]) -> str:
    normalized = (loai_loi or "").strip()
    if normalized in COMBO_SHARED_CAP_TYPES and ((ma_loi or "").strip() or (vi_tri or "").strip()):
        return "CAP theo combo lỗi"
    return normalized

@app.get("/api/qc/employees")
def get_qc_employees():
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT ma_nv, ho_ten, chuc_vu, don_vi, bo_phan, station FROM public.quality_employees WHERE chuc_vu = 'QC' ORDER BY ma_nv")
            rows = cur.fetchall()
            return {"status": "ok", "rows": rows}

@app.post("/api/qc/employees")
def upsert_qc_employee(payload: QCEmployee):
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT ma_nv FROM public.quality_employees WHERE ma_nv = %s", (payload.ma_nv,))
            existing = cur.fetchone()
            if existing:
                cur.execute("""
                    UPDATE public.quality_employees
                    SET ho_ten = %s, chuc_vu = %s, don_vi = %s, bo_phan = %s, station = %s
                    WHERE ma_nv = %s
                """, (payload.ho_ten, payload.chuc_vu, payload.don_vi, payload.bo_phan, json.dumps(payload.station or []), payload.ma_nv))
            else:
                cur.execute("""
                    INSERT INTO public.quality_employees (ma_nv, ho_ten, chuc_vu, don_vi, bo_phan, station)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (payload.ma_nv, payload.ho_ten, payload.chuc_vu, payload.don_vi, payload.bo_phan, json.dumps(payload.station or [])))
        conn.commit()
    return {"status": "ok"}

@app.post("/api/qc/cap/action")
def upsert_qc_cap_action(
    payload: QcErrorDpsPayload,
    _user: Dict = Depends(require_qaqt_api_user),
):
    loai_loi = normalize_qc_cap_loai_loi(payload.loai_loi, payload.ma_loi, payload.vi_tri)
    if not loai_loi:
        raise HTTPException(status_code=400, detail="Loai loi khong duoc de trong")
    time_bucket = (payload.time_bucket or "").strip()
    if not time_bucket:
        raise HTTPException(status_code=400, detail="Time bucket khong duoc de trong")
    ma_loi = (payload.ma_loi or "").strip() or None
    vi_tri = (payload.vi_tri or "").strip() or None
    station = (payload.station or "").strip() or None
    bo_phan = (payload.bo_phan or "").strip() or None

    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT id
                FROM public.qc_error_dps
                WHERE plan_id = %s
                  AND date = %s
                  AND time_bucket = %s
                  AND COALESCE(station, '') = COALESCE(%s, '')
                  AND COALESCE(bo_phan, '') = COALESCE(%s, '')
                  AND loai_loi = %s
                  AND COALESCE(ma_loi, '') = COALESCE(%s, '')
                  AND COALESCE(vi_tri, '') = COALESCE(%s, '')
                """,
                (payload.plan_id, payload.date, time_bucket, station, bo_phan, loai_loi, ma_loi, vi_tri),
            )
            existing = cur.fetchone()

            if existing:
                cur.execute(
                    """
                    UPDATE public.qc_error_dps
                    SET nn_cong_nhan = %s,
                        nn_may_moc = %s,
                        nn_phuong_phap = %s,
                        nn_nguyen_phu_lieu = %s,
                        nn_moi_truong = %s,
                        mo_ta = %s,
                        giai_phap = %s,
                        tram_ap_dung = %s,
                        tien_do = %s,
                        ngay_hoan_thanh = %s,
                        ket_luan = %s,
                        ghi_chu = %s,
                        updated_at = NOW()
                    WHERE id = %s
                    RETURNING id
                    """,
                    (
                        payload.nn_cong_nhan,
                        payload.nn_may_moc,
                        payload.nn_phuong_phap,
                        payload.nn_nguyen_phu_lieu,
                        payload.nn_moi_truong,
                        payload.mo_ta,
                        payload.giai_phap,
                        payload.tram_ap_dung,
                        payload.tien_do,
                        payload.ngay_hoan_thanh,
                        payload.ket_luan,
                        payload.ghi_chu,
                        existing["id"],
                    ),
                )
                action_id = cur.fetchone()["id"]
            else:
                cur.execute(
                    """
                    INSERT INTO public.qc_error_dps (
                        plan_id, date, time_bucket, station, bo_phan, loai_loi, ma_loi, vi_tri,
                        nn_cong_nhan, nn_may_moc, nn_phuong_phap, nn_nguyen_phu_lieu, nn_moi_truong,
                        mo_ta, giai_phap, tram_ap_dung, tien_do, ngay_hoan_thanh, ket_luan, ghi_chu
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                    """,
                    (
                        payload.plan_id,
                        payload.date,
                        time_bucket,
                        station,
                        bo_phan,
                        loai_loi,
                        ma_loi,
                        vi_tri,
                        payload.nn_cong_nhan,
                        payload.nn_may_moc,
                        payload.nn_phuong_phap,
                        payload.nn_nguyen_phu_lieu,
                        payload.nn_moi_truong,
                        payload.mo_ta,
                        payload.giai_phap,
                        payload.tram_ap_dung,
                        payload.tien_do,
                        payload.ngay_hoan_thanh,
                        payload.ket_luan,
                        payload.ghi_chu,
                    ),
                )
                action_id = cur.fetchone()["id"]
        conn.commit()

    return {"status": "ok", "id": action_id}


@app.get("/api/qc/cap/action")
def get_qc_cap_action(
    request: Request,
    plan_id: int = Query(...),
    date: str = Query(...),
    time_bucket: str = Query(...),
    station: Optional[str] = Query(None),
    bo_phan: Optional[str] = Query(None),
    loai_loi: str = Query(...),
    ma_loi: Optional[str] = Query(None),
    vi_tri: Optional[str] = Query(None),
):
    effective_don_vi, user = resolve_qc_don_vi_scope(request, None)
    loai_loi = normalize_qc_cap_loai_loi(loai_loi, ma_loi, vi_tri)
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            if is_qc_don_vi_scoped_role(user):
                cur.execute(
                    "SELECT 1 FROM public.prod_plan WHERE id = %s AND don_vi = %s",
                    (plan_id, effective_don_vi),
                )
                if not cur.fetchone():
                    raise HTTPException(status_code=403, detail="Không có quyền xem dữ liệu ngoài đơn vị được gán.")
            cur.execute(
                """
                SELECT *
                FROM public.qc_error_dps
                WHERE plan_id = %s
                  AND date = %s
                  AND time_bucket = %s
                  AND COALESCE(station, '') = COALESCE(%s, '')
                  AND COALESCE(bo_phan, '') = COALESCE(%s, '')
                  AND loai_loi = %s
                  AND COALESCE(ma_loi, '') = COALESCE(%s, '')
                  AND COALESCE(vi_tri, '') = COALESCE(%s, '')
                LIMIT 1
                """,
                (plan_id, date, time_bucket, station, bo_phan, loai_loi, ma_loi, vi_tri),
            )
            row = cur.fetchone()
    if not row:
        return {"status": "not_found"}
    row["date"] = str(row["date"]) if row.get("date") else None
    row["ngay_hoan_thanh"] = str(row["ngay_hoan_thanh"]) if row.get("ngay_hoan_thanh") else None
    return {"status": "ok", "row": row}

@app.delete("/api/qc/employees/{ma_nv}")
def delete_qc_employee(ma_nv: str):
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM public.quality_employees WHERE ma_nv = %s AND chuc_vu = 'QC'", (ma_nv,))
        conn.commit()
    return {"status": "ok"}

@app.get("/api/qc/cap")
def api_qc_cap(
    request: Request,
    date: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    don_vi: Optional[str] = Query(None),
    bo_phan: Optional[str] = Query(None),
    station: Optional[str] = Query(None),
):
    """Tổng hợp dữ liệu CAP theo ngày và mốc thời gian."""
    don_vi, _user = resolve_qc_don_vi_scope(request, don_vi)
    if date:
        date_from = date_to = date
    if not date_from or not date_to:
        today = datetime.now().strftime("%Y-%m-%d")
        date_from = date_from or today
        date_to = date_to or today

    def _bucket_case(time_expr: str) -> str:
        return f"""
            CASE
                WHEN {time_expr} < '09:30' THEN '7H30 - 9H30'
                WHEN {time_expr} >= '09:30' AND {time_expr} < '11:30' THEN '9H30 - 11H30'
                WHEN {time_expr} >= '11:30' AND {time_expr} < '14:30' THEN '12H30 - 14H30'
                WHEN {time_expr} >= '14:30' AND {time_expr} < '16:30' THEN '14H30 - 16H30'
                ELSE 'Sau 16H30'
            END
        """

    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            # Detect prod_plan.bo_phan type to avoid jsonb/text mismatch across deployments
            cur.execute("""
                SELECT data_type
                FROM information_schema.columns
                WHERE table_schema = 'public'
                  AND table_name = 'prod_plan'
                  AND column_name = 'bo_phan'
                LIMIT 1
            """)
            bo_phan_type = (cur.fetchone() or {}).get("data_type")
            use_jsonb_bo_phan = bo_phan_type == "jsonb"

            bo_phan_expr = "bp.bo_phan_text" if use_jsonb_bo_phan else "p.bo_phan"
            bo_phan_join = (
                """
                LEFT JOIN LATERAL (
                    SELECT string_agg(value, ', ') AS bo_phan_text
                    FROM jsonb_array_elements_text(COALESCE(p.bo_phan, '[]'::jsonb)) AS t(value)
                ) bp ON true
                """
                if use_jsonb_bo_phan
                else ""
            )
            bo_phan_order = bo_phan_expr

            # Output tổng hợp theo mốc thời gian
            output_params = [date_from, date_to]
            output_where = []
            if station:
                output_where.append("o.station = %s")
                output_params.append(station)

            cur.execute(f"""
                WITH output_logs AS (
                    SELECT
                        o.plan_id,
                        o.date,
                        o.station,
                        COALESCE(qe.bo_phan, '') AS qc_bo_phan,
                        {_bucket_case("timezone('Asia/Bangkok', o.created_at)::time")} AS time_bucket,
                        SUM(o.delta) AS output_total,
                        SUM(CASE WHEN o.status = 'Failed' THEN o.delta ELSE 0 END) AS defect_total
                    FROM public.qc_output_sp_log o
                    LEFT JOIN public.quality_employees qe ON qe.ma_nv = o.ma_nv
                    WHERE o.date BETWEEN %s AND %s
                    {("AND " + " AND ".join(output_where)) if output_where else ""}
                    GROUP BY o.plan_id, o.date, o.station, qc_bo_phan, time_bucket
                )
                SELECT
                    o.*,
                    p.don_vi,
                    COALESCE(NULLIF(o.qc_bo_phan, ''), COALESCE({bo_phan_expr}, ''), '') AS bo_phan,
                    p.ke_hoach,
                    p.ma_hang,
                    p.ngay_rc,
                    p.san_luong,
                    p.loai_hang
                FROM output_logs o
                JOIN public.prod_plan p ON p.id = o.plan_id
                {bo_phan_join}
                WHERE 1=1
                {("AND p.don_vi = %s") if don_vi else ""}
                {("AND COALESCE(NULLIF(o.qc_bo_phan, ''), COALESCE(" + bo_phan_expr + ", ''), '') = %s") if bo_phan else ""}
                ORDER BY o.date, o.time_bucket, p.don_vi, COALESCE(NULLIF(o.qc_bo_phan, ''), {bo_phan_order}, ''), o.station
            """, tuple(output_params + ([don_vi] if don_vi else []) + ([bo_phan] if bo_phan else [])))
            output_rows = cur.fetchall()

            # Target theo loại hàng
            cur.execute("""
                SELECT lh.ten_loai, tgt.target_percent
                FROM public.dm_loai_hang lh
                LEFT JOIN public.dm_loai_hang_target tgt ON tgt.id_type = lh.id_type
            """)
            target_map = {r["ten_loai"]: r["target_percent"] for r in cur.fetchall() if r.get("ten_loai")}

            # Lỗi nghiêm trọng
            cur.execute(f"""
                SELECT
                    sp.plan_id,
                    sp.date,
                    sp.station,
                    COALESCE(qe.bo_phan, '') AS qc_bo_phan,
                    {_bucket_case("timezone('Asia/Bangkok', d.created_at)::time")} AS time_bucket,
                    COALESCE(ml.ten_ma, d.ma_loi_id::text) AS ma_loi,
                    COALESCE(bp.ten_bo_phan, '') AS bo_phan,
                    COALESCE(ct.ten_chi_tiet, '') AS chi_tiet,
                    COUNT(*) AS qty
                FROM public.qc_defect d
                JOIN public.qc_error_log_sp sp ON sp.id = d.error_log_sp_id
                LEFT JOIN public.quality_employees qe ON qe.ma_nv = sp.ma_nv
                LEFT JOIN public.dm_ma_loi ml ON ml.id = d.ma_loi_id
                LEFT JOIN public.dm_bo_phan bp ON bp.id = d.bo_phan_id
                LEFT JOIN public.dm_chi_tiet ct ON ct.id = d.chi_tiet_id
                WHERE sp.date BETWEEN %s AND %s
                  AND d.muc_do ILIKE '%%nghiêm%%'
                GROUP BY
                    sp.plan_id,
                    sp.date,
                    sp.station,
                    qc_bo_phan,
                    time_bucket,
                    COALESCE(ml.ten_ma, d.ma_loi_id::text),
                    COALESCE(bp.ten_bo_phan, ''),
                    COALESCE(ct.ten_chi_tiet, '')
                ORDER BY sp.date, time_bucket, sp.station
            """, (date_from, date_to))
            serious_rows = cur.fetchall()

            # Lỗi đại trà (mới): combo (Vị trí [Bộ phận - Chi tiết] + Mã lỗi) tính theo tỉ lệ trên tổng kiểm trong từng mốc thời gian.
            # Đếm theo số sản phẩm (DISTINCT error_log_sp_id + sp_index) mắc combo đó trong bucket.
            cur.execute(f"""
                SELECT
                    t.plan_id,
                    t.date,
                    t.station,
                    t.qc_bo_phan,
                    t.time_bucket,
                    t.ma_loi,
                    t.bo_phan,
                    t.chi_tiet,
                    COUNT(*) AS qty
                FROM (
                    SELECT DISTINCT
                        sp.plan_id,
                        sp.date,
                        sp.station,
                        COALESCE(qe.bo_phan, '') AS qc_bo_phan,
                        {_bucket_case("timezone('Asia/Bangkok', d.created_at)::time")} AS time_bucket,
                        COALESCE(ml.ten_ma, d.ma_loi_id::text) AS ma_loi,
                        COALESCE(bp.ten_bo_phan, '') AS bo_phan,
                        COALESCE(ct.ten_chi_tiet, '') AS chi_tiet,
                        d.error_log_sp_id,
                        d.sp_index
                    FROM public.qc_defect d
                    JOIN public.qc_error_log_sp sp ON sp.id = d.error_log_sp_id
                    LEFT JOIN public.quality_employees qe ON qe.ma_nv = sp.ma_nv
                    LEFT JOIN public.dm_ma_loi ml ON ml.id = d.ma_loi_id
                    LEFT JOIN public.dm_bo_phan bp ON bp.id = d.bo_phan_id
                    LEFT JOIN public.dm_chi_tiet ct ON ct.id = d.chi_tiet_id
                    WHERE sp.date BETWEEN %s AND %s
                      AND d.error_log_sp_id IS NOT NULL
                      AND d.sp_index IS NOT NULL
                ) t
                GROUP BY
                    t.plan_id,
                    t.date,
                    t.station,
                    t.qc_bo_phan,
                    t.time_bucket,
                    t.ma_loi,
                    t.bo_phan,
                    t.chi_tiet
                ORDER BY t.date, t.time_bucket, t.station
            """, (date_from, date_to))
            mass_combo_rows = cur.fetchall()

            # Ảnh lỗi theo plan/date/bucket/ma_loi/vi_tri
            cur.execute(f"""
                SELECT
                    sp.plan_id,
                    sp.date,
                    sp.station,
                    COALESCE(qe.bo_phan, '') AS qc_bo_phan,
                    {_bucket_case("timezone('Asia/Bangkok', d.created_at)::time")} AS time_bucket,
                    COALESCE(ml.ten_ma, d.ma_loi_id::text) AS ma_loi,
                    COALESCE(bp.ten_bo_phan, '') AS bo_phan,
                    COALESCE(ct.ten_chi_tiet, '') AS chi_tiet,
                    ARRAY_REMOVE(ARRAY_AGG(DISTINCT d.image_path), NULL) AS images
                FROM public.qc_defect d
                JOIN public.qc_error_log_sp sp ON sp.id = d.error_log_sp_id
                LEFT JOIN public.quality_employees qe ON qe.ma_nv = sp.ma_nv
                LEFT JOIN public.dm_ma_loi ml ON ml.id = d.ma_loi_id
                LEFT JOIN public.dm_bo_phan bp ON bp.id = d.bo_phan_id
                LEFT JOIN public.dm_chi_tiet ct ON ct.id = d.chi_tiet_id
                WHERE sp.date BETWEEN %s AND %s
                  AND d.image_path IS NOT NULL
                GROUP BY
                    sp.plan_id,
                    sp.date,
                    sp.station,
                    qc_bo_phan,
                    time_bucket,
                    COALESCE(ml.ten_ma, d.ma_loi_id::text),
                    COALESCE(bp.ten_bo_phan, ''),
                    COALESCE(ct.ten_chi_tiet, '')
            """, (date_from, date_to))
            image_rows = cur.fetchall()

            # Lỗi nguy cơ hàng loạt
            cur.execute(f"""
                SELECT
                    plan_id,
                    date,
                    qdm.station,
                    COALESCE(qe.bo_phan, '') AS qc_bo_phan,
                    {_bucket_case("qdm.time")} AS time_bucket,
                    COALESCE(ma_loi, '') AS ma_loi,
                    COALESCE(qdm.bo_phan, '') AS bo_phan,
                    COALESCE(qdm.chi_tiet, '') AS chi_tiet,
                    COUNT(*) AS occurrences
                FROM public.qc_defect_multi qdm
                LEFT JOIN public.quality_employees qe ON qe.ma_nv = qdm.ma_nv
                WHERE date BETWEEN %s AND %s
                GROUP BY plan_id, date, qdm.station, qc_bo_phan, time_bucket, ma_loi, qdm.bo_phan, qdm.chi_tiet
                ORDER BY date, time_bucket, qdm.station
            """, (date_from, date_to))
            multi_rows = cur.fetchall()

            # Rework verified count (#SP đã xác nhận)
            cur.execute(f"""
                SELECT
                    sp.plan_id,
                    sp.date,
                    sp.station,
                    COALESCE(qe.bo_phan, '') AS qc_bo_phan,
                    {_bucket_case("timezone('Asia/Bangkok', per_sp.first_created_at)::time")} AS time_bucket,
                    COUNT(*) FILTER (WHERE per_sp.done) AS rework_done_count
                FROM public.qc_error_log_sp sp
                LEFT JOIN public.quality_employees qe ON qe.ma_nv = sp.ma_nv
                JOIN (
                    SELECT
                        d.error_log_sp_id,
                        d.sp_index,
                        MIN(d.created_at) AS first_created_at,
                        BOOL_AND(COALESCE(d.rework_done, FALSE)) AS done
                    FROM public.qc_defect d
                    GROUP BY d.error_log_sp_id, d.sp_index
                ) per_sp ON per_sp.error_log_sp_id = sp.id
                WHERE sp.date BETWEEN %s AND %s
                GROUP BY sp.plan_id, sp.date, sp.station, qc_bo_phan, time_bucket
            """, (date_from, date_to))
            rework_rows = cur.fetchall()

    def _vi_tri(bo_phan: str, chi_tiet: str) -> str:
        parts = [p for p in [bo_phan, chi_tiet] if p]
        return " - ".join(parts) if parts else "--"

    def _cap_key(row):
        return (
            row["plan_id"],
            str(row["date"]),
            row["time_bucket"],
            row.get("station") or "",
            row.get("qc_bo_phan") or "",
        )

    serious_map = {}
    for r in serious_rows:
        key = _cap_key(r)
        serious_map.setdefault(key, []).append({
            "ma_loi": r["ma_loi"] or "--",
            "vi_tri": _vi_tri(r.get("bo_phan"), r.get("chi_tiet")),
            "qty": int(r["qty"] or 0)
        })

    image_map = {}
    for r in image_rows:
        key = (
            r["plan_id"],
            str(r["date"]),
            r["time_bucket"],
            r.get("station") or "",
            r.get("qc_bo_phan") or "",
            r["ma_loi"] or "--",
            _vi_tri(r.get("bo_phan"), r.get("chi_tiet")),
        )
        images = [f"/api/images/{p}" for p in (r.get("images") or []) if p]
        image_map[key] = images

    multi_map = {}
    for r in multi_rows:
        key = _cap_key(r)
        multi_map.setdefault(key, []).append({
            "ma_loi": r["ma_loi"] or "--",
            "vi_tri": _vi_tri(r.get("bo_phan"), r.get("chi_tiet")),
            "consecutive": 3,
            "occurrences": int(r["occurrences"] or 0)
        })

    mass_combo_map = {}
    for r in mass_combo_rows:
        key = _cap_key(r)
        mass_combo_map.setdefault(key, []).append({
            "ma_loi": r.get("ma_loi") or "--",
            "vi_tri": _vi_tri(r.get("bo_phan"), r.get("chi_tiet")),
            "qty": int(r.get("qty") or 0),
        })

    rework_map = {}
    for r in rework_rows:
        key = _cap_key(r)
        rework_map[key] = int(r.get("rework_done_count") or 0)

    # Group output rows by date/time bucket
    dates = {}
    for r in output_rows:
        date_key = str(r["date"])
        bucket = r["time_bucket"]
        plan_id = r["plan_id"]
        key = (plan_id, date_key, bucket, r.get("station") or "", r.get("qc_bo_phan") or "")

        output_total = int(r["output_total"] or 0)
        defect_total = int(r["defect_total"] or 0)
        rate = (defect_total / output_total * 100) if output_total > 0 else 0
        target_percent = target_map.get(r.get("loai_hang"))
        target_percent_float = float(target_percent) if target_percent is not None else None
        bucket_over_target = (
            output_total > 0
            and target_percent_float is not None
            and rate > target_percent_float
        )

        combo_items = list(mass_combo_map.get(key) or [])
        mass_list = []
        over_target_list = []
        for item in combo_items:
            qty = int(item.get("qty") or 0)
            if qty <= 0:
                continue
            combo_rate = (qty * 100.0 / output_total) if output_total > 0 else 0.0
            combo_entry = {
                "ma_loi": item.get("ma_loi") or "--",
                "vi_tri": item.get("vi_tri") or "--",
                "qty": qty,
                "rate_percent": combo_rate,
            }
            if output_total <= 0:
                continue
            if (qty / output_total) >= 0.15:
                mass_list.append(dict(combo_entry))
            if target_percent_float is not None and combo_rate > target_percent_float:
                over_target_list.append(dict(combo_entry))
        mass_list.sort(key=lambda x: (-int(x.get("qty") or 0), str(x.get("ma_loi") or ""), str(x.get("vi_tri") or "")))

        combo_rate_list = []
        if defect_total > 0:
            for item in combo_items:
                qty = int(item.get("qty") or 0)
                if qty <= 0:
                    continue
                combo_rate_list.append({
                    "ma_loi": item.get("ma_loi") or "--",
                    "vi_tri": item.get("vi_tri") or "--",
                    "qty": qty,
                    "total_qty": defect_total,
                    "ratio_percent": (qty * 100.0 / defect_total),
                    "product_rate_percent": rate,
                    "needs_cap": bucket_over_target,
                    "cap_status_text": "Cần làm CAP" if bucket_over_target else "Không cần làm CAP",
                })
        combo_rate_list.sort(key=lambda x: (-float(x.get("ratio_percent") or 0), -int(x.get("qty") or 0), str(x.get("ma_loi") or ""), str(x.get("vi_tri") or "")))
        combo_rate_top3 = combo_rate_list[:3]
        over_target_list.sort(key=lambda x: (-float(x.get("rate_percent") or 0), -int(x.get("qty") or 0), str(x.get("ma_loi") or ""), str(x.get("vi_tri") or "")))

        row = {
            "plan_id": plan_id,
            "date": date_key,
            "time_bucket": bucket,
            "don_vi": r.get("don_vi") or "--",
            "bo_phan": r.get("bo_phan") or "--",
            "qc_bo_phan": r.get("qc_bo_phan") or "",
            "station": r.get("station") or "",
            "ke_hoach": r.get("ke_hoach") or "--",
            "ke_hoach_hien_thi": build_prod_plan_display_name(r),
            "ma_hang": r.get("ma_hang") or "",
            "ngay_rc": str(r.get("ngay_rc")) if r.get("ngay_rc") else None,
            "san_luong": r.get("san_luong"),
            "output_total": output_total,
            "defect_total": defect_total,
            "rework_done_count": rework_map.get(key, 0),
            "rate_percent": rate,
            "target_percent": target_percent_float,
            "over_target": (len(over_target_list) > 0) or bucket_over_target,
            "over_target_list": over_target_list,
            "is_mass": len(mass_list) > 0,
            "mass_list": mass_list,
            "combo_rate_top3": combo_rate_top3,
            "serious_list": serious_map.get(key, []),
            "multi_list": multi_map.get(key, []),
            "has_serious": len(serious_map.get(key, [])) > 0,
            "has_multi": len(multi_map.get(key, [])) > 0,
        }

        dates.setdefault(date_key, {}).setdefault(bucket, []).append(row)

    # Sort buckets in desired order
    bucket_order = ["7H30 - 9H30", "9H30 - 11H30", "12H30 - 14H30", "14H30 - 16H30", "Sau 16H30"]
    result = []
    for date_key in sorted(dates.keys()):
        buckets = []
        for b in bucket_order:
            if b in dates[date_key]:
                buckets.append({"time_bucket": b, "rows": dates[date_key][b]})
        result.append({"date": date_key, "buckets": buckets})

    # Pre-create qc_error_dps records for rows that have errors
    precreate_keys = set()
    for date_key, buckets in dates.items():
        for bucket, rows in buckets.items():
            for r in rows:
                bo_phan_key = r.get("qc_bo_phan") or r.get("bo_phan") or ""
                for item in (r.get("serious_list") or []):
                    precreate_keys.add((
                        r["plan_id"],
                        date_key,
                        bucket,
                        r.get("station") or "",
                        bo_phan_key,
                        "Lỗi nghiêm trọng",
                        item.get("ma_loi") or None,
                        item.get("vi_tri") or None,
                    ))
                for item in (r.get("multi_list") or []):
                    precreate_keys.add((
                        r["plan_id"],
                        date_key,
                        bucket,
                        r.get("station") or "",
                        bo_phan_key,
                        "Lỗi nguy cơ hàng loạt",
                        item.get("ma_loi") or None,
                        item.get("vi_tri") or None,
                    ))
                for item in (r.get("over_target_list") or []):
                    precreate_keys.add((
                        r["plan_id"],
                        date_key,
                        bucket,
                        r.get("station") or "",
                        bo_phan_key,
                        "CAP theo combo lỗi",
                        item.get("ma_loi") or None,
                        item.get("vi_tri") or None,
                    ))
                for item in (r.get("mass_list") or []):
                    precreate_keys.add((
                        r["plan_id"],
                        date_key,
                        bucket,
                        r.get("station") or "",
                        bo_phan_key,
                        "CAP theo combo lỗi",
                        item.get("ma_loi") or None,
                        item.get("vi_tri") or None,
                    ))

    if precreate_keys:
        with get_db_connection() as conn:
            with conn.cursor() as cur:
                psycopg2.extras.execute_values(
                    cur,
                    """
                    INSERT INTO public.qc_error_dps
                        (plan_id, date, time_bucket, station, bo_phan, loai_loi, ma_loi, vi_tri)
                    VALUES %s
                    ON CONFLICT DO NOTHING
                    """,
                    list(precreate_keys),
                )
            conn.commit()

    # Mark which rows already have input data
    action_map = {}
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT
                    id,
                    plan_id,
                    date,
                    time_bucket,
                    station,
                    bo_phan,
                    loai_loi,
                    COALESCE(ma_loi, '') AS ma_loi,
                    COALESCE(vi_tri, '') AS vi_tri,
                    hdkp_pdf,
                    (
                        nn_cong_nhan OR nn_may_moc OR nn_phuong_phap OR nn_nguyen_phu_lieu OR nn_moi_truong
                        OR NULLIF(TRIM(COALESCE(mo_ta, '')), '') IS NOT NULL
                        OR NULLIF(TRIM(COALESCE(giai_phap, '')), '') IS NOT NULL
                        OR NULLIF(TRIM(COALESCE(tram_ap_dung, '')), '') IS NOT NULL
                        OR ngay_hoan_thanh IS NOT NULL
                        OR NULLIF(TRIM(COALESCE(ket_luan, '')), '') IS NOT NULL
                        OR NULLIF(TRIM(COALESCE(ghi_chu, '')), '') IS NOT NULL
                    ) AS has_input
                FROM public.qc_error_dps
                WHERE date BETWEEN %s AND %s
                """,
                (date_from, date_to),
            )
            for r in cur.fetchall():
                normalized_loai_loi = normalize_qc_cap_loai_loi(r["loai_loi"], r.get("ma_loi"), r.get("vi_tri"))
                key = (
                    r["plan_id"],
                    str(r["date"]),
                    r["time_bucket"],
                    r.get("station") or "",
                    r.get("bo_phan") or "",
                    normalized_loai_loi,
                    r.get("ma_loi") or "",
                    r.get("vi_tri") or "",
                )
                action_map[key] = {
                    "has_input": bool(r.get("has_input")),
                    "id": r.get("id"),
                    "hdkp_pdf": r.get("hdkp_pdf"),
                }

    # Attach action_done flags
    for date_key, buckets in dates.items():
        for bucket, rows in buckets.items():
            for row in rows:
                bo_phan_key = row.get("qc_bo_phan") or row.get("bo_phan") or ""
                for item in row.get("serious_list", []):
                    key = (
                        row["plan_id"],
                        date_key,
                        bucket,
                        row.get("station") or "",
                        bo_phan_key,
                        "Lỗi nghiêm trọng",
                        item.get("ma_loi") or "",
                        item.get("vi_tri") or "",
                    )
                    item["action_done"] = action_map.get(key, {}).get("has_input", False)
                    item["action_id"] = action_map.get(key, {}).get("id")
                    item["hdkp_pdf"] = action_map.get(key, {}).get("hdkp_pdf")
                    img_key = (
                        row["plan_id"],
                        date_key,
                        bucket,
                        row.get("station") or "",
                        bo_phan_key,
                        item.get("ma_loi") or "--",
                        item.get("vi_tri") or "--",
                    )
                    item["images"] = image_map.get(img_key, [])
                for item in row.get("multi_list", []):
                    key = (
                        row["plan_id"],
                        date_key,
                        bucket,
                        row.get("station") or "",
                        bo_phan_key,
                        "Lỗi nguy cơ hàng loạt",
                        item.get("ma_loi") or "",
                        item.get("vi_tri") or "",
                    )
                    item["action_done"] = action_map.get(key, {}).get("has_input", False)
                    item["action_id"] = action_map.get(key, {}).get("id")
                    item["hdkp_pdf"] = action_map.get(key, {}).get("hdkp_pdf")
                    img_key = (
                        row["plan_id"],
                        date_key,
                        bucket,
                        row.get("station") or "",
                        bo_phan_key,
                        item.get("ma_loi") or "--",
                        item.get("vi_tri") or "--",
                    )
                    item["images"] = image_map.get(img_key, [])
                for item in row.get("mass_list", []):
                    key = (
                        row["plan_id"],
                        date_key,
                        bucket,
                        row.get("station") or "",
                        bo_phan_key,
                        "CAP theo combo lỗi",
                        item.get("ma_loi") or "",
                        item.get("vi_tri") or "",
                    )
                    item["action_done"] = action_map.get(key, {}).get("has_input", False)
                    item["action_id"] = action_map.get(key, {}).get("id")
                    item["hdkp_pdf"] = action_map.get(key, {}).get("hdkp_pdf")
                    img_key = (
                        row["plan_id"],
                        date_key,
                        bucket,
                        row.get("station") or "",
                        bo_phan_key,
                        item.get("ma_loi") or "--",
                        item.get("vi_tri") or "--",
                    )
                    item["images"] = image_map.get(img_key, [])

                for item in row.get("over_target_list", []):
                    key = (
                        row["plan_id"],
                        date_key,
                        bucket,
                        row.get("station") or "",
                        bo_phan_key,
                        "CAP theo combo lỗi",
                        item.get("ma_loi") or "",
                        item.get("vi_tri") or "",
                    )
                    item["action_done"] = action_map.get(key, {}).get("has_input", False)
                    item["action_id"] = action_map.get(key, {}).get("id")
                    item["hdkp_pdf"] = action_map.get(key, {}).get("hdkp_pdf")
                    img_key = (
                        row["plan_id"],
                        date_key,
                        bucket,
                        row.get("station") or "",
                        bo_phan_key,
                        item.get("ma_loi") or "--",
                        item.get("vi_tri") or "--",
                    )
                    item["images"] = image_map.get(img_key, [])

                for item in row.get("combo_rate_top3", []):
                    key = (
                        row["plan_id"],
                        date_key,
                        bucket,
                        row.get("station") or "",
                        bo_phan_key,
                        "CAP theo combo lỗi",
                        item.get("ma_loi") or "",
                        item.get("vi_tri") or "",
                    )
                    item["action_done"] = action_map.get(key, {}).get("has_input", False)
                    item["action_id"] = action_map.get(key, {}).get("id")
                    item["hdkp_pdf"] = action_map.get(key, {}).get("hdkp_pdf")
                    img_key = (
                        row["plan_id"],
                        date_key,
                        bucket,
                        row.get("station") or "",
                        bo_phan_key,
                        item.get("ma_loi") or "--",
                        item.get("vi_tri") or "--",
                    )
                    item["images"] = image_map.get(img_key, [])

    return {"dates": result}


@app.get("/api/qc/cap/heartbeat")
def api_qc_cap_heartbeat(
    date: Optional[str] = Query(None),
):
    """Return latest update marker for QC output (SP) on a date."""
    if not date:
        date = datetime.now().strftime("%Y-%m-%d")
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT MAX(created_at) AS latest
                FROM public.qc_output_sp_log
                WHERE date = %s
                """,
                (date,),
            )
            row = cur.fetchone()
            latest = row[0] if row else None
    return {"date": date, "latest": latest.isoformat() if latest else None}

@app.get("/api/qc/cap/filters")
def api_qc_cap_filters(request: Request, don_vi: Optional[str] = Query(None)):
    """Return filter options for QC CAP (don_vi, bo_phan, station)."""
    don_vi, _user = resolve_qc_don_vi_scope(request, don_vi)
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT data_type
                FROM information_schema.columns
                WHERE table_schema = 'public'
                  AND table_name = 'prod_plan'
                  AND column_name = 'bo_phan'
                LIMIT 1
            """)
            bo_phan_type = (cur.fetchone() or {}).get("data_type")
            use_jsonb_bo_phan = bo_phan_type == "jsonb"

            # don_vi options
            don_vi_filter_params: List[Any] = []
            don_vi_filter_sql = ""
            if don_vi:
                don_vi_filter_sql = "AND don_vi = %s"
                don_vi_filter_params.append(don_vi)
            cur.execute(
                f"""
                SELECT DISTINCT don_vi
                FROM public.prod_plan
                WHERE don_vi IS NOT NULL AND don_vi <> ''
                  {don_vi_filter_sql}
                ORDER BY don_vi
                """,
                tuple(don_vi_filter_params),
            )
            don_vi_options = [r["don_vi"] for r in cur.fetchall()]

            # station options
            station_params: List[Any] = []
            station_scope_sql = ""
            if don_vi:
                station_scope_sql = "JOIN public.prod_plan p ON p.id = o.plan_id AND p.don_vi = %s"
                station_params.append(don_vi)
            cur.execute(
                f"""
                SELECT DISTINCT o.station
                FROM public.qc_output_sp_log o
                {station_scope_sql}
                WHERE o.station IS NOT NULL AND o.station <> ''
                ORDER BY o.station
                """,
                tuple(station_params),
            )
            station_options = [r["station"] for r in cur.fetchall()]

            # bo_phan options (child of don_vi if provided)
            if use_jsonb_bo_phan:
                if don_vi:
                    cur.execute("""
                        SELECT DISTINCT value AS bo_phan
                        FROM public.prod_plan p,
                             jsonb_array_elements_text(COALESCE(p.bo_phan, '[]'::jsonb)) AS t(value)
                        WHERE p.don_vi = %s AND value <> ''
                        ORDER BY value
                    """, (don_vi,))
                else:
                    cur.execute("""
                        SELECT DISTINCT value AS bo_phan
                        FROM public.prod_plan p,
                             jsonb_array_elements_text(COALESCE(p.bo_phan, '[]'::jsonb)) AS t(value)
                        WHERE value <> ''
                        ORDER BY value
                    """)
            else:
                if don_vi:
                    cur.execute("""
                        SELECT DISTINCT bo_phan
                        FROM public.prod_plan
                        WHERE don_vi = %s AND bo_phan IS NOT NULL AND bo_phan <> ''
                        ORDER BY bo_phan
                    """, (don_vi,))
                else:
                    cur.execute("""
                        SELECT DISTINCT bo_phan
                        FROM public.prod_plan
                        WHERE bo_phan IS NOT NULL AND bo_phan <> ''
                        ORDER BY bo_phan
                    """)

            bo_phan_options = [r["bo_phan"] for r in cur.fetchall()]

    return {
        "don_vi": don_vi_options,
        "bo_phan": bo_phan_options,
        "station": station_options,
    }


@app.post("/api/qc/cap/backfill")
def backfill_qc_cap(
    request: Request,
    date: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
):
    """Backfill qc_error_dps from existing CAP data."""
    _ = api_qc_cap(request=request, date=date, date_from=date_from, date_to=date_to)
    return {"status": "ok", "date_from": date_from or date, "date_to": date_to or date}


QTCN_LOAI_HANG_MAP = {
    "AOVES": "Áo vest",
    "QUANVES": "Quần tây",
}
QTCN_SYNC_PLAN_PREFIX = "SYNC-QTCN-XNV2-"
QTCN_SOURCE_SYSTEM = "prod_factory"
_qtcn_auto_sync_started = False
_qtcn_auto_sync_lock = threading.Lock()


def normalize_prod_plan_bo_phan_list(raw: Any) -> List[str]:
    if raw is None:
        return []
    if isinstance(raw, list):
        parts = [str(x) for x in raw]
    else:
        parts = re.split(r"[;,]|\s+và\s+|\s+&\s+|\s+\+\s+", str(raw), flags=re.IGNORECASE)

    seen = set()
    out: List[str] = []
    for item in parts:
        s = (item or "").strip()
        if not s:
            continue
        num_match = re.search(r"(\d+)", s)
        if s.isdigit():
            s = f"Tổ {int(s)}"
        elif s.lower().startswith(("tổ", "to")) and num_match:
            s = f"Tổ {int(num_match.group(1))}"
        s = re.sub(r"\s+", " ", s)
        if s not in seen:
            seen.add(s)
            out.append(s)
    return out


def build_prod_plan_display_name(row: Dict[str, Any]) -> str:
    ma_hang = str(row.get("ma_hang") or "").strip()
    san_luong = row.get("san_luong")
    ngay_rc = row.get("ngay_rc")

    if not ma_hang:
        return str(row.get("ke_hoach") or "--")

    parts = [ma_hang]
    if san_luong is not None and str(san_luong).strip() != "":
        parts.append(str(san_luong).strip())
    if ngay_rc:
        if isinstance(ngay_rc, (date, datetime)):
            parts.append(ngay_rc.strftime("%d-%m-%y"))
        else:
            try:
                parts.append(date.fromisoformat(str(ngay_rc)[:10]).strftime("%d-%m-%y"))
            except ValueError:
                parts.append(str(ngay_rc))
    return "_".join(parts) or str(row.get("ke_hoach") or "--")


def normalize_qtcn_to_sx_list(raw: Any) -> List[str]:
    if raw is None:
        return []
    if isinstance(raw, list):
        items = raw
    elif isinstance(raw, str):
        text = raw.strip()
        if not text:
            return []
        try:
            parsed = json.loads(text)
            if isinstance(parsed, list):
                items = parsed
            else:
                items = [text]
        except Exception:
            items = [text]
    else:
        items = [raw]

    out: List[str] = []
    seen = set()
    for item in items:
        text = (str(item or "")).strip()
        if not text:
            continue
        match = re.search(r"(\d+)", text)
        if not match:
            continue
        bo_phan = f"Tổ {int(match.group(1))}"
        if bo_phan not in seen:
            seen.add(bo_phan)
            out.append(bo_phan)
    return out


def normalize_po_info(raw: Any) -> List[Any]:
    if raw is None:
        return []
    if isinstance(raw, list):
        return raw
    if isinstance(raw, str):
        text = raw.strip()
        if not text:
            return []
        try:
            parsed = json.loads(text)
            return parsed if isinstance(parsed, list) else []
        except Exception:
            return []
    return []


def compute_po_info_total(po_info: List[Any], fallback: Any = None) -> Optional[int]:
    total = 0
    has_quantity = False
    for item in po_info:
        if isinstance(item, dict):
            qty = item.get("quantity")
            try:
                qty_int = int(qty)
            except (TypeError, ValueError):
                continue
            total += qty_int
            has_quantity = True
    if has_quantity:
        return total
    if fallback is None or fallback == "":
        return None
    try:
        return int(fallback)
    except (TypeError, ValueError):
        return None


def build_qtcn_sync_plan_code(source_id: str) -> str:
    safe_id = re.sub(r"[^A-Za-z0-9]+", "", str(source_id or "")).upper()
    return f"{QTCN_SYNC_PLAN_PREFIX}{safe_id}" if safe_id else QTCN_SYNC_PLAN_PREFIX.rstrip("-")


def ensure_prod_plan_is_active(plan_id: Any) -> None:
    try:
        plan_id_int = int(plan_id)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="plan_id không hợp lệ")

    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT id, COALESCE(is_active, TRUE) AS is_active
                FROM public.prod_plan
                WHERE id = %s
                """,
                (plan_id_int,),
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Không tìm thấy kế hoạch")
            if not row.get("is_active"):
                raise HTTPException(status_code=409, detail="Kế hoạch đã ngừng hiệu lực. Vui lòng chọn lại kế hoạch khác.")


def sync_qtcn_prod_plan() -> Dict[str, Any]:
    synced = 0
    inserted = 0
    updated = 0
    skipped = 0
    deactivated = 0
    warnings: List[str] = []

    with get_prod_factory_connection() as prod_conn:
        with prod_conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as prod_cur:
            prod_cur.execute(
                """
                SELECT id, khach_hang, ma_hang, to_sx, ngay_sx, loai_hang, so_po, sl_ke_hoach, status
                FROM public.qtcn_input
                ORDER BY ngay_sx DESC NULLS LAST, created_at DESC
                """
            )
            source_rows = prod_cur.fetchall()

    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            for row in source_rows:
                source_id = str(row.get("id") or "").strip()
                if not source_id:
                    skipped += 1
                    warnings.append("Bỏ qua 1 record qtcn_input vì thiếu id nguồn")
                    continue

                bo_phan_list = normalize_qtcn_to_sx_list(row.get("to_sx"))
                if not bo_phan_list:
                    skipped += 1
                    warnings.append(f"Bỏ qua {source_id}: không tách được tổ sản xuất")
                    continue

                loai_hang_source = (row.get("loai_hang") or "").strip().upper()
                loai_hang = QTCN_LOAI_HANG_MAP.get(loai_hang_source)
                if not loai_hang:
                    skipped += 1
                    warnings.append(f"Bỏ qua {source_id}: chưa map loại hàng {row.get('loai_hang')}")
                    continue

                source_status = ((row.get("status") or "active").strip() or "active").lower()
                is_active = source_status == "active"
                po_info = normalize_po_info(row.get("so_po"))
                san_luong = compute_po_info_total(po_info, row.get("sl_ke_hoach"))
                ke_hoach = build_qtcn_sync_plan_code(source_id)

                synced += 1
                cur.execute(
                    """
                    UPDATE public.prod_plan
                    SET ke_hoach = %s,
                        don_vi = %s,
                        bo_phan = %s::jsonb,
                        khach_hang = %s,
                        ma_hang = %s,
                        loai_hang = %s,
                        ngay_rc = %s,
                        san_luong = %s,
                        mau = COALESCE(mau, ''),
                        size = COALESCE(size, ''),
                        po_info = %s::jsonb,
                        source_status = %s,
                        is_active = %s,
                        last_synced_at = NOW(),
                        updated_at = NOW()
                    WHERE source_system = %s AND source_record_id = %s
                    """,
                    (
                        ke_hoach,
                        "XNV2",
                        json.dumps(bo_phan_list),
                        (row.get("khach_hang") or "").strip(),
                        (row.get("ma_hang") or "").strip(),
                        loai_hang,
                        row.get("ngay_sx"),
                        san_luong,
                        json.dumps(po_info),
                        source_status,
                        is_active,
                        QTCN_SOURCE_SYSTEM,
                        source_id,
                    ),
                )
                if cur.rowcount:
                    updated += 1
                    if not is_active:
                        deactivated += 1
                    continue

                # Backward compatibility for earlier synced rows keyed only by ke_hoach.
                cur.execute(
                    """
                    UPDATE public.prod_plan
                    SET source_system = %s,
                        source_record_id = %s,
                        ke_hoach = %s,
                        don_vi = %s,
                        bo_phan = %s::jsonb,
                        khach_hang = %s,
                        ma_hang = %s,
                        loai_hang = %s,
                        ngay_rc = %s,
                        san_luong = %s,
                        mau = COALESCE(mau, ''),
                        size = COALESCE(size, ''),
                        po_info = %s::jsonb,
                        source_status = %s,
                        is_active = %s,
                        last_synced_at = NOW(),
                        updated_at = NOW()
                    WHERE ke_hoach = %s
                    """,
                    (
                        QTCN_SOURCE_SYSTEM,
                        source_id,
                        ke_hoach,
                        "XNV2",
                        json.dumps(bo_phan_list),
                        (row.get("khach_hang") or "").strip(),
                        (row.get("ma_hang") or "").strip(),
                        loai_hang,
                        row.get("ngay_sx"),
                        san_luong,
                        json.dumps(po_info),
                        source_status,
                        is_active,
                        ke_hoach,
                    ),
                )
                if cur.rowcount:
                    updated += 1
                    if not is_active:
                        deactivated += 1
                    continue

                cur.execute(
                    """
                    INSERT INTO public.prod_plan
                        (ke_hoach, don_vi, bo_phan, khach_hang, ma_hang, loai_hang,
                         ngay_rc, san_luong, mau, size, po_info,
                         source_system, source_record_id, source_status, is_active, last_synced_at)
                    VALUES (%s, %s, %s::jsonb, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, %s, %s, %s, %s, NOW())
                    """,
                    (
                        ke_hoach,
                        "XNV2",
                        json.dumps(bo_phan_list),
                        (row.get("khach_hang") or "").strip(),
                        (row.get("ma_hang") or "").strip(),
                        loai_hang,
                        row.get("ngay_sx"),
                        san_luong,
                        "",
                        "",
                        json.dumps(po_info),
                        QTCN_SOURCE_SYSTEM,
                        source_id,
                        source_status,
                        is_active,
                    ),
                )
                inserted += 1
                if not is_active:
                    deactivated += 1
            conn.commit()

    return {
        "status": "ok",
        "synced": synced,
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "deactivated": deactivated,
        "warnings": warnings[:20],
    }


def run_qtcn_auto_sync_loop() -> None:
    logger.info("QTCN auto-sync started with interval %s minutes", QTCN_AUTO_SYNC_INTERVAL_MINUTES)
    while True:
        try:
            result = sync_qtcn_prod_plan()
            logger.info(
                "QTCN auto-sync completed: synced=%s inserted=%s updated=%s deactivated=%s skipped=%s",
                result.get("synced"),
                result.get("inserted"),
                result.get("updated"),
                result.get("deactivated"),
                result.get("skipped"),
            )
        except Exception:
            logger.exception("QTCN auto-sync failed")
        time.sleep(QTCN_AUTO_SYNC_INTERVAL_MINUTES * 60)


def start_qtcn_auto_sync_if_enabled() -> None:
    global _qtcn_auto_sync_started
    if not QTCN_AUTO_SYNC_ENABLED:
        return
    with _qtcn_auto_sync_lock:
        if _qtcn_auto_sync_started:
            return
        thread = threading.Thread(target=run_qtcn_auto_sync_loop, name="qtcn-auto-sync", daemon=True)
        thread.start()
        _qtcn_auto_sync_started = True


@app.get("/api/prod-plan")
def api_prod_plan_list(
    request: Request,
    don_vi: Optional[str] = Query(None),
    bo_phan: Optional[str] = Query(None),
    only_active: bool = Query(False),
):
    """List prod_plan rows, optionally filtered by don_vi and bo_phan."""

    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT data_type
                FROM information_schema.columns
                WHERE table_schema = 'public'
                  AND table_name = 'prod_plan'
                  AND column_name = 'bo_phan'
                LIMIT 1
            """)
            bo_phan_type = (cur.fetchone() or {}).get("data_type")
            use_jsonb_bo_phan = bo_phan_type == "jsonb"

            clauses = []
            params = []
            if only_active:
                clauses.append("COALESCE(is_active, TRUE) = TRUE")
            if don_vi:
                clauses.append("don_vi = %s")
                params.append(don_vi)
            if bo_phan:
                if use_jsonb_bo_phan:
                    items = normalize_prod_plan_bo_phan_list(bo_phan)
                    if items:
                        or_clauses = []
                        for item in items:
                            or_clauses.append("bo_phan @> %s::jsonb")
                            params.append(json.dumps([item]))
                        clauses.append("(" + " OR ".join(or_clauses) + ")")
                else:
                    clauses.append("bo_phan = %s")
                    params.append(bo_phan)
            where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
            cur.execute(
                f"""
                SELECT id, ke_hoach, don_vi,
                       {
                           "COALESCE(NULLIF((SELECT string_agg(value, ', ') FROM jsonb_array_elements_text(COALESCE(bo_phan, '[]'::jsonb)) AS t(value)), ''), '')"
                           if use_jsonb_bo_phan
                           else "COALESCE(bo_phan, '')"
                       } AS bo_phan,
                       khach_hang, ma_hang,
                       loai_hang, ngay_rc, san_luong, mau, size, po_info,
                       source_system, source_record_id, source_status, COALESCE(is_active, TRUE) AS is_active, last_synced_at,
                       created_at, updated_at
                FROM public.prod_plan
                {where}
                ORDER BY ngay_rc DESC NULLS LAST, id DESC
                """,
                params
            )
            rows = cur.fetchall()
            # Serialize dates
            for r in rows:
                for k in ('ngay_rc', 'created_at', 'updated_at'):
                    if r.get(k):
                        r[k] = str(r[k])
    return {"rows": rows}


@app.post("/api/prod-plan")
async def api_prod_plan_create(request: Request):
    """Create a new prod_plan entry."""
    body = await request.json()
    ke_hoach = (body.get("ke_hoach") or "").strip()
    don_vi = (body.get("don_vi") or "").strip()
    bo_phan_raw = (body.get("bo_phan") or "").strip()
    khach_hang = (body.get("khach_hang") or "").strip()
    ma_hang = (body.get("ma_hang") or "").strip()
    loai_hang = (body.get("loai_hang") or "").strip()
    ngay_rc = body.get("ngay_rc") or None
    san_luong = body.get("san_luong")
    mau = (body.get("mau") or "").strip()
    size = (body.get("size") or "").strip()
    po_info = normalize_po_info(body.get("po_info"))

    if not don_vi:
        raise HTTPException(status_code=400, detail="Đơn vị không được để trống")
    if don_vi not in DON_VI_OPTIONS:
        raise HTTPException(status_code=400, detail=f"Đơn vị không hợp lệ: {don_vi}")

    if not bo_phan_raw:
        raise HTTPException(status_code=400, detail="Bộ phận không được để trống")
    bo_phan_list = normalize_prod_plan_bo_phan_list(bo_phan_raw)
    if not bo_phan_list:
        raise HTTPException(status_code=400, detail="Bộ phận không hợp lệ")

    if san_luong is not None:
        try:
            san_luong = int(san_luong)
        except (ValueError, TypeError):
            raise HTTPException(status_code=400, detail="Sản lượng phải là số")

    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                INSERT INTO public.prod_plan
                    (ke_hoach, don_vi, bo_phan, khach_hang, ma_hang, loai_hang,
                     ngay_rc, san_luong, mau, size, po_info)
                VALUES (%s, %s, %s::jsonb, %s, %s, %s, %s, %s, %s, %s, %s::jsonb)
                RETURNING id
                """,
                (ke_hoach, don_vi, json.dumps(bo_phan_list), khach_hang, ma_hang, loai_hang,
                 ngay_rc, san_luong, mau, size, json.dumps(po_info))
            )
            new_id = cur.fetchone()["id"]
            conn.commit()
    return {"status": "ok", "id": new_id}


@app.patch("/api/prod-plan/{plan_id}")
async def api_prod_plan_update(plan_id: int, request: Request):
    """Update an existing prod_plan entry."""
    body = await request.json()

    # Build dynamic SET clause
    allowed_fields = ["ke_hoach", "don_vi", "khach_hang", "ma_hang",
                      "loai_hang", "ngay_rc", "san_luong", "mau", "size"]
    sets = ["updated_at = NOW()"]
    params = []

    for field in allowed_fields:
        if field in body:
            val = body[field]
            if field == "don_vi" and val and val not in DON_VI_OPTIONS:
                raise HTTPException(status_code=400, detail=f"Đơn vị không hợp lệ: {val}")
            if field == "san_luong" and val is not None:
                try:
                    val = int(val)
                except (ValueError, TypeError):
                    raise HTTPException(status_code=400, detail="Sản lượng phải là số")
            if field == "ngay_rc" and val == "":
                val = None
            sets.append(f"{field} = %s")
            params.append(val)

    # Handle bo_phan separately (parse list)
    if "bo_phan" in body:
        bo_phan_raw = (body["bo_phan"] or "").strip()
        if bo_phan_raw:
            out = normalize_prod_plan_bo_phan_list(bo_phan_raw)
            if not out:
                raise HTTPException(status_code=400, detail="Bộ phận không hợp lệ")
            sets.append("bo_phan = %s::jsonb")
            params.append(json.dumps(out))

    if "po_info" in body:
        sets.append("po_info = %s::jsonb")
        params.append(json.dumps(normalize_po_info(body.get("po_info"))))

    params.append(plan_id)

    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"UPDATE public.prod_plan SET {', '.join(sets)} WHERE id = %s",
                params
            )
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="Không tìm thấy kế hoạch")
            conn.commit()
    return {"status": "ok"}


@app.post("/api/prod-plan/sync-qtcn")
def api_prod_plan_sync_qtcn(request: Request):
    """Sync active XNV2 plans from prod_factory.qtcn_input into qlcl.prod_plan."""
    user = get_authenticated_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Bạn chưa đăng nhập")
    return sync_qtcn_prod_plan()


@app.delete("/api/prod-plan/{plan_id}")
def api_prod_plan_delete(plan_id: int):
    """Delete a prod_plan entry."""
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM public.prod_plan WHERE id = %s", (plan_id,))
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="Không tìm thấy kế hoạch")
            conn.commit()
    return {"status": "ok"}



# --- Hierarchical QC Error Reporting ---

@app.get("/api/dm/loai-hang")
def api_dm_loai_hang():
    """Lấy danh sách Loại hàng."""
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT lh.id, lh.ten_loai, lh.id_type, tgt.type AS type_name, tgt.target_percent
                FROM public.dm_loai_hang lh
                LEFT JOIN public.dm_loai_hang_target tgt ON tgt.id_type = lh.id_type
                ORDER BY lh.ten_loai
            """)
            return {"rows": cur.fetchall()}

@app.post("/api/dm/loai-hang")
async def api_dm_loai_hang_create(request: Request):
    body = await request.json()
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            id_type = body.get("id_type")
            if id_type is None and body.get("type"):
                cur.execute("SELECT id_type FROM public.dm_loai_hang_target WHERE type = %s", (body["type"],))
                row = cur.fetchone()
                id_type = row[0] if row else None

            if id_type is None:
                raise HTTPException(status_code=400, detail="Thiếu phân loại (id_type) hơp lệ")

            cur.execute(
                "INSERT INTO public.dm_loai_hang (ten_loai, id_type) VALUES (%s, %s) RETURNING id",
                (body["ten_loai"], id_type)
            )
            new_id = cur.fetchone()[0]
            conn.commit()
    return {"id": new_id}

@app.delete("/api/dm/loai-hang/{id}")
def api_dm_loai_hang_delete(id: int):
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM public.dm_loai_hang WHERE id = %s", (id,))
            conn.commit()
    return {"status": "ok"}

@app.get("/api/dm/loai-hang-target")
def api_dm_loai_hang_target():
    """Lấy danh sách phân loại loại hàng và target."""
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT id_type, type, target_percent
                FROM public.dm_loai_hang_target
                ORDER BY id_type
            """)
            return {"rows": cur.fetchall()}

@app.get("/api/dm/bo-phan")
def api_dm_bo_phan(loai_hang_id: int = Query(...)):
    """Lấy danh sách Bộ phận theo loai_hang_id."""
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT id, ten_bo_phan FROM public.dm_bo_phan WHERE loai_hang_id = %s ORDER BY ten_bo_phan", (loai_hang_id,))
            return {"rows": cur.fetchall()}

@app.get("/api/dm/qc-cum")
def api_dm_qc_cum(
    loai_hang_id: int = Query(...),
    active_only: bool = Query(True),
):
    """Lấy danh sách cụm QC theo loại hàng."""
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            clauses = ["loai_hang_id = %s"]
            params: List[Any] = [loai_hang_id]
            if active_only:
                clauses.append("is_active = TRUE")
            cur.execute(
                f"""
                SELECT id, loai_hang_id, ten_cum, sort_order, is_active
                FROM public.dm_qc_cum
                WHERE {" AND ".join(clauses)}
                ORDER BY sort_order, ten_cum
                """,
                tuple(params),
            )
            return {"rows": cur.fetchall()}

@app.post("/api/dm/bo-phan")
async def api_dm_bo_phan_create(request: Request):
    body = await request.json()
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("INSERT INTO public.dm_bo_phan (loai_hang_id, ten_bo_phan) VALUES (%s, %s) RETURNING id", 
                        (body["loai_hang_id"], body["ten_bo_phan"]))
            new_id = cur.fetchone()[0]
            conn.commit()
    return {"id": new_id}

@app.delete("/api/dm/bo-phan/{id}")
def api_dm_bo_phan_delete(id: int):
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM public.dm_bo_phan WHERE id = %s", (id,))
            conn.commit()
    return {"status": "ok"}

@app.get("/api/dm/chi-tiet")
def api_dm_chi_tiet(bo_phan_id: int = Query(...)):
    """Lấy danh sách Chi tiết theo bo_phan_id."""
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT id, ten_chi_tiet FROM public.dm_chi_tiet WHERE bo_phan_id = %s ORDER BY ten_chi_tiet", (bo_phan_id,))
            return {"rows": cur.fetchall()}

@app.post("/api/dm/chi-tiet")
async def api_dm_chi_tiet_create(request: Request):
    body = await request.json()
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("INSERT INTO public.dm_chi_tiet (bo_phan_id, ten_chi_tiet) VALUES (%s, %s) RETURNING id", 
                        (body["bo_phan_id"], body["ten_chi_tiet"]))
            new_id = cur.fetchone()[0]
            conn.commit()
    return {"id": new_id}

@app.delete("/api/dm/chi-tiet/{id}")
def api_dm_chi_tiet_delete(id: int):
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM public.dm_chi_tiet WHERE id = %s", (id,))
            conn.commit()
    return {"status": "ok"}


# Visual position picker: returns nhóm -> khối (image) -> hotspots (codes).
# Used by /qc-input for loại hàng that has been "số hoá" (e.g. Áo vest).
NHOM_LABELS = {"chinh": "Chính", "lot": "Lót", "nhan_dien": "Nhận diện"}

@app.get("/api/qc/visual-picker")
def api_qc_visual_picker(loai_hang_id: int = Query(...)):
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT bp.id AS bo_phan_id, bp.ten_bo_phan, bp.nhom, bp.image_png, bp.image_svg, bp.sort_order,
                       ct.id AS chi_tiet_id, ct.ma_vi_tri, ct.ten_chi_tiet,
                       ct.x_pct, ct.y_pct, ct.w_pct, ct.h_pct, ct.rotation
                FROM public.dm_bo_phan bp
                LEFT JOIN public.dm_chi_tiet ct ON ct.bo_phan_id = bp.id
                WHERE bp.loai_hang_id = %s
                ORDER BY bp.sort_order, bp.id, ct.ma_vi_tri NULLS LAST, ct.id
                """,
                (loai_hang_id,),
            )
            rows = cur.fetchall()

    has_picker = any(r["image_png"] and r["nhom"] for r in rows)
    if not has_picker:
        return {"has_visual_picker": False, "nhoms": []}

    by_nhom: Dict[str, Dict[int, dict]] = {}
    for r in rows:
        nhom = r["nhom"] or "khac"
        bp_id = r["bo_phan_id"]
        if nhom not in by_nhom:
            by_nhom[nhom] = {}
        if bp_id not in by_nhom[nhom]:
            by_nhom[nhom][bp_id] = {
                "bo_phan_id": bp_id,
                "ten_khoi": r["ten_bo_phan"],
                "image_png": f"/api/images/{r['image_png']}" if r["image_png"] else None,
                "image_svg": f"/api/images/{r['image_svg']}" if r["image_svg"] else None,
                "sort_order": r["sort_order"],
                "hotspots": [],
            }
        if r["chi_tiet_id"] is not None:
            by_nhom[nhom][bp_id]["hotspots"].append({
                "chi_tiet_id": r["chi_tiet_id"],
                "ma_vi_tri": r["ma_vi_tri"],
                "label": r["ten_chi_tiet"],
                "x_pct": float(r["x_pct"]) if r["x_pct"] is not None else None,
                "y_pct": float(r["y_pct"]) if r["y_pct"] is not None else None,
                "w_pct": float(r["w_pct"]) if r["w_pct"] is not None else None,
                "h_pct": float(r["h_pct"]) if r["h_pct"] is not None else None,
                "rotation": float(r["rotation"]) if r["rotation"] is not None else 0.0,
            })

    nhom_order = ["chinh", "lot", "nhan_dien"]
    nhoms = []
    for key in nhom_order:
        if key not in by_nhom:
            continue
        khoi_list = sorted(by_nhom[key].values(), key=lambda x: (x["sort_order"], x["bo_phan_id"]))
        nhoms.append({"nhom": key, "label": NHOM_LABELS.get(key, key), "khoi": khoi_list})
    # Any other nhom (e.g. NULL) goes last
    for key, bps in by_nhom.items():
        if key in nhom_order:
            continue
        khoi_list = sorted(bps.values(), key=lambda x: (x["sort_order"], x["bo_phan_id"]))
        nhoms.append({"nhom": key, "label": NHOM_LABELS.get(key, key.title()), "khoi": khoi_list})

    return {"has_visual_picker": True, "nhoms": nhoms}


@app.get("/api/qc/visual-picker/loai-hang-list")
def api_qc_visual_picker_loai_hang_list():
    """Liệt kê loại hàng có visual picker để hiển thị trong selector admin."""
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT DISTINCT lh.id, lh.ten_loai
                FROM public.dm_loai_hang lh
                JOIN public.dm_bo_phan bp ON bp.loai_hang_id = lh.id
                WHERE bp.image_png IS NOT NULL AND bp.nhom IS NOT NULL
                ORDER BY lh.ten_loai
                """
            )
            return {"rows": cur.fetchall()}


@app.patch("/api/qc/visual-picker/hotspots-batch")
async def api_qc_visual_picker_update_hotspots(
    request: Request,
    _user: Dict = Depends(require_qaqt_api_user),
):
    """Bulk update x_pct/y_pct cho 1 nhóm chi_tiet (1 khoi). Chỉ QAQT.
    Body: {"items": [{"chi_tiet_id": int, "x_pct": float, "y_pct": float}, ...]}
    """
    body = await request.json()
    items = body.get("items") or []
    if not items:
        return {"status": "ok", "updated": 0}
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            for it in items:
                cid = int(it["chi_tiet_id"])
                x = max(0.0, min(1.0, float(it["x_pct"])))
                y = max(0.0, min(1.0, float(it["y_pct"])))
                cur.execute(
                    "UPDATE public.dm_chi_tiet SET x_pct = %s, y_pct = %s WHERE id = %s",
                    (x, y, cid),
                )
            conn.commit()
    return {"status": "ok", "updated": len(items)}


@app.get("/qc/settings/visual-picker")
def qc_settings_visual_picker_page(request: Request):
    user = get_authenticated_user(request)
    if not user:
        return RedirectResponse(url="/qc-login", status_code=303)
    if (user.get("department") or "").upper() != "QAQT":
        return templates.TemplateResponse(
            "qc_settings_visual_picker.html",
            build_qc_template_context(request, user, forbidden=True),
            status_code=403,
        )
    return templates.TemplateResponse(
        "qc_settings_visual_picker.html",
        build_qc_template_context(request, user, forbidden=False),
    )


@app.get("/api/dm/khach-hang")
def api_dm_khach_hang():
    """Lấy danh sách Khách hàng."""
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT id, ten_khach_hang FROM public.dm_khach_hang ORDER BY ten_khach_hang")
            return {"rows": cur.fetchall()}

@app.post("/api/dm/khach-hang")
async def api_dm_khach_hang_create(request: Request):
    body = await request.json()
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("INSERT INTO public.dm_khach_hang (ten_khach_hang) VALUES (%s) RETURNING id", (body["ten_khach_hang"],))
            new_id = cur.fetchone()[0]
            conn.commit()
    return {"id": new_id}

@app.delete("/api/dm/khach-hang/{id}")
def api_dm_khach_hang_delete(id: int):
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM public.dm_khach_hang WHERE id = %s", (id,))
            conn.commit()
    return {"status": "ok"}

@app.get("/api/dm/ma-hang")
def api_dm_ma_hang(khach_hang_id: int = Query(...)):
    """Lấy danh sách Mã hàng theo khach_hang_id."""
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT id, ten_ma_hang FROM public.dm_ma_hang WHERE khach_hang_id = %s ORDER BY ten_ma_hang", (khach_hang_id,))
            return {"rows": cur.fetchall()}

@app.post("/api/dm/ma-hang")
async def api_dm_ma_hang_create(request: Request):
    body = await request.json()
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("INSERT INTO public.dm_ma_hang (khach_hang_id, ten_ma_hang) VALUES (%s, %s) RETURNING id", 
                        (body["khach_hang_id"], body["ten_ma_hang"]))
            new_id = cur.fetchone()[0]
            conn.commit()
    return {"id": new_id}

@app.delete("/api/dm/ma-hang/{id}")
def api_dm_ma_hang_delete(id: int):
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM public.dm_ma_hang WHERE id = %s", (id,))
            conn.commit()
    return {"status": "ok"}

# --- Error Classification ---

@app.get("/api/dm/nhom-loi")
def api_dm_nhom_loi():
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT id, ten_nhom FROM public.dm_nhom_loi ORDER BY ten_nhom")
            return {"rows": cur.fetchall()}

@app.post("/api/dm/nhom-loi")
async def api_dm_nhom_loi_create(request: Request):
    body = await request.json()
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("INSERT INTO public.dm_nhom_loi (ten_nhom) VALUES (%s) RETURNING id", (body["ten_nhom"],))
            new_id = cur.fetchone()[0]
            conn.commit()
    return {"id": new_id}

@app.delete("/api/dm/nhom-loi/{id}")
def api_dm_nhom_loi_delete(id: int):
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM public.dm_nhom_loi WHERE id = %s", (id,))
            conn.commit()
    return {"status": "ok"}

@app.get("/api/dm/ma-loi")
def api_dm_ma_loi(nhom_loi_id: int = Query(...)):
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT id, ten_ma FROM public.dm_ma_loi WHERE nhom_loi_id = %s ORDER BY ten_ma", (nhom_loi_id,))
            return {"rows": cur.fetchall()}

@app.post("/api/dm/ma-loi")
async def api_dm_ma_loi_create(request: Request):
    body = await request.json()
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("INSERT INTO public.dm_ma_loi (nhom_loi_id, ten_ma) VALUES (%s, %s) RETURNING id", 
                        (body["nhom_loi_id"], body["ten_ma"]))
            new_id = cur.fetchone()[0]
            conn.commit()
    return {"id": new_id}

@app.delete("/api/dm/ma-loi/{id}")
def api_dm_ma_loi_delete(id: int):
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM public.dm_ma_loi WHERE id = %s", (id,))
            conn.commit()
    return {"status": "ok"}

@app.get("/api/dm/mo-ta-loi")
def api_dm_mo_ta_loi(ma_loi_id: int = Query(...)):
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT id, ten_mo_ta FROM public.dm_mo_ta_loi WHERE ma_loi_id = %s ORDER BY ten_mo_ta", (ma_loi_id,))
            return {"rows": cur.fetchall()}

@app.post("/api/dm/mo-ta-loi")
async def api_dm_mo_ta_loi_create(request: Request):
    body = await request.json()
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("INSERT INTO public.dm_mo_ta_loi (ma_loi_id, ten_mo_ta) VALUES (%s, %s) RETURNING id", 
                        (body["ma_loi_id"], body["ten_mo_ta"]))
            new_id = cur.fetchone()[0]
            conn.commit()
    return {"id": new_id}

@app.delete("/api/dm/mo-ta-loi/{id}")
def api_dm_mo_ta_loi_delete(id: int):
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM public.dm_mo_ta_loi WHERE id = %s", (id,))
            conn.commit()
    return {"status": "ok"}

@app.get("/api/dm/ma-loi-options")
def api_dm_ma_loi_options():
    """
    Lấy danh sách Mã lỗi + Mô tả lỗi dạng gộp (ten_ma: ten_mo_ta)
    """
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT 
                    ml.id AS ma_loi_id,
                    mt.id AS mo_ta_loi_id,
                    ml.ten_ma,
                    mt.ten_mo_ta,
                    mt.muc_do
                FROM public.dm_ma_loi ml
                JOIN public.dm_mo_ta_loi mt ON mt.ma_loi_id = ml.id
                ORDER BY ml.ten_ma, mt.ten_mo_ta
            """)
            rows = cur.fetchall()

            def _sort_key(ten_ma: str):
                if not ten_ma:
                    return ("", 0, ten_ma)
                m = re.match(r"([A-Za-z]+)(\d+)", ten_ma.strip())
                if m:
                    return (m.group(1), int(m.group(2)), ten_ma)
                return (ten_ma, 0, ten_ma)

            rows.sort(key=lambda r: _sort_key(r.get("ten_ma")))
            for r in rows:
                r["label"] = f"{r['ten_ma']}: {r['ten_mo_ta']}"
            return {"rows": rows}


@app.post("/api/qc/error-log-sp")
async def api_qc_error_log_sp_create(request: Request):
    """Ghi nhận lỗi QC theo sản phẩm (real-time, không chọn mốc thời gian)."""
    body = await request.json()
    ma_nv_cookie = request.cookies.get("ma_nv")
    ma_nv_db = None
    if ma_nv_cookie:
        ma_nv = decode_ma_nv_cookie(ma_nv_cookie)
        ma_nv_variants = generate_ma_nv_variants(ma_nv)
        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(
                    """
                    SELECT ma_nv
                    FROM public.quality_employees
                    WHERE ma_nv = ANY(%s) AND chuc_vu = 'QC'
                    ORDER BY CASE WHEN ma_nv = %s THEN 0 ELSE 1 END
                    LIMIT 1
                    """,
                    (ma_nv_variants, ma_nv),
                )
                row_nv = cur.fetchone()
                if row_nv:
                    ma_nv_db = row_nv.get("ma_nv")
    plan_id = body.get("plan_id")
    ensure_prod_plan_is_active(plan_id)
    date_str = body.get("date") or datetime.now().strftime("%Y-%m-%d")
    station = (body.get("station") or "").strip()
    defect_products = body.get("defect_products", []) or []
    append_only = bool(body.get("append_only"))
    pending_failed_count = body.get("pending_failed_count", 0) or 0
    try:
        pending_failed_count = int(pending_failed_count)
    except (TypeError, ValueError):
        pending_failed_count = 0
    if pending_failed_count < 0:
        pending_failed_count = 0
    if pending_failed_count > 0 and len(defect_products) == 0:
        raise HTTPException(status_code=400, detail="pending_failed_count requires defect_products")

    output = 0

    def _defect_key(d: Dict) -> Tuple:
        return (
            d.get("bo_phan_id"),
            d.get("chi_tiet_id"),
            d.get("ma_loi_id"),
            d.get("mo_ta_loi_id"),
            (d.get("muc_do") or "").strip(),
        )

    def _defect_key_multi(d: Dict) -> Tuple:
        return (
            d.get("bo_phan_id"),
            d.get("chi_tiet_id"),
            d.get("ma_loi_id"),
        )

    def _has_defect(prod: Dict, key: Tuple) -> bool:
        for x in (prod.get("defects") or []):
            if _defect_key(x) == key:
                return True
        return False

    # Auto detect lap_lai_3 when 3 consecutive products share same defect combo
    for idx, prod in enumerate(defect_products):
        defects = prod.get("defects", []) or []
        for d in defects:
            d["lap_lai_3"] = False
            if idx >= 2:
                key = _defect_key(d)
                if _has_defect(defect_products[idx - 1], key) and _has_defect(defect_products[idx - 2], key):
                    d["lap_lai_3"] = True

    defect_count = len(defect_products)

    with get_db_connection() as conn:
        with conn.cursor() as cur:
            # Calculate output from real-time log (+1/-1)
            cur.execute(
                """
                SELECT COALESCE(SUM(delta), 0)
                FROM public.qc_output_sp_log
                WHERE plan_id = %s AND date = %s AND COALESCE(station, '') = COALESCE(%s, '')
                """,
                (plan_id, date_str, station)
            )
            output = cur.fetchone()[0] or 0

            cur.execute(
                """
                SELECT id
                FROM public.qc_error_log_sp
                WHERE plan_id = %s AND date = %s AND COALESCE(station, '') = COALESCE(%s, '')
                ORDER BY id DESC
                LIMIT 1
                """,
                (plan_id, date_str, station)
            )
            row = cur.fetchone()
            if row:
                log_id = row[0]
                if append_only:
                    cur.execute(
                        """
                        SELECT COALESCE(MAX(sp_index), 0)
                        FROM public.qc_defect
                        WHERE error_log_sp_id = %s
                        """,
                        (log_id,)
                    )
                    base_index = cur.fetchone()[0] or 0
                    total_after = base_index + defect_count
                    cur.execute(
                        """
                        UPDATE public.qc_error_log_sp
                        SET output = %s,
                            defect_count = %s,
                            ma_nv = COALESCE(ma_nv, %s),
                            updated_at = NOW()
                        WHERE id = %s
                        """,
                        (output, total_after, ma_nv_db, log_id)
                    )
                else:
                    cur.execute(
                        """
                        UPDATE public.qc_error_log_sp
                        SET output = %s,
                            defect_count = %s,
                            ma_nv = COALESCE(ma_nv, %s),
                            updated_at = NOW()
                        WHERE id = %s
                        """,
                        (output, defect_count, ma_nv_db, log_id)
                    )
                    cur.execute("DELETE FROM public.qc_defect WHERE error_log_sp_id = %s", (log_id,))
                updated = True
            else:
                cur.execute(
                    """
                    INSERT INTO public.qc_error_log_sp (plan_id, date, station, output, defect_count, ma_nv)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    RETURNING id
                    """,
                    (plan_id, date_str, station, output, defect_count, ma_nv_db)
                )
                log_id = cur.fetchone()[0]
                updated = False
                base_index = 0

            if not append_only:
                base_index = 0
            for idx, prod in enumerate(defect_products, start=1):
                defects = prod.get("defects", []) or []
                for d in defects:
                    cur.execute(
                        """
                        INSERT INTO public.qc_defect
                        (error_log_sp_id, sp_index, bo_phan_id, chi_tiet_id, ma_loi_id, mo_ta_loi_id, muc_do, lap_lai_3, image_path)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """,
                        (
                            log_id,
                            base_index + idx,
                            d.get("bo_phan_id"),
                            d.get("chi_tiet_id"),
                            d.get("ma_loi_id"),
                            d.get("mo_ta_loi_id"),
                            d.get("muc_do"),
                            d.get("lap_lai_3", False),
                            d.get("image_path"),
                        )
                    )
            if pending_failed_count > 0:
                failed_times = []
                for _ in range(pending_failed_count):
                    cur.execute(
                        """
                        INSERT INTO public.qc_output_sp_log (plan_id, date, station, delta, status, ma_nv)
                        VALUES (%s, %s, %s, 1, 'Failed', %s)
                        RETURNING created_at
                        """,
                        (plan_id, date_str, station, ma_nv_db)
                    )
                    row = cur.fetchone()
                    failed_times.append(row[0] if row else None)
                cur.execute(
                    """
                    SELECT COALESCE(COUNT(*), 0)
                    FROM public.qc_output_sp_log
                    WHERE plan_id = %s AND date = %s AND COALESCE(station, '') = COALESCE(%s, '')
                    """
                    , (plan_id, date_str, station)
                )
                output = cur.fetchone()[0] or 0
                cur.execute(
                    """
                    UPDATE public.qc_error_log_sp
                    SET output = %s, updated_at = NOW()
                    WHERE id = %s
                    """
                    , (output, log_id)
                )
            else:
                failed_times = []

            # Detect 3 consecutive failed products sharing same (bo_phan, chi_tiet, ma_loi)
            # and log to qc_defect_multi at the 3rd occurrence only.
            if defect_products and pending_failed_count > 0:
                min_sp = max(1, base_index - 2)
                max_sp = base_index + defect_count
                cur.execute(
                    """
                    SELECT sp_index, bo_phan_id, chi_tiet_id, ma_loi_id
                    FROM public.qc_defect
                    WHERE error_log_sp_id = %s
                      AND sp_index BETWEEN %s AND %s
                    """,
                    (log_id, min_sp, max_sp)
                )
                sp_map: Dict[int, set] = {}
                for sp_index, bp_id, ct_id, ml_id in cur.fetchall():
                    sp_map.setdefault(int(sp_index), set()).add((bp_id, ct_id, ml_id))

                def _get_keys(sp_idx: int) -> set:
                    return sp_map.get(sp_idx, set())

                new_sp_indices = [base_index + i for i in range(1, defect_count + 1)]
                time_by_sp = {}
                for i, sp_idx in enumerate(new_sp_indices):
                    t = failed_times[i] if i < len(failed_times) else None
                    time_by_sp[sp_idx] = t

                records = []
                for sp_idx in new_sp_indices:
                    keys_cur = _get_keys(sp_idx)
                    if not keys_cur:
                        continue
                    keys_prev1 = _get_keys(sp_idx - 1)
                    keys_prev2 = _get_keys(sp_idx - 2)
                    keys_prev3 = _get_keys(sp_idx - 3)
                    for key in keys_cur:
                        if key in keys_prev1 and key in keys_prev2 and key not in keys_prev3:
                            records.append((sp_idx, key))

                if records:
                    bp_ids = {k[0] for _, k in records if k[0] is not None}
                    ct_ids = {k[1] for _, k in records if k[1] is not None}
                    ml_ids = {k[2] for _, k in records if k[2] is not None}

                    bp_map = {}
                    ct_map = {}
                    ml_map = {}
                    if bp_ids:
                        cur.execute(
                            "SELECT id, ten_bo_phan FROM public.dm_bo_phan WHERE id = ANY(%s)",
                            (list(bp_ids),)
                        )
                        bp_map = {row[0]: row[1] for row in cur.fetchall()}
                    if ct_ids:
                        cur.execute(
                            "SELECT id, ten_chi_tiet FROM public.dm_chi_tiet WHERE id = ANY(%s)",
                            (list(ct_ids),)
                        )
                        ct_map = {row[0]: row[1] for row in cur.fetchall()}
                    if ml_ids:
                        cur.execute(
                            "SELECT id, ten_ma FROM public.dm_ma_loi WHERE id = ANY(%s)",
                            (list(ml_ids),)
                        )
                        ml_map = {row[0]: row[1] for row in cur.fetchall()}

                    for sp_idx, key in records:
                        bp_id, ct_id, ml_id = key
                        t = time_by_sp.get(sp_idx)
                        time_val = t.time() if hasattr(t, "time") else datetime.now().time()
                        cur.execute(
                            """
                            INSERT INTO public.qc_defect_multi
                            (plan_id, date, station, time, bo_phan, chi_tiet, ma_loi, ma_nv)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                            """,
                            (
                                plan_id,
                                date_str,
                                station,
                                time_val,
                                bp_map.get(bp_id) if bp_id is not None else None,
                                ct_map.get(ct_id) if ct_id is not None else None,
                                ml_map.get(ml_id) if ml_id is not None else None,
                                ma_nv_db,
                            )
                        )
            conn.commit()
    return {"status": "ok", "id": log_id, "updated": updated}


@app.post("/api/qc/upload-sp-image")
async def api_qc_upload_sp_image(file: UploadFile = File(...)):
    """Upload anh chup san pham loi QC (SP mode)."""
    import uuid
    sub_dir = os.path.join(IMAGES_STORAGE_DIR, "qc_sp")
    os.makedirs(sub_dir, exist_ok=True)
    ext = os.path.splitext(file.filename or "photo.jpg")[1] or ".jpg"
    unique_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}{ext}"
    file_path = os.path.join(sub_dir, unique_name)
    contents = await file.read()
    with open(file_path, "wb") as f:
        f.write(contents)
    relative_path = f"qc_sp/{unique_name}"
    return {"status": "ok", "image_path": relative_path, "url": f"/api/images/{relative_path}"}



@app.post("/api/qc/output-sp-log")
async def api_qc_output_sp_log_create(request: Request):
    """Ghi nhận thao tác +1/-1 cho số sản phẩm đã kiểm (theo sản phẩm)."""
    body = await request.json()
    ma_nv_cookie = request.cookies.get("ma_nv")
    if not ma_nv_cookie:
        raise HTTPException(status_code=401, detail="Chưa đăng nhập")
    ma_nv = decode_ma_nv_cookie(ma_nv_cookie)
    ma_nv_variants = generate_ma_nv_variants(ma_nv)
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT ma_nv
                FROM public.quality_employees
                WHERE ma_nv = ANY(%s) AND chuc_vu = 'QC'
                ORDER BY CASE WHEN ma_nv = %s THEN 0 ELSE 1 END
                LIMIT 1
                """,
                (ma_nv_variants, ma_nv),
            )
            row_nv = cur.fetchone()
            if not row_nv:
                raise HTTPException(status_code=403, detail="Không có quyền (QC only)")
            ma_nv_db = row_nv["ma_nv"]
    plan_id = body.get("plan_id")
    ensure_prod_plan_is_active(plan_id)
    date_str = body.get("date") or datetime.now().strftime("%Y-%m-%d")
    station = (body.get("station") or "").strip()
    delta = body.get("delta", 1)
    status = (body.get("status") or "Passed").strip()
    try:
        delta = int(delta)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="delta phải là số (+1 hoặc -1)")
    if delta not in (1, -1):
        raise HTTPException(status_code=400, detail="delta chỉ nhận +1 hoặc -1")
    if status not in ("Passed", "Failed"):
        raise HTTPException(status_code=400, detail="status must be Passed or Failed")

    with get_db_connection() as conn:
        with conn.cursor() as cur:
            if delta == 1:
                cur.execute(
                    """
                    INSERT INTO public.qc_output_sp_log (plan_id, date, station, delta, status, ma_nv)
                    VALUES (%s, %s, %s, 1, %s, %s)
                    """,
                    (plan_id, date_str, station, status, ma_nv_db)
                )
            else:
                # Remove latest +1 log (same plan/date/station) for the current QC user
                cur.execute(
                    """
                    DELETE FROM public.qc_output_sp_log
                    WHERE id = (
                        SELECT id FROM public.qc_output_sp_log
                        WHERE plan_id = %s
                          AND date = %s
                          AND COALESCE(station, '') = COALESCE(%s, '')
                          AND COALESCE(ma_nv, '') = COALESCE(%s, '')
                        ORDER BY created_at DESC, id DESC
                        LIMIT 1
                    )
                    """,
                    (plan_id, date_str, station, ma_nv_db)
                )
            cur.execute(
                """
                SELECT 
                    COALESCE(COUNT(*), 0) AS total,
                    COALESCE(SUM(CASE WHEN status = 'Passed' THEN 1 ELSE 0 END), 0) AS passed_count,
                    COALESCE(SUM(CASE WHEN status = 'Failed' THEN 1 ELSE 0 END), 0) AS failed_count
                FROM public.qc_output_sp_log
                WHERE plan_id = %s AND date = %s AND COALESCE(station, '') = COALESCE(%s, '')
                """,
                (plan_id, date_str, station)
            )
            row = cur.fetchone()
        conn.commit()

    return {
        "status": "ok",
        "total": row[0],
        "passed_count": row[1],
        "failed_count": row[2],
    }


@app.get("/api/qc/output-sp")
def api_qc_output_sp_get(
    plan_id: int = Query(...),
    date_str: str = Query(..., alias="date"),
    station: Optional[str] = Query(None),
):
    """Lấy tổng số sản phẩm đã kiểm (tính theo log +1/-1)."""
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT 
                    COALESCE(COUNT(*), 0) AS total,
                    COALESCE(SUM(CASE WHEN status = 'Passed' THEN 1 ELSE 0 END), 0) AS passed_count,
                    COALESCE(SUM(CASE WHEN status = 'Failed' THEN 1 ELSE 0 END), 0) AS failed_count
                FROM public.qc_output_sp_log
                WHERE plan_id = %s AND date = %s AND COALESCE(station, '') = COALESCE(%s, '')
                """,
                (plan_id, date_str, station)
            )
            row = cur.fetchone()
            return {"total": row[0], "passed_count": row[1], "failed_count": row[2]}


@app.get("/api/qc/input/pos-summary")
def api_qc_input_pos_summary(
    plan_id: int = Query(...),
    date_str: str = Query(..., alias="date"),
    station: Optional[str] = Query(None),
):
    """Thống kê lỗi theo vị trí trong từng mốc thời gian cho trang /qc-input.

    Các loại lỗi:
    - Lỗi đại trà: theo combo (Vị trí [Bộ phận - Chi tiết] + Mã lỗi) / tổng kiểm trong mốc >= 15%
      (mẫu số: qc_output_sp_log; đếm theo số sản phẩm DISTINCT (error_log_sp_id, sp_index))
    - Lỗi nghiêm trọng: qc_defect.muc_do ILIKE '%nghiêm%'
    - Lỗi nguy cơ hàng loạt: qc_defect_multi
    """

    def _bucket_case(time_expr: str) -> str:
        return f"""
            CASE
                WHEN {time_expr} < '09:30' THEN '7H30 - 9H30'
                WHEN {time_expr} >= '09:30' AND {time_expr} < '11:30' THEN '9H30 - 11H30'
                WHEN {time_expr} >= '11:30' AND {time_expr} < '14:30' THEN '12H30 - 14H30'
                WHEN {time_expr} >= '14:30' AND {time_expr} < '16:30' THEN '14H30 - 16H30'
                ELSE 'Sau 16H30'
            END
        """

    def _vi_tri(bo_phan: str, chi_tiet: str) -> str:
        parts = [p for p in [bo_phan, chi_tiet] if p]
        return " - ".join(parts) if parts else "--"

    bucket_order = ["7H30 - 9H30", "9H30 - 11H30", "12H30 - 14H30", "14H30 - 16H30", "Sau 16H30"]

    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            output_params: List = [plan_id, date_str]
            output_where = ["o.plan_id = %s", "o.date = %s"]
            if station:
                output_where.append("COALESCE(o.station, '') = COALESCE(%s, '')")
                output_params.append(station)
            cur.execute(
                f"""
                SELECT
                    {_bucket_case("timezone('Asia/Bangkok', o.created_at)::time")} AS time_bucket,
                    SUM(o.delta) AS output_total,
                    SUM(CASE WHEN o.status = 'Failed' THEN o.delta ELSE 0 END) AS defect_total
                FROM public.qc_output_sp_log o
                WHERE {" AND ".join(output_where)}
                GROUP BY time_bucket
                """,
                tuple(output_params),
            )
            output_rows = cur.fetchall()
            output_by_bucket = {r["time_bucket"]: int(r.get("output_total") or 0) for r in output_rows}
            failed_by_bucket = {r["time_bucket"]: int(r.get("defect_total") or 0) for r in output_rows}

            sp_params: List = [plan_id, date_str]
            sp_where = ["sp.plan_id = %s", "sp.date = %s"]
            if station:
                sp_where.append("COALESCE(sp.station, '') = COALESCE(%s, '')")
                sp_params.append(station)
            sp_where_sql = " AND ".join(sp_where)

            # Serious
            cur.execute(
                f"""
                SELECT
                    {_bucket_case("timezone('Asia/Bangkok', d.created_at)::time")} AS time_bucket,
                    COALESCE(ml.ten_ma, d.ma_loi_id::text) AS ma_loi,
                    COALESCE(bp.ten_bo_phan, '') AS bo_phan,
                    COALESCE(ct.ten_chi_tiet, '') AS chi_tiet,
                    COUNT(*) AS qty
                FROM public.qc_defect d
                JOIN public.qc_error_log_sp sp ON sp.id = d.error_log_sp_id
                LEFT JOIN public.dm_ma_loi ml ON ml.id = d.ma_loi_id
                LEFT JOIN public.dm_bo_phan bp ON bp.id = d.bo_phan_id
                LEFT JOIN public.dm_chi_tiet ct ON ct.id = d.chi_tiet_id
                WHERE {sp_where_sql}
                  AND d.muc_do ILIKE '%%nghiêm%%'
                GROUP BY time_bucket, ma_loi, bo_phan, chi_tiet
                """,
                tuple(sp_params),
            )
            serious_rows = cur.fetchall()

            # Multi risk
            multi_params: List = [plan_id, date_str]
            multi_where = ["qdm.plan_id = %s", "qdm.date = %s"]
            if station:
                multi_where.append("COALESCE(qdm.station, '') = COALESCE(%s, '')")
                multi_params.append(station)
            cur.execute(
                f"""
                SELECT
                    {_bucket_case("qdm.time")} AS time_bucket,
                    COALESCE(qdm.ma_loi, '') AS ma_loi,
                    COALESCE(qdm.bo_phan, '') AS bo_phan,
                    COALESCE(qdm.chi_tiet, '') AS chi_tiet,
                    COUNT(*) AS occurrences
                FROM public.qc_defect_multi qdm
                WHERE {" AND ".join(multi_where)}
                GROUP BY time_bucket, ma_loi, bo_phan, chi_tiet
                """,
                tuple(multi_params),
            )
            multi_rows = cur.fetchall()

            # Mass combos
            cur.execute(
                f"""
                SELECT
                    {_bucket_case("timezone('Asia/Bangkok', d.created_at)::time")} AS time_bucket,
                    COALESCE(ml.ten_ma, d.ma_loi_id::text) AS ma_loi,
                    COALESCE(bp.ten_bo_phan, '') AS bo_phan,
                    COALESCE(ct.ten_chi_tiet, '') AS chi_tiet,
                    COUNT(DISTINCT (d.error_log_sp_id, d.sp_index)) AS qty
                FROM public.qc_defect d
                JOIN public.qc_error_log_sp sp ON sp.id = d.error_log_sp_id
                LEFT JOIN public.dm_ma_loi ml ON ml.id = d.ma_loi_id
                LEFT JOIN public.dm_bo_phan bp ON bp.id = d.bo_phan_id
                LEFT JOIN public.dm_chi_tiet ct ON ct.id = d.chi_tiet_id
                WHERE {sp_where_sql}
                  AND d.error_log_sp_id IS NOT NULL
                  AND d.sp_index IS NOT NULL
                GROUP BY time_bucket, ma_loi, bo_phan, chi_tiet
                """,
                tuple(sp_params),
            )
            mass_rows = cur.fetchall()

    bucket_map: Dict[str, List[Dict]] = {b: [] for b in bucket_order}
    loai_order = {"Lỗi đại trà": 0, "Lỗi nghiêm trọng": 1, "Lỗi nguy cơ hàng loạt": 2}

    for r in serious_rows:
        bucket = r.get("time_bucket")
        if bucket not in bucket_map:
            continue
        ma_loi = r.get("ma_loi") or "--"
        vi_tri = _vi_tri(r.get("bo_phan") or "", r.get("chi_tiet") or "")
        bucket_map[bucket].append({
            "loai_loi": "Lỗi nghiêm trọng",
            "ma_loi": ma_loi,
            "bo_phan": r.get("bo_phan") or "",
            "chi_tiet": r.get("chi_tiet") or "",
            "vi_tri": f"{ma_loi} - {vi_tri}",
            "qty": int(r.get("qty") or 0),
        })

    for r in multi_rows:
        bucket = r.get("time_bucket")
        if bucket not in bucket_map:
            continue
        ma_loi = (r.get("ma_loi") or "").strip() or "--"
        vi_tri = _vi_tri(r.get("bo_phan") or "", r.get("chi_tiet") or "")
        bucket_map[bucket].append({
            "loai_loi": "Lỗi nguy cơ hàng loạt",
            "ma_loi": ma_loi,
            "bo_phan": r.get("bo_phan") or "",
            "chi_tiet": r.get("chi_tiet") or "",
            "vi_tri": f"{ma_loi} - {vi_tri}",
            "qty": int(r.get("occurrences") or 0),
        })

    for r in mass_rows:
        bucket = r.get("time_bucket")
        if bucket not in bucket_map:
            continue
        qty = int(r.get("qty") or 0)
        if qty <= 0:
            continue
        output_total = int(output_by_bucket.get(bucket) or 0)
        if output_total <= 0:
            continue
        if (qty / output_total) < 0.15:
            continue
        ma_loi = r.get("ma_loi") or "--"
        vi_tri = _vi_tri(r.get("bo_phan") or "", r.get("chi_tiet") or "")
        bucket_map[bucket].append({
            "loai_loi": "Lỗi đại trà",
            "ma_loi": ma_loi,
            "bo_phan": r.get("bo_phan") or "",
            "chi_tiet": r.get("chi_tiet") or "",
            "vi_tri": f"{ma_loi} - {vi_tri}",
            "qty": qty,
        })

    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT
                    dps.time_bucket,
                    dps.loai_loi,
                    COALESCE(dps.ma_loi, '') AS ma_loi,
                    COALESCE(dps.vi_tri, '') AS vi_tri,
                    CASE WHEN hm.qc_error_dps_id IS NOT NULL THEN TRUE ELSE FALSE END AS has_hdkp
                FROM public.qc_error_dps dps
                LEFT JOIN public.qc_hdkp_mota hm ON hm.qc_error_dps_id = dps.id
                WHERE dps.plan_id = %s
                  AND dps.date = %s
                  AND COALESCE(dps.station, '') = COALESCE(%s, '')
                """,
                (plan_id, date_str, station),
            )
            hdkp_map = {
                (
                    r.get("time_bucket") or "",
                    r.get("loai_loi") or "",
                    r.get("ma_loi") or "",
                    r.get("vi_tri") or "",
                ): bool(r.get("has_hdkp"))
                for r in cur.fetchall()
            }

    buckets = []
    for b in bucket_order:
        items = bucket_map.get(b) or []
        items = [x for x in items if int(x.get("qty") or 0) > 0]
        for item in items:
            raw_vi_tri = " - ".join([p for p in [item.get("bo_phan") or "", item.get("chi_tiet") or ""] if p])
            item["has_hdkp"] = hdkp_map.get(
                (
                    b,
                    item.get("loai_loi") or "",
                    item.get("ma_loi") or "",
                    raw_vi_tri,
                ),
                False,
            )
        items.sort(key=lambda x: (loai_order.get(x.get("loai_loi") or "", 99), -int(x.get("qty") or 0), str(x.get("vi_tri") or "")))
        output_total = int(output_by_bucket.get(b) or 0)
        defect_total = int(failed_by_bucket.get(b) or 0)
        defect_rate = round((defect_total / output_total) * 100, 2) if output_total > 0 else 0.0
        buckets.append({
            "time_bucket": b,
            "output_total": output_total,
            "defect_total": defect_total,
            "defect_rate": defect_rate,
            "items": items,
        })

    return {"status": "ok", "plan_id": plan_id, "date": date_str, "station": station or "", "buckets": buckets}


@app.get("/api/qc/input/quick-defect-combos")
def api_qc_input_quick_defect_combos(
    plan_id: int = Query(...),
    station: str = Query(...),
):
    """Top defect combos by historical count for the selected plan type and QC cluster."""
    station_clean = (station or "").strip()
    if not station_clean:
        return {"status": "ok", "rows": [], "message": "Chưa chọn cụm."}

    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                WITH selected AS (
                    SELECT p.id AS plan_id,
                           p.loai_hang,
                           lh.id AS loai_hang_id,
                           qc.ten_cum
                    FROM public.prod_plan p
                    JOIN public.dm_loai_hang lh ON lh.ten_loai = p.loai_hang
                    JOIN public.dm_qc_cum qc
                      ON qc.loai_hang_id = lh.id
                     AND qc.ten_cum = %s
                     AND qc.is_active = TRUE
                    WHERE p.id = %s
                    LIMIT 1
                )
                SELECT
                    d.bo_phan_id,
                    bp.ten_bo_phan,
                    d.chi_tiet_id,
                    ct.ten_chi_tiet,
                    d.ma_loi_id,
                    ml.ten_ma,
                    d.mo_ta_loi_id,
                    mt.ten_mo_ta,
                    MAX(mt.muc_do::text) AS muc_do_text,
                    COUNT(*)::int AS defect_qty
                FROM selected s
                JOIN public.prod_plan p
                  ON p.loai_hang = s.loai_hang
                JOIN public.qc_error_log_sp sp
                  ON sp.plan_id = p.id
                 AND COALESCE(sp.station, '') = s.ten_cum
                JOIN public.qc_defect d
                  ON d.error_log_sp_id = sp.id
                JOIN public.dm_bo_phan bp
                  ON bp.id = d.bo_phan_id
                 AND bp.loai_hang_id = s.loai_hang_id
                JOIN public.dm_chi_tiet ct
                  ON ct.id = d.chi_tiet_id
                 AND ct.bo_phan_id = bp.id
                JOIN public.dm_ma_loi ml
                  ON ml.id = d.ma_loi_id
                JOIN public.dm_mo_ta_loi mt
                  ON mt.id = d.mo_ta_loi_id
                 AND mt.ma_loi_id = ml.id
                WHERE d.bo_phan_id IS NOT NULL
                  AND d.chi_tiet_id IS NOT NULL
                  AND d.ma_loi_id IS NOT NULL
                  AND d.mo_ta_loi_id IS NOT NULL
                GROUP BY
                    d.bo_phan_id, bp.ten_bo_phan,
                    d.chi_tiet_id, ct.ten_chi_tiet,
                    d.ma_loi_id, ml.ten_ma,
                    d.mo_ta_loi_id, mt.ten_mo_ta
                ORDER BY defect_qty DESC, bp.ten_bo_phan, ct.ten_chi_tiet, ml.ten_ma, mt.ten_mo_ta
                LIMIT 15
                """,
                (station_clean, plan_id),
            )
            rows = cur.fetchall()

    for r in rows:
        muc_do_text = r.pop("muc_do_text", None)
        try:
            r["muc_do"] = json.loads(muc_do_text) if muc_do_text else None
        except (TypeError, ValueError):
            r["muc_do"] = muc_do_text
        r["label"] = (
            f"{r.get('ten_bo_phan') or ''} - "
            f"{r.get('ten_chi_tiet') or ''} - "
            f"{r.get('ten_ma') or ''}. {r.get('ten_mo_ta') or ''}"
        )
    return {"status": "ok", "rows": rows}


@app.get("/api/qc/error-log-sp")
def api_qc_error_log_sp_get(
    plan_id: int = Query(...),
    date_str: str = Query(..., alias="date"),
    station: Optional[str] = Query(None),
):
    """Lấy dữ liệu QC theo sản phẩm để sửa."""
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT id, plan_id, date, station, output, defect_count
                FROM public.qc_error_log_sp
                WHERE plan_id = %s AND date = %s AND COALESCE(station, '') = COALESCE(%s, '')
                ORDER BY id DESC
                LIMIT 1
                """,
                (plan_id, date_str, station)
            )
            log = cur.fetchone()
            if not log:
                return {"exists": False}

            cur.execute(
                """
                SELECT sp_index, bo_phan_id, chi_tiet_id, ma_loi_id, mo_ta_loi_id, muc_do, lap_lai_3
                FROM public.qc_defect
                WHERE error_log_sp_id = %s
                ORDER BY sp_index ASC, id ASC
                """,
                (log["id"],)
            )
            rows = cur.fetchall()

            grouped: Dict[int, List[Dict]] = {}
            for r in rows:
                idx = r.get("sp_index") or 0
                grouped.setdefault(idx, []).append({
                    "bo_phan_id": r.get("bo_phan_id"),
                    "chi_tiet_id": r.get("chi_tiet_id"),
                    "ma_loi_id": r.get("ma_loi_id"),
                    "mo_ta_loi_id": r.get("mo_ta_loi_id"),
                    "muc_do": r.get("muc_do"),
                    "lap_lai_3": r.get("lap_lai_3"),
                })

            defect_products = [
                {"product_no": idx, "defects": grouped[idx]}
                for idx in sorted(grouped.keys())
                if idx
            ]

            return {
                "exists": True,
                "data": {
                    "output": log.get("output") or 0,
                    "defect_count": log.get("defect_count") or 0,
                    "date": str(log.get("date")) if log.get("date") else date_str,
                    "defect_products": defect_products,
                }
            }

@app.get("/qc-input/rework")
def qc_rework_page(request: Request):
    ma_nv_cookie = request.cookies.get("ma_nv")
    if not ma_nv_cookie:
        return RedirectResponse(url="/qc-login", status_code=303)
    user_data = None
    ma_nv = decode_ma_nv_cookie(ma_nv_cookie)
    ma_nv_variants = generate_ma_nv_variants(ma_nv)
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT ma_nv, ho_ten as name, chuc_vu, don_vi, bo_phan
                FROM public.quality_employees
                WHERE ma_nv = ANY(%s)
                ORDER BY CASE WHEN ma_nv = %s THEN 0 ELSE 1 END
                LIMIT 1
                """,
                (ma_nv_variants, ma_nv)
            )
            user_data = cur.fetchone()
    if not user_data:
        return RedirectResponse(url="/qc-login", status_code=303)
        
    return templates.TemplateResponse("qc_rework_2.html", {"request": request, "user": user_data})


@app.get("/qc-input-2/rework")
def qc_rework_sp_page(request: Request):
    return RedirectResponse(url="/qc-input/rework", status_code=303)


@app.get("/api/qc/rework-2")
def api_qc_rework_sp(
    request: Request,
    date_str: str = Query(None),
    plan_id: Optional[int] = Query(None),
):
    if not date_str:
        date_str = datetime.now().strftime("%Y-%m-%d")
    ma_nv_cookie = request.cookies.get("ma_nv")
    ma_nv = decode_ma_nv_cookie(ma_nv_cookie)
    ma_nv_variants = generate_ma_nv_variants(ma_nv)
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            params: List[Any] = [date_str, ma_nv_variants]
            plan_where = ""
            if plan_id:
                plan_where = " AND e.plan_id = %s"
                params.append(plan_id)
            cur.execute("""
                SELECT 
                    e.id as error_log_sp_id,
                    e.plan_id,
                    e.created_at,
                    p.ke_hoach,
                    p.ma_hang,
                    p.loai_hang,
                    d.sp_index,
                    COUNT(*) as defect_items,
                    BOOL_AND(COALESCE(d.rework_done, FALSE)) as rework_done,
                    ARRAY_AGG(DISTINCT CONCAT_WS(' - ', bp.ten_bo_phan, ct.ten_chi_tiet, ml.ten_ma)) as defect_list
                FROM public.qc_defect d
                JOIN public.qc_error_log_sp e ON e.id = d.error_log_sp_id
                JOIN public.prod_plan p ON e.plan_id = p.id
                LEFT JOIN public.dm_bo_phan bp ON d.bo_phan_id = bp.id
                LEFT JOIN public.dm_chi_tiet ct ON d.chi_tiet_id = ct.id
                LEFT JOIN public.dm_ma_loi ml ON d.ma_loi_id = ml.id
                WHERE e.date = %s
                  AND e.ma_nv = ANY(%s)
                  {plan_where}
                GROUP BY e.id, e.plan_id, e.created_at, p.ke_hoach, p.ma_hang, p.loai_hang, d.sp_index
                ORDER BY e.created_at ASC, d.sp_index ASC
            """.format(plan_where=plan_where), tuple(params))
            rows = cur.fetchall()
            for r in rows:
                if r.get('created_at'):
                    r['created_at'] = str(r['created_at'])
            return {"rows": rows}


@app.patch("/api/qc/rework-2/{log_id}/{sp_index}")
async def api_update_rework_sp(log_id: int, sp_index: int, request: Request):
    body = await request.json()
    rework_done = bool(body.get("rework_done"))
    ma_nv_cookie = request.cookies.get("ma_nv")
    ma_nv = decode_ma_nv_cookie(ma_nv_cookie)
    ma_nv_variants = generate_ma_nv_variants(ma_nv)
    with get_db_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE public.qc_defect
                SET rework_done = %s
                WHERE error_log_sp_id = %s
                  AND sp_index = %s
                  AND error_log_sp_id IN (
                      SELECT id FROM public.qc_error_log_sp WHERE ma_nv = ANY(%s)
                  )
                """,
                (rework_done, log_id, sp_index, ma_nv_variants)
            )
        conn.commit()
    return {"status": "ok", "rework_done": rework_done}

@app.get("/healthz")
def healthz():
    return {"status": "ok"}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8008, reload=True)
